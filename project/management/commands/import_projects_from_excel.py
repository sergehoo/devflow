from __future__ import annotations
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from openpyxl import load_workbook

from project import models as dm


class Command(BaseCommand):
    help = (
        "Importe un fichier Excel de portefeuille applicatif et crée/met à jour les "
        "projets, roadmaps, jalons, sprints et tâches."
    )

    HEADER_ALIASES = {
        "numero": "numero",
        "n": "numero",
        "nom de l'application": "project_name",
        "nom application": "project_name",
        "acronyme": "code",
        "description / fonction": "description",
        "description": "description",
        "filiale / client": "client",
        "client": "client",
        "processus métier couvert": "business_process",
        "processus metier couvert": "business_process",
        "utilisateurs": "users",
        "type d'application (web/mobile/api)": "app_type",
        "type dapplication (web/mobile/api)": "app_type",
        "technologie / framework": "tech_stack",
        "base de données": "database",
        "base de donnees": "database",
        "api / intégrations": "integrations",
        "api / integrations": "integrations",
        "hébergement (cloud / serveur kaydan)": "hosting",
        "hebergement (cloud / serveur kaydan)": "hosting",
        "url": "url",
        "environnement (prod / preprod / dev)": "environment",
        "criticité (faible/moyen/critique)": "criticality",
        "criticite (faible/moyen/critique)": "criticality",
        "responsable métier/point focal client": "business_owner",
        "responsable metier/point focal client": "business_owner",
        "developpeurs": "developers",
        "statut (production / projet / obsolète)": "project_status",
        "statut (production / projet / obsolete)": "project_status",
        "version actuelle": "version",
        "date de mise en production": "go_live_date",
        "sensibilité des données": "data_sensitivity",
        "sensibilite des donnees": "data_sensitivity",
        "backup /sauvegarde": "backup",
        "backup / sauvegarde": "backup",
        "commentaires": "comments",
    }

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            required=True,
            help="Chemin absolu ou relatif du fichier Excel à importer.",
        )
        parser.add_argument(
            "--workspace",
            required=True,
            help="ID ou nom exact du workspace cible.",
        )
        parser.add_argument(
            "--owner-email",
            dest="owner_email",
            help="Email du user propriétaire à utiliser si aucun owner n'est trouvé dans le fichier.",
        )
        parser.add_argument(
            "--sheet",
            default=None,
            help="Nom de feuille à importer. Par défaut: première feuille.",
        )
        parser.add_argument(
            "--roadmap-name",
            default="Roadmap importée",
            help="Nom de la roadmap de regroupement si aucune roadmap spécifique n'est trouvée.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Analyse le fichier sans écrire en base.",
        )
        parser.add_argument(
            "--update-existing",
            action="store_true",
            help="Met à jour les projets existants trouvés par workspace + nom.",
        )
        parser.add_argument(
            "--create-sprint",
            action="store_true",
            help="Crée automatiquement un sprint initial pour chaque projet importé.",
        )
        parser.add_argument(
            "--create-tasks",
            action="store_true",
            help="Crée automatiquement des tâches bootstrap par projet.",
        )
        parser.add_argument(
            "--create-milestone",
            action="store_true",
            help="Crée automatiquement un jalon principal par projet.",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"]).expanduser()
        if not file_path.exists():
            raise CommandError(f"Fichier introuvable: {file_path}")

        workspace = self.get_workspace(options["workspace"])
        fallback_owner = self.get_user_by_email(options.get("owner_email")) if options.get("owner_email") else None

        workbook = load_workbook(filename=file_path, data_only=True)
        sheet = workbook[options["sheet"]] if options.get("sheet") else workbook[workbook.sheetnames[0]]

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise CommandError("Le fichier Excel est vide.")

        headers = [self.normalize_header(v) for v in rows[0]]
        mapped_headers = [self.HEADER_ALIASES.get(h, h) for h in headers]

        created = {
            "projects": 0,
            "roadmap_items": 0,
            "milestones": 0,
            "sprints": 0,
            "tasks": 0,
        }
        updated = {"projects": 0}
        skipped = []

        roadmap = self.get_or_create_default_roadmap(
            workspace=workspace,
            roadmap_name=options["roadmap_name"],
        )

        context = transaction.atomic if not options["dry_run"] else self.dry_run_atomic

        with context():
            for line_no, raw_row in enumerate(rows[1:], start=2):
                payload = self.row_to_dict(mapped_headers, raw_row)

                if not self.has_meaningful_content(payload):
                    continue

                project_name = (payload.get("project_name") or "").strip()
                if not project_name:
                    skipped.append(f"Ligne {line_no}: nom de projet vide")
                    continue

                team = self.get_or_create_team(workspace, payload.get("client"))
                owner = self.resolve_owner(payload, fallback_owner)
                product_manager = owner

                project_defaults = self.build_project_defaults(
                    workspace=workspace,
                    team=team,
                    owner=owner,
                    product_manager=product_manager,
                    payload=payload,
                )

                project_qs = dm.Project.objects.filter(workspace=workspace, name=project_name)
                project = project_qs.first()

                if project and options["update_existing"]:
                    for field, value in project_defaults.items():
                        setattr(project, field, value)
                    project.save()
                    updated["projects"] += 1
                elif project:
                    skipped.append(f"Ligne {line_no}: projet déjà existant '{project_name}'")
                    continue
                else:
                    project = dm.Project.objects.create(name=project_name, **project_defaults)
                    created["projects"] += 1

                self.sync_project_label(project, payload.get("environment"))
                self.sync_project_label(project, payload.get("project_status"))
                self.sync_project_label(project, payload.get("criticality"))

                if payload.get("client"):
                    self.create_or_update_roadmap_item(
                        roadmap=roadmap,
                        project=project,
                        payload=payload,
                    )
                    created["roadmap_items"] += 1

                if options["create_milestone"]:
                    if self.create_default_milestone(project, payload, owner):
                        created["milestones"] += 1

                if options["create_sprint"]:
                    sprint = self.create_default_sprint(project, payload)
                    if sprint:
                        created["sprints"] += 1
                    else:
                        sprint = None
                else:
                    sprint = None

                if options["create_tasks"]:
                    task_count = self.create_default_tasks(project, payload, sprint=sprint, owner=owner)
                    created["tasks"] += task_count

            if options["dry_run"]:
                raise RuntimeError("DRY_RUN_ROLLBACK")

        self.stdout.write(self.style.SUCCESS("Import terminé."))
        self.stdout.write(f"Projets créés: {created['projects']}")
        self.stdout.write(f"Projets mis à jour: {updated['projects']}")
        self.stdout.write(f"Roadmap items créés: {created['roadmap_items']}")
        self.stdout.write(f"Jalons créés: {created['milestones']}")
        self.stdout.write(f"Sprints créés: {created['sprints']}")
        self.stdout.write(f"Tâches créées: {created['tasks']}")

        if skipped:
            self.stdout.write(self.style.WARNING("Lignes ignorées:"))
            for item in skipped:
                self.stdout.write(f" - {item}")

    class dry_run_atomic:
        def __enter__(self):
            self.ctx = transaction.atomic()
            self.ctx.__enter__()
            return self

        def __exit__(self, exc_type, exc, tb):
            if exc and str(exc) == "DRY_RUN_ROLLBACK":
                transaction.set_rollback(True)
                self.ctx.__exit__(None, None, None)
                return True
            return self.ctx.__exit__(exc_type, exc, tb)

    def normalize_header(self, value: Any) -> str:
        value = "" if value is None else str(value)
        value = value.strip().lower()
        value = (
            value.replace("é", "e")
            .replace("è", "e")
            .replace("ê", "e")
            .replace("à", "a")
            .replace("ù", "u")
            .replace("ô", "o")
            .replace("ï", "i")
            .replace("/ ", "/")
        )
        value = re.sub(r"\s+", " ", value)
        return value

    def row_to_dict(self, headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
        payload = {}
        for idx, key in enumerate(headers):
            payload[key] = row[idx] if idx < len(row) else None
        return payload

    def has_meaningful_content(self, payload: dict[str, Any]) -> bool:
        values = [v for v in payload.values() if v not in (None, "")]
        return bool(values)

    def get_workspace(self, value: str) -> dm.Workspace:
        qs = dm.Workspace.objects.all()
        workspace = None
        if str(value).isdigit():
            workspace = qs.filter(pk=int(value)).first()
        if not workspace:
            workspace = qs.filter(name=value).first()
        if not workspace:
            raise CommandError(f"Workspace introuvable: {value}")
        return workspace

    def get_user_by_email(self, email: str):
        User = get_user_model()
        user = User.objects.filter(email__iexact=email.strip()).first()
        if not user:
            raise CommandError(f"Utilisateur introuvable pour l'email {email}")
        return user

    def resolve_owner(self, payload: dict[str, Any], fallback_owner=None):
        raw = payload.get("business_owner") or payload.get("developers") or ""
        emails = re.findall(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", str(raw), flags=re.I)
        User = get_user_model()
        for email in emails:
            user = User.objects.filter(email__iexact=email).first()
            if user:
                return user

        if raw:
            cleaned = re.sub(r"[<(].*?[>)]", "", str(raw)).strip()
            parts = [p for p in re.split(r"\s+", cleaned) if p]
            if len(parts) >= 2:
                first_name = parts[0]
                last_name = " ".join(parts[1:])
                user = User.objects.filter(first_name__iexact=first_name, last_name__iexact=last_name).first()
                if user:
                    return user

        return fallback_owner

    def get_or_create_team(self, workspace: dm.Workspace, raw_name: Any):
        name = (str(raw_name).strip() if raw_name else "")
        if not name:
            return None
        team, _ = dm.Team.objects.get_or_create(
            workspace=workspace,
            name=name,
            defaults={"team_type": dm.Team.TeamType.OTHER},
        )
        return team

    def map_project_status(self, value: Any) -> str:
        raw = (str(value or "").strip().lower())
        mapping = {
            "production": dm.Project.Status.DONE,
            "prod": dm.Project.Status.DONE,
            "projet": dm.Project.Status.IN_PROGRESS,
            "en projet": dm.Project.Status.IN_PROGRESS,
            "obsolète": dm.Project.Status.CANCELLED,
            "obsolete": dm.Project.Status.CANCELLED,
            "dev": dm.Project.Status.PLANNED,
            "preprod": dm.Project.Status.IN_DELIVERY,
        }
        return mapping.get(raw, dm.Project.Status.PLANNED)

    def map_priority(self, value: Any) -> str:
        raw = (str(value or "").strip().lower())
        if raw == "critique":
            return dm.Project.Priority.CRITICAL
        if raw == "moyen":
            return dm.Project.Priority.MEDIUM
        if raw == "faible":
            return dm.Project.Priority.LOW
        if raw == "eleve":
            return dm.Project.Priority.HIGH
        if raw == "élevé":
            return dm.Project.Priority.HIGH
        return dm.Project.Priority.MEDIUM

    def map_health(self, value: Any) -> str:
        raw = (str(value or "").strip().lower())
        if raw == "critique":
            return dm.Project.HealthStatus.RED
        if raw == "moyen":
            return dm.Project.HealthStatus.AMBER
        if raw == "faible":
            return dm.Project.HealthStatus.GREEN
        return dm.Project.HealthStatus.GRAY

    def parse_date(self, value: Any):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value).strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    def build_project_defaults(self, workspace, team, owner, product_manager, payload):
        go_live_date = self.parse_date(payload.get("go_live_date"))
        environment = str(payload.get("environment") or "").strip()
        comments = str(payload.get("comments") or "").strip()
        description_parts = [
            str(payload.get("description") or "").strip(),
            f"Processus: {payload.get('business_process')}" if payload.get("business_process") else "",
            f"Utilisateurs: {payload.get('users')}" if payload.get("users") else "",
            f"Type: {payload.get('app_type')}" if payload.get("app_type") else "",
            f"BD: {payload.get('database')}" if payload.get("database") else "",
            f"Hébergement: {payload.get('hosting')}" if payload.get("hosting") else "",
            f"URL: {payload.get('url')}" if payload.get("url") else "",
            f"Commentaires: {comments}" if comments else "",
            f"Version: {payload.get('version')}" if payload.get("version") else "",
            f"Sensibilité: {payload.get('data_sensitivity')}" if payload.get("data_sensitivity") else "",
            f"Backup: {payload.get('backup')}" if payload.get("backup") else "",
            f"Environnement: {environment}" if environment else "",
        ]
        description = "\n".join([p for p in description_parts if p])

        start_date = timezone.localdate()
        target_date = go_live_date or (start_date + timedelta(days=90))

        return {
            "workspace": workspace,
            "team": team,
            "description": description,
            "tech_stack": str(payload.get("tech_stack") or "").strip(),
            "owner": owner,
            "product_manager": product_manager,
            "status": self.map_project_status(payload.get("project_status") or environment),
            "priority": self.map_priority(payload.get("criticality")),
            "health_status": self.map_health(payload.get("criticality")),
            "progress_percent": 100 if self.map_project_status(payload.get("project_status")) == dm.Project.Status.DONE else 0,
            "ai_risk_label": str(payload.get("criticality") or "").strip(),
            "start_date": start_date,
            "target_date": target_date,
            "delivered_at": go_live_date,
        }

    def get_or_create_default_roadmap(self, workspace: dm.Workspace, roadmap_name: str):
        today = timezone.localdate()
        roadmap, _ = dm.Roadmap.objects.get_or_create(
            workspace=workspace,
            name=roadmap_name,
            defaults={
                "start_date": today,
                "end_date": today + timedelta(days=365),
                "is_public": False,
                "description": "Roadmap générée automatiquement par import Excel.",
            },
        )
        return roadmap

    def create_or_update_roadmap_item(self, roadmap, project, payload):
        start_date = project.start_date or timezone.localdate()
        end_date = project.target_date or (start_date + timedelta(days=90))
        status_map = {
            dm.Project.Status.DONE: dm.RoadmapItem.ItemStatus.DONE,
            dm.Project.Status.IN_PROGRESS: dm.RoadmapItem.ItemStatus.IN_PROGRESS,
            dm.Project.Status.IN_DELIVERY: dm.RoadmapItem.ItemStatus.IN_PROGRESS,
            dm.Project.Status.BLOCKED: dm.RoadmapItem.ItemStatus.AT_RISK,
            dm.Project.Status.DELAYED: dm.RoadmapItem.ItemStatus.AT_RISK,
        }
        status = status_map.get(project.status, dm.RoadmapItem.ItemStatus.PLANNED)
        row = project.pk or 0
        dm.RoadmapItem.objects.update_or_create(
            roadmap=roadmap,
            project=project,
            title=project.name,
            defaults={
                "milestone": None,
                "color": self.environment_color(payload.get("environment")),
                "status": status,
                "start_date": max(start_date, roadmap.start_date),
                "end_date": min(end_date, roadmap.end_date),
                "row": row,
            },
        )

    def environment_color(self, raw_value: Any) -> str:
        value = str(raw_value or "").strip().lower()
        return {
            "prod": "#16A34A",
            "production": "#16A34A",
            "preprod": "#0EA5C9",
            "dev": "#F59E0B",
        }.get(value, "#F4722B")

    def create_default_milestone(self, project, payload, owner):
        due_date = project.target_date or (timezone.localdate() + timedelta(days=90))
        name = f"Mise en production {project.name}"
        milestone, created = dm.Milestone.objects.get_or_create(
            project=project,
            name=name,
            defaults={
                "workspace": project.workspace,
                "description": f"Jalon généré depuis l'import Excel. Version: {payload.get('version') or 'N/A'}",
                "status": dm.Milestone.Status.DONE if project.status == dm.Project.Status.DONE else dm.Milestone.Status.PLANNED,
                "due_date": due_date,
                "completed_at": project.delivered_at,
                "progress_percent": 100 if project.status == dm.Project.Status.DONE else project.progress_percent,
                "owner": owner,
            },
        )
        return created

    def create_default_sprint(self, project, payload):
        start_date = project.start_date or timezone.localdate()
        end_date = min(project.target_date or (start_date + timedelta(days=14)), start_date + timedelta(days=14))
        sprint, created = dm.Sprint.objects.get_or_create(
            project=project,
            number=1,
            defaults={
                "workspace": project.workspace,
                "team": project.team,
                "name": f"Sprint 1 · {project.name}",
                "goal": str(payload.get("business_process") or payload.get("description") or "Initialisation du projet"),
                "status": dm.Sprint.Status.DONE if project.status == dm.Project.Status.DONE else dm.Sprint.Status.PLANNED,
                "start_date": start_date,
                "end_date": end_date,
            },
        )
        return sprint if created else None

    def create_default_tasks(self, project, payload, sprint=None, owner=None):
        task_specs = []
        if payload.get("business_process"):
            task_specs.append((
                "Analyser le processus métier",
                str(payload.get("business_process")),
                dm.Task.Priority.HIGH,
            ))
        if payload.get("integrations"):
            task_specs.append((
                "Configurer les intégrations",
                str(payload.get("integrations")),
                dm.Task.Priority.MEDIUM,
            ))
        if payload.get("backup"):
            task_specs.append((
                "Mettre en place la sauvegarde",
                f"Sauvegarde: {payload.get('backup')}",
                dm.Task.Priority.MEDIUM,
            ))
        if payload.get("comments"):
            task_specs.append((
                "Traiter les remarques importées",
                str(payload.get("comments")),
                dm.Task.Priority.LOW,
            ))
        if not task_specs:
            task_specs.append((
                "Initialiser le projet",
                str(payload.get("description") or project.description or "Initialisation"),
                dm.Task.Priority.MEDIUM,
            ))

        created_count = 0
        for index, (title, description, priority) in enumerate(task_specs, start=1):
            task_title = f"{title} · {project.name}"
            task, created = dm.Task.objects.get_or_create(
                project=project,
                title=task_title,
                defaults={
                    "workspace": project.workspace,
                    "sprint": sprint,
                    "description": description[:5000],
                    "status": dm.Task.Status.DONE if project.status == dm.Project.Status.DONE else dm.Task.Status.TODO,
                    "priority": priority,
                    "risk_score": 20 if priority == dm.Task.Priority.HIGH else 10,
                    "progress_percent": 100 if project.status == dm.Project.Status.DONE else 0,
                    "estimate_hours": Decimal("8.00"),
                    "reporter": owner,
                    "assignee": owner,
                    "position": index,
                },
            )
            created_count += 1 if created else 0
        return created_count

    def sync_project_label(self, project, raw_label: Any):
        value = (str(raw_label).strip() if raw_label else "")
        if not value:
            return
        label, _ = dm.Label.objects.get_or_create(
            workspace=project.workspace,
            name=value[:60],
            defaults={"color": "#F4722B"},
        )
        dm.ProjectLabel.objects.get_or_create(project=project, label=label)
