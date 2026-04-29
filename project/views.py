from __future__ import annotations

import json
from collections import OrderedDict
from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django import forms
from django.contrib.auth import get_user_model, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.views import PasswordChangeView
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models.functions import Coalesce

from django.http import HttpResponse, JsonResponse, Http404
from django.views.decorators.http import require_POST
from django.views.generic import TemplateView
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Count, Sum, ExpressionWrapper, F, DecimalField, Prefetch, Min, Max, Avg, Value, \
    FloatField
from django.shortcuts import redirect, get_object_or_404, render
from django.urls import path, reverse_lazy, reverse
from django.utils import timezone
from django.utils.text import slugify
from django.views import View
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Side, Border, Alignment

from . import models as dm
from .forms import (
    WorkspaceForm, TeamForm, TeamMembershipForm, ProjectForm, ProjectMemberForm,
    SprintForm, SprintMetricForm, BacklogItemForm, TaskForm, TaskAssignmentForm,
    TaskCommentForm, TaskAttachmentForm, PullRequestForm, RiskForm, AInsightForm,
    NotificationForm, ActivityLogForm, DirectChannelForm, ChannelMembershipForm,
    MessageForm, TimesheetEntryForm, DashboardSnapshotForm, UserPreferenceForm,
    LabelForm, TaskLabelForm, ProjectLabelForm, TaskDependencyForm,
    TaskChecklistForm, ChecklistItemForm, MilestoneForm, MilestoneTaskForm,
    ReleaseForm, RoadmapForm, RoadmapItemForm, BoardColumnForm,
    WorkspaceInvitationForm, IntegrationForm, WebhookForm, ReactionForm,
    MessageAttachmentForm, SprintReviewForm, SprintRetrospectiveForm,
    APIKeyForm, WorkspaceSettingsForm, ObjectiveForm, KeyResultForm, DevFlowPasswordChangeForm, UserAccountForm,
    UserProfileForm, TaskCommentQuickForm, ProjectDocumentImportForm,
    InvoiceForm, InvoiceLineForm, InvoicePaymentForm, InvoiceClientForm,
    InvoiceGenerateForm,
)
from .services.budget import ProjectBudgetService
from .utils.workspaces import ensure_workspace

APP_LABEL = "project"



class WorkspaceAutoAssignMixin:
    def assign_workspace(self, obj):
        ensure_workspace(obj, user=self.request.user)
        return obj

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.assign_workspace(self.object)
        self.object.save()
        form.save_m2m()
        return super().form_valid(form)


class HomeView(TemplateView, LoginRequiredMixin):
    template_name = "layout/base.html"
    login_url = reverse_lazy("login")


# =============================================================================
# MIXINS
# =============================================================================
class WorkspaceSecurityMixin:
    current_workspace = None

    def get_profile_workspace(self):
        profile = getattr(self.request.user, "profile", None)
        if not profile:
            return None
        if not profile.workspace_id:
            return None
        return profile.workspace

    def get_user_workspaces(self):
        """
        Retourne les workspaces autorisés pour l'utilisateur connecté.

        Priorité :
        1. workspace du profil utilisateur
        2. workspaces dont il est owner
        3. workspaces liés à ses memberships (si ton modèle TeamMembership est utilisé)
        """
        workspace_ids = set()

        profile_workspace = self.get_profile_workspace()
        if profile_workspace and not profile_workspace.is_archived:
            workspace_ids.add(profile_workspace.id)

        owned_ids = (
            dm.Workspace.objects.filter(
                owner=self.request.user,
                is_archived=False,
            ).values_list("id", flat=True)
        )
        workspace_ids.update(owned_ids)

        membership_ids = (
            dm.Workspace.objects.filter(
                memberships__user=self.request.user,
                is_archived=False,
            ).values_list("id", flat=True)
        )
        workspace_ids.update(membership_ids)

        return dm.Workspace.objects.filter(id__in=workspace_ids).order_by("name")

    def get_current_workspace(self):
        if self.current_workspace is not None:
            return self.current_workspace

        workspace_id = self.kwargs.get("workspace_id") or self.request.GET.get("workspace")
        user_workspaces = self.get_user_workspaces()

        if workspace_id:
            workspace = user_workspaces.filter(pk=workspace_id).first()
            if not workspace:
                raise Http404("Workspace introuvable ou accès interdit.")
            self.current_workspace = workspace
            return self.current_workspace

        profile_workspace = self.get_profile_workspace()
        if profile_workspace and user_workspaces.filter(pk=profile_workspace.pk).exists():
            self.current_workspace = profile_workspace
            return self.current_workspace

        self.current_workspace = user_workspaces.first()
        return self.current_workspace


class DevflowBaseMixin(LoginRequiredMixin):
    section = "dashboard"
    page_title = "DevFlow"
    success_list_url_name: str | None = None
    search_fields: tuple[str, ...] = ()

    def get_workspace_id(self):
        return self.kwargs.get("workspace_id") or self.request.GET.get("workspace")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["section"] = self.section
        ctx["page_title"] = self.page_title
        ctx["workspace_id"] = self.get_workspace_id()

        if hasattr(self, "get_current_workspace"):
            current_workspace = self.get_current_workspace()
            ctx["current_workspace"] = current_workspace
            ctx["workspace"] = current_workspace

        if hasattr(self, "get_user_workspaces"):
            ctx["workspaces"] = self.get_user_workspaces()

        return ctx

    def filter_by_workspace(self, queryset):
        workspace_id = self.get_workspace_id()

        if not workspace_id:
            if hasattr(self, "get_current_workspace"):
                current_workspace = self.get_current_workspace()
                if current_workspace:
                    workspace_id = current_workspace.id
                else:
                    return queryset.none()
            else:
                return queryset

        model = queryset.model
        direct_field_names = {f.name for f in model._meta.fields}

        if "workspace" in direct_field_names:
            return queryset.filter(workspace_id=workspace_id)
        if "project" in direct_field_names:
            return queryset.filter(project__workspace_id=workspace_id)
        if "team" in direct_field_names:
            return queryset.filter(team__workspace_id=workspace_id)
        if "task" in direct_field_names:
            return queryset.filter(task__workspace_id=workspace_id)
        if "sprint" in direct_field_names:
            return queryset.filter(sprint__workspace_id=workspace_id)
        if "channel" in direct_field_names:
            return queryset.filter(channel__workspace_id=workspace_id)
        if "roadmap" in direct_field_names:
            return queryset.filter(roadmap__workspace_id=workspace_id)
        if "milestone" in direct_field_names:
            return queryset.filter(milestone__workspace_id=workspace_id)
        if "objective" in direct_field_names:
            return queryset.filter(objective__workspace_id=workspace_id)
        if "message" in direct_field_names:
            return queryset.filter(message__channel__workspace_id=workspace_id)
        if "invoice" in direct_field_names:
            return queryset.filter(invoice__workspace_id=workspace_id)

        return queryset

    def build_search_query(self, term: str):
        query = Q()
        for field in self.search_fields:
            query |= Q(**{f"{field}__icontains": term})
        return query


class DevflowListView(WorkspaceSecurityMixin, DevflowBaseMixin, ListView):
    """
    ListView DevFlow avec :
      - filtrage workspace (via WorkspaceSecurityMixin)
      - recherche textuelle (search_fields → ?q=)
      - filtres déclaratifs (filter_fields → ?<param>=)

    Déclaration des filtres par sous-classe :

        filter_fields = [
            {"name": "status", "label": "Statut", "type": "choices",
             "choices": dm.Task.Status.choices},
            {"name": "project", "label": "Projet", "type": "model",
             "queryset": "_projects_qs", "lookup": "project_id"},
            {"name": "is_active", "label": "Active", "type": "bool"},
        ]

    `type` peut valoir "choices" (liste TextChoices), "model" (FK,
    queryset string ou callable), "bool" ou "exact" (texte brut).
    """

    paginate_by = 25
    context_object_name = "items"
    filter_fields: list = []
    search_placeholder: str = ""

    # ---- Filtres déclaratifs ----------------------------------------------
    def _resolve_filter_queryset(self, definition):
        """Permet d'utiliser une chaîne ('_my_qs_method') ou un callable."""
        qs = definition.get("queryset")
        if callable(qs):
            return qs(self)
        if isinstance(qs, str) and hasattr(self, qs):
            attr = getattr(self, qs)
            return attr() if callable(attr) else attr
        return qs

    def _filter_definitions_resolved(self):
        resolved = []
        for f in self.filter_fields or []:
            data = dict(f)
            if data.get("type") == "model":
                data["queryset"] = self._resolve_filter_queryset(data)
            resolved.append(data)
        return resolved

    def apply_filters(self, queryset):
        for f in self.filter_fields or []:
            name = f["name"]
            raw = (self.request.GET.get(name) or "").strip()
            if not raw:
                continue
            lookup = f.get("lookup") or name
            ftype = f.get("type", "exact")
            try:
                if ftype == "bool":
                    if raw in ("1", "true", "True"):
                        queryset = queryset.filter(**{lookup: True})
                    elif raw in ("0", "false", "False"):
                        queryset = queryset.filter(**{lookup: False})
                elif ftype == "model":
                    queryset = queryset.filter(**{lookup: raw})
                elif ftype == "choices":
                    queryset = queryset.filter(**{lookup: raw})
                else:
                    queryset = queryset.filter(**{f"{lookup}__icontains": raw})
            except Exception:
                # Filtre cassé → on ignore plutôt que crasher la liste
                continue
        return queryset

    def get_active_filters(self):
        return {
            f["name"]: (self.request.GET.get(f["name"]) or "").strip()
            for f in self.filter_fields or []
        }

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = self.filter_by_workspace(queryset)

        term = self.request.GET.get("q", "").strip()
        if term and self.search_fields:
            queryset = queryset.filter(self.build_search_query(term))

        queryset = self.apply_filters(queryset)
        return queryset.distinct()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["search_q"] = self.request.GET.get("q", "").strip()
        ctx["search_placeholder"] = self.search_placeholder or "Rechercher…"
        ctx["filter_definitions"] = self._filter_definitions_resolved()
        ctx["active_filters"] = self.get_active_filters()
        ctx["has_filters_or_search"] = bool(
            ctx["search_q"]
            or any(v for v in ctx["active_filters"].values())
        )
        return ctx


class DevflowDetailView(WorkspaceSecurityMixin, DevflowBaseMixin, DetailView):
    context_object_name = "item"

    def get_queryset(self):
        return self.filter_by_workspace(super().get_queryset())


class DevflowCreateView(WorkspaceSecurityMixin, DevflowBaseMixin, CreateView):

    template_name = "project/create.html"

    def get_success_url(self):

        return reverse_lazy(self.success_list_url_name)

    def get_form_kwargs(self):

        kwargs = super().get_form_kwargs()

        if hasattr(self, "get_current_workspace"):

            kwargs["current_workspace"] = self.get_current_workspace()

        if hasattr(self, "get_user_workspaces"):

            kwargs["allowed_workspaces"] = self.get_user_workspaces()

        kwargs["request"] = self.request

        return kwargs

    def get_form(self, form_class=None):

        form = super().get_form(form_class)

        current_workspace = self.get_current_workspace() if hasattr(self, "get_current_workspace") else None

        allowed_workspaces = self.get_user_workspaces() if hasattr(self, "get_user_workspaces") else None

        if "workspace" in form.fields:

            if allowed_workspaces is not None:

                form.fields["workspace"].queryset = allowed_workspaces

            if current_workspace:

                form.fields["workspace"].initial = current_workspace.pk

                form.initial.setdefault("workspace", current_workspace.pk)

                # si le workspace est déjà déterminé par le contexte,

                # on évite de bloquer la validation

                form.fields["workspace"].required = False

                # optionnel : masquer si un seul workspace

                if allowed_workspaces is not None and allowed_workspaces.count() == 1:

                    form.fields["workspace"].widget = forms.HiddenInput()

        return form

    def form_valid(self, form):

        self.object = form.save(commit=False)

        if hasattr(self.object, "slug") and not getattr(self.object, "slug", None):

            source = getattr(self.object, "name", None) or getattr(self.object, "title", None)

            if source:

                self.object.slug = slugify(source)

        if hasattr(self.object, "workspace_id"):

            workspace = None

            if "workspace" in form.cleaned_data:

                workspace = form.cleaned_data.get("workspace")

            if not workspace and hasattr(self, "get_current_workspace"):

                workspace = self.get_current_workspace()

            if not workspace:

                form.add_error("workspace", "Ce champ est obligatoire.")

                form.add_error(None, "Aucun workspace actif n'a pu être déterminé.")

                return self.form_invalid(form)

            self.object.workspace = workspace

        self.object.save()

        form.save_m2m()

        messages.success(self.request, f"{self.model._meta.verbose_name.title()} créé avec succès.")

        return redirect(self.get_success_url())


class DevflowUpdateView(WorkspaceSecurityMixin, DevflowBaseMixin, UpdateView):
    template_name = "project/update.html"

    def get_queryset(self):
        return self.filter_by_workspace(super().get_queryset())

    def get_success_url(self):
        return reverse_lazy(self.success_list_url_name)

    def form_valid(self, form):
        messages.success(self.request, f"{self.model._meta.verbose_name.title()} mis à jour avec succès.")
        return super().form_valid(form)


class DevflowDeleteView(WorkspaceSecurityMixin, DevflowBaseMixin, DeleteView):
    template_name = "project/crud/confirm_delete.html"
    context_object_name = "item"

    def get_queryset(self):
        return self.filter_by_workspace(super().get_queryset())

    def get_success_url(self):
        return reverse_lazy(self.success_list_url_name)

    def delete(self, request, *args, **kwargs):
        messages.success(request, f"{self.model._meta.verbose_name.title()} supprimé avec succès.")
        return super().delete(request, *args, **kwargs)


class ArchiveObjectView(WorkspaceSecurityMixin, DevflowBaseMixin, View):
    model = None
    success_list_url_name = None

    def post(self, request, pk):
        obj = self.filter_by_workspace(self.model.objects.all()).get(pk=pk)
        if hasattr(obj, "archive"):
            obj.archive()
            messages.success(request, f"{self.model._meta.verbose_name.title()} archivé avec succès.")
        return redirect(reverse_lazy(self.success_list_url_name))


class ProfileDetailView(DevflowBaseMixin, WorkspaceSecurityMixin, LoginRequiredMixin, DetailView):
    model = dm.UserProfile
    template_name = "account/profile_detail.html"
    context_object_name = "profile"
    section = "profile"
    page_title = "Mon profil"

    def get_object(self, queryset=None):
        return self.request.user.profile

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        profile = self.object

        current_memberships = (
            self.request.user.devflow_memberships
            .select_related("workspace", "team")
            .order_by("workspace__name", "team__name")
        )

        owned_projects = (
            self.request.user.owned_projects
            .select_related("workspace", "team")
            .filter(is_archived=False)
            .order_by("name")[:8]
        )

        managed_projects = (
            self.request.user.managed_projects
            .select_related("workspace", "team")
            .filter(is_archived=False)
            .order_by("name")[:8]
        )

        project_memberships = (
            self.request.user.project_memberships
            .select_related("project", "team")
            .order_by("project__name")[:8]
        )

        # Allocation cumulée sur projets actifs
        from django.db.models import Sum, Count
        total_allocation = (
            self.request.user.project_memberships
            .filter(project__is_archived=False)
            .aggregate(s=Sum("allocation_percent"))["s"] or 0
        )
        capacity_left = max(0, 100 - total_allocation)

        # Tâches assignées
        active_tasks_count = (
            self.request.user.assigned_tasks
            .filter(is_archived=False)
            .exclude(status=dm.Task.Status.DONE)
            .count()
        )
        overdue_tasks_count = (
            self.request.user.assigned_tasks
            .filter(is_archived=False, due_date__lt=timezone.localdate())
            .exclude(status=dm.Task.Status.DONE)
            .count()
        )

        recent_tasks = (
            self.request.user.assigned_tasks
            .filter(is_archived=False)
            .select_related("project")
            .order_by("-updated_at")[:6]
        )

        # Factures émises par l'utilisateur (si module billing actif)
        recent_invoices = (
            dm.Invoice.objects
            .filter(issued_by=self.request.user, is_archived=False)
            .select_related("project", "client")
            .order_by("-issue_date")[:5]
        )

        # Activité récente
        try:
            recent_activity = (
                dm.ActivityLog.objects
                .filter(actor=self.request.user)
                .select_related("project", "task")
                .order_by("-created_at")[:8]
            )
        except Exception:
            recent_activity = []

        ctx.update({
            "current_memberships": current_memberships,
            "owned_projects": owned_projects,
            "managed_projects": managed_projects,
            "project_memberships": project_memberships,
            "workspace": profile.workspace,
            "total_allocation": total_allocation,
            "capacity_left": capacity_left,
            "active_tasks_count": active_tasks_count,
            "overdue_tasks_count": overdue_tasks_count,
            "recent_tasks": recent_tasks,
            "recent_invoices": recent_invoices,
            "recent_activity": recent_activity,
        })
        return ctx


class ProfileUpdateView(DevflowBaseMixin, WorkspaceSecurityMixin, LoginRequiredMixin, UpdateView):
    model = dm.UserProfile
    form_class = UserProfileForm
    template_name = "account/profile_update.html"
    context_object_name = "profile"
    section = "profile"
    page_title = "Modifier mon profil"

    def get_object(self, queryset=None):
        return self.request.user.profile

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Optionnel : limiter le choix au workspace du profil actuel
        form.fields["workspace"].queryset = dm.Workspace.objects.filter(is_archived=False).order_by("name")
        return form

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if "account_form" not in ctx:
            ctx["account_form"] = UserAccountForm(instance=self.request.user)
        return ctx

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        account_form = UserAccountForm(request.POST, instance=request.user)

        if form.is_valid() and account_form.is_valid():
            return self.forms_valid(form, account_form)
        return self.forms_invalid(form, account_form)

    def forms_valid(self, form, account_form):
        account_form.save()
        form.save()
        messages.success(self.request, "Profil mis à jour avec succès.")
        return redirect("profile_detail")

    def forms_invalid(self, form, account_form):
        return self.render_to_response(
            self.get_context_data(form=form, account_form=account_form)
        )


class ProfilePasswordChangeView(DevflowBaseMixin, WorkspaceSecurityMixin, LoginRequiredMixin, PasswordChangeView):
    form_class = DevFlowPasswordChangeForm
    template_name = "account/profile_password_change.html"
    success_url = reverse_lazy("profile_detail")
    section = "profile"
    page_title = "Changer mon mot de passe"

    def form_valid(self, form):
        response = super().form_valid(form)
        update_session_auth_hash(self.request, form.user)
        messages.success(self.request, "Mot de passe modifié avec succès.")
        return response


# =============================================================================
# DASHBOARD
# =============================================================================

class DashboardView(WorkspaceSecurityMixin, DevflowBaseMixin, TemplateView):
    template_name = "dashboard/index.html"
    section = "dashboard"
    page_title = "Tableau de bord"

    @staticmethod
    def pct(value, total):
        if not total:
            return 0
        try:
            return round((value / total) * 100)
        except Exception:
            return 0

    def build_analysis_cards(
            self,
            stats,
            finance,
            active_sprint,
            delayed_projects,
            blocked_tasks,
            critical_risks,
            overloaded_members_count,
    ):
        cards = []

        if critical_risks > 0:
            cards.append({
                "severity": "critical",
                "title": f"{critical_risks} risque(s) critique(s)",
                "text": "Des risques majeurs nécessitent une escalade ou un plan de mitigation immédiat.",
                "action": "Prioriser la revue des risques",
            })

        if delayed_projects > 0:
            cards.append({
                "severity": "high",
                "title": f"{delayed_projects} projet(s) en retard",
                "text": "Le portefeuille comporte des échéances dépassées ou des livraisons glissantes.",
                "action": "Réviser les deadlines et arbitrer les priorités",
            })

        if blocked_tasks > 0:
            cards.append({
                "severity": "medium",
                "title": f"{blocked_tasks} tâche(s) bloquée(s)",
                "text": "Des blocages ralentissent l’avancement opérationnel du workspace.",
                "action": "Lever les dépendances critiques",
            })

        if overloaded_members_count > 0:
            cards.append({
                "severity": "medium",
                "title": f"{overloaded_members_count} membre(s) en surcharge",
                "text": "Certaines ressources dépassent un seuil de charge élevé.",
                "action": "Rééquilibrer l’allocation ou délester les tâches",
            })

        if finance["budget_usage_percent"] >= 100:
            cards.append({
                "severity": "critical",
                "title": "Budget consommé à plus de 100%",
                "text": "Les dépenses ou coûts estimés dépassent le budget approuvé consolidé.",
                "action": "Réviser les arbitrages financiers",
            })
        elif finance["budget_usage_percent"] >= 80:
            cards.append({
                "severity": "high",
                "title": "Seuil budgétaire d’alerte dépassé",
                "text": "Le niveau de consommation du budget devient sensible.",
                "action": "Surveiller les coûts et geler les dépenses non essentielles",
            })

        if stats["done_ratio_percent"] >= 70 and stats["blocked_tasks_count"] == 0:
            cards.append({
                "severity": "positive",
                "title": "Bonne dynamique de livraison",
                "text": "Le taux de tâches terminées est élevé et les blocages restent maîtrisés.",
                "action": "Maintenir le rythme de livraison",
            })

        if active_sprint and active_sprint.remaining_days < 0:
            cards.append({
                "severity": "high",
                "title": f"{active_sprint.name} est dépassé",
                "text": "Le sprint actif a franchi sa date de fin prévue.",
                "action": "Clôturer ou prolonger le sprint selon l’état réel",
            })

        return cards[:6]

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = timezone.localdate()

        money_field = DecimalField(max_digits=14, decimal_places=2)
        hours_field = DecimalField(max_digits=12, decimal_places=2)

        workspaces = self.get_user_workspaces()
        current_workspace = self.get_current_workspace()

        if not current_workspace:
            ctx.update({
                "workspaces": workspaces,
                "current_workspace": None,
                "workspace": None,
                "stats": {},
                "finance": {},
                "task_distribution": {},
                "task_distribution_pct": {},
                "projects": [],
                "recent_projects": [],
                "high_risk_projects": [],
                "late_projects": [],
                "blocked_tasks": [],
                "due_soon_tasks": [],
                "active_sprint": None,
                "kanban_columns": [],
                "activities": [],
                "notifications": [],
                "ai_insights": [],
                "team_velocity": [],
                "team_load": [],
                "top_cost_members": [],
                "upcoming_milestones": [],
                "analysis_cards": [],
                "debug_info": {
                    "workspace_found": False,
                    "workspace_name": None,
                },
            })
            return ctx

        projects = (
            current_workspace.projects
            .filter(is_archived=False)
            .select_related("team", "owner", "product_manager")
            .prefetch_related(
                Prefetch(
                    "board_columns",
                    queryset=dm.BoardColumn.objects.order_by("position"),
                )
            )
            .order_by("name")
        )

        tasks = (
            current_workspace.tasks
            .filter(is_archived=False)
            .select_related("project", "sprint", "assignee", "reporter")
            .order_by("-updated_at")
        )

        memberships = (
            current_workspace.memberships
            .select_related("user", "team")
            .order_by("user__username")
        )

        user_profiles = (
            current_workspace.user_profiles
            .select_related("user", "workspace")
            .order_by("-cost_per_day", "user__username")
        )

        sprints = (
            current_workspace.sprints
            .filter(is_archived=False)
            .select_related("project", "team")
            .order_by("-start_date")
        )

        risks = (
            current_workspace.risks
            .filter(is_archived=False)
            .select_related("project", "task", "owner")
            .order_by("-created_at")
        )

        ai_insights = (
            current_workspace.ai_insights
            .filter(is_dismissed=False)
            .select_related("project", "sprint", "task")
            .order_by("-detected_at")[:6]
        )

        notifications = (
            current_workspace.notifications
            .filter(recipient=self.request.user, is_read=False)
            .order_by("-created_at")[:6]
        )

        activities = (
            current_workspace.activity_logs
            .select_related("actor", "project", "task", "sprint")
            .order_by("-created_at")[:10]
        )

        active_sprint = (
            sprints.filter(status=dm.Sprint.Status.ACTIVE)
            .order_by("end_date", "start_date")
            .first()
        )
        sprint_tasks = tasks.filter(sprint=active_sprint) if active_sprint else tasks.none()

        project_stats = projects.aggregate(
            projects_total=Count("id"),
            active_projects_count=Count(
                "id",
                filter=~Q(status__in=[dm.Project.Status.DONE, dm.Project.Status.CANCELLED]),
            ),
            delayed_projects_count=Count("id", filter=Q(status=dm.Project.Status.DELAYED)),
            completed_projects_count=Count("id", filter=Q(status=dm.Project.Status.DONE)),
            avg_project_progress=Coalesce(
                Avg("progress_percent"),
                Value(0.0),
                output_field=FloatField(),
            ),
            avg_project_risk=Coalesce(
                Avg("risk_score"),
                Value(0.0),
                output_field=FloatField(),
            ),
        )

        task_stats = tasks.aggregate(
            tasks_total=Count("id"),
            todo_count=Count("id", filter=Q(status=dm.Task.Status.TODO)),
            in_progress_count=Count("id", filter=Q(status=dm.Task.Status.IN_PROGRESS)),
            review_count=Count("id", filter=Q(status=dm.Task.Status.REVIEW)),
            done_tasks_count=Count("id", filter=Q(status=dm.Task.Status.DONE)),
            blocked_tasks_count=Count("id", filter=Q(status=dm.Task.Status.BLOCKED)),
            cancelled_tasks_count=Count("id", filter=Q(status=dm.Task.Status.CANCELLED)),
            flagged_tasks_count=Count("id", filter=Q(is_flagged=True)),
            total_estimate_hours=Coalesce(
                Sum("estimate_hours"),
                Value(Decimal("0.00")),
                output_field=hours_field,
            ),
            total_spent_hours=Coalesce(
                Sum("spent_hours"),
                Value(Decimal("0.00")),
                output_field=hours_field,
            ),
        )

        member_stats = memberships.aggregate(
            active_members_count=Count("id", filter=Q(status=dm.TeamMembership.Status.ACTIVE)),
            remote_members_count=Count("id", filter=Q(status=dm.TeamMembership.Status.REMOTE)),
            on_leave_members_count=Count("id", filter=Q(status=dm.TeamMembership.Status.ON_LEAVE)),
            inactive_members_count=Count("id", filter=Q(status=dm.TeamMembership.Status.INACTIVE)),
            avg_member_load=Coalesce(
                Avg("current_load_percent"),
                Value(0.0),
                output_field=FloatField(),
            ),
        )

        risk_stats = risks.aggregate(
            open_risks_count=Count("id", filter=~Q(status=dm.Risk.Status.CLOSED)),
            critical_risks_count=Count("id", filter=Q(severity=dm.Risk.Severity.CRITICAL)),
            escalated_risks_count=Count("id", filter=Q(status=dm.Risk.Status.ESCALATED)),
            avg_risk_priority=Coalesce(
                Avg("impact_score"),
                Value(0.0),
                output_field=FloatField(),
            ),
        )

        revenues_qs = dm.ProjectRevenue.objects.filter(project__workspace=current_workspace)
        expenses_qs = dm.ProjectExpense.objects.filter(project__workspace=current_workspace)
        budgets_qs = dm.ProjectBudget.objects.filter(project__workspace=current_workspace, is_archived=False)
        timesheets_qs = current_workspace.timesheet_entries.select_related("user", "project", "task")

        approved_budget_total = budgets_qs.aggregate(
            total=Coalesce(
                Sum("approved_budget"),
                Value(Decimal("0.00")),
                output_field=money_field,
            )
        )["total"]

        estimated_cost_total = budgets_qs.aggregate(
            total=(
                    Coalesce(Sum("estimated_labor_cost"), Value(Decimal("0.00")), output_field=money_field) +
                    Coalesce(Sum("estimated_software_cost"), Value(Decimal("0.00")), output_field=money_field) +
                    Coalesce(Sum("estimated_infra_cost"), Value(Decimal("0.00")), output_field=money_field) +
                    Coalesce(Sum("estimated_subcontract_cost"), Value(Decimal("0.00")), output_field=money_field) +
                    Coalesce(Sum("estimated_other_cost"), Value(Decimal("0.00")), output_field=money_field) +
                    Coalesce(Sum("contingency_amount"), Value(Decimal("0.00")), output_field=money_field)
            )
        )["total"]

        planned_revenue_total = budgets_qs.aggregate(
            total=Coalesce(
                Sum("planned_revenue"),
                Value(Decimal("0.00")),
                output_field=money_field,
            )
        )["total"]

        received_revenue_total = revenues_qs.aggregate(
            total=Coalesce(
                Sum("amount", filter=Q(is_received=True)),
                Value(Decimal("0.00")),
                output_field=money_field,
            )
        )["total"]

        expected_revenue_total = revenues_qs.aggregate(
            total=Coalesce(
                Sum("amount"),
                Value(Decimal("0.00")),
                output_field=money_field,
            )
        )["total"]

        estimated_expenses_total = expenses_qs.aggregate(
            total=Coalesce(
                Sum("amount", filter=Q(status=dm.ProjectExpense.ExpenseStatus.ESTIMATED)),
                Value(Decimal("0.00")),
                output_field=money_field,
            )
        )["total"]

        forecast_expenses_total = expenses_qs.aggregate(
            total=Coalesce(
                Sum("amount", filter=Q(status=dm.ProjectExpense.ExpenseStatus.FORECAST)),
                Value(Decimal("0.00")),
                output_field=money_field,
            )
        )["total"]

        committed_expenses_total = expenses_qs.aggregate(
            total=Coalesce(
                Sum("amount", filter=Q(status=dm.ProjectExpense.ExpenseStatus.COMMITTED)),
                Value(Decimal("0.00")),
                output_field=money_field,
            )
        )["total"]

        accrued_expenses_total = expenses_qs.aggregate(
            total=Coalesce(
                Sum("amount", filter=Q(status=dm.ProjectExpense.ExpenseStatus.ACCRUED)),
                Value(Decimal("0.00")),
                output_field=money_field,
            )
        )["total"]

        actual_expenses_total = expenses_qs.aggregate(
            total=Coalesce(
                Sum("amount", filter=Q(status=dm.ProjectExpense.ExpenseStatus.PAID)),
                Value(Decimal("0.00")),
                output_field=money_field,
            )
        )["total"]

        rejected_expenses_total = expenses_qs.aggregate(
            total=Coalesce(
                Sum("amount", filter=Q(status=dm.ProjectExpense.ExpenseStatus.REJECTED)),
                Value(Decimal("0.00")),
                output_field=money_field,
            )
        )["total"]

        logged_hours_total = timesheets_qs.aggregate(
            total=Coalesce(
                Sum("hours"),
                Value(Decimal("0.00")),
                output_field=hours_field,
            )
        )["total"]

        logged_cost_total = timesheets_qs.aggregate(
            total=Coalesce(
                Sum("cost_snapshot__computed_cost"),
                Value(Decimal("0.00")),
                output_field=money_field,
            )
        )["total"]

        budget_usage_percent = (
            self.pct(float(actual_expenses_total), float(approved_budget_total))
            if approved_budget_total else 0
        )

        margin_estimate = planned_revenue_total - estimated_cost_total
        actual_margin = received_revenue_total - actual_expenses_total

        finance = {
            "approved_budget_total": approved_budget_total,
            "estimated_cost_total": estimated_cost_total,
            "planned_revenue_total": planned_revenue_total,
            "expected_revenue_total": expected_revenue_total,
            "received_revenue_total": received_revenue_total,
            "actual_expenses_total": actual_expenses_total,
            # "draft_expenses_total": draft_expenses_total,
            "logged_hours_total": logged_hours_total,
            "logged_cost_total": logged_cost_total,
            "budget_usage_percent": budget_usage_percent,
            "margin_estimate": margin_estimate,
            "actual_margin": actual_margin,
            "currency": "XOF",
        }

        total_tasks = task_stats["tasks_total"] or 0
        done_tasks_count = task_stats["done_tasks_count"] or 0
        pending_tasks_count = max(total_tasks - done_tasks_count, 0)

        task_distribution = {
            "todo": task_stats["todo_count"],
            "in_progress": task_stats["in_progress_count"],
            "review": task_stats["review_count"],
            "done": task_stats["done_tasks_count"],
            "blocked": task_stats["blocked_tasks_count"],
            "cancelled": task_stats["cancelled_tasks_count"],
        }

        task_distribution_pct = {
            key: self.pct(value, total_tasks)
            for key, value in task_distribution.items()
        }

        stats = {
            **project_stats,
            **task_stats,
            **member_stats,
            **risk_stats,
            "pending_tasks_count": pending_tasks_count,
            "done_ratio_percent": self.pct(done_tasks_count, total_tasks),
            "blocked_ratio_percent": self.pct(task_stats["blocked_tasks_count"], total_tasks),
            "remote_ratio_percent": self.pct(
                member_stats["remote_members_count"],
                member_stats["active_members_count"],
            ),
            "project_completion_percent": self.pct(
                project_stats["completed_projects_count"],
                project_stats["projects_total"],
            ),
            "estimate_consumption_percent": self.pct(
                float(task_stats["total_spent_hours"]),
                float(task_stats["total_estimate_hours"]),
            ) if task_stats["total_estimate_hours"] else 0,
        }

        high_risk_projects = projects.filter(risk_score__gte=70).order_by("-risk_score", "target_date")[:6]
        late_projects = projects.filter(status=dm.Project.Status.DELAYED).order_by("target_date")[:6]
        blocked_tasks = tasks.filter(status=dm.Task.Status.BLOCKED).order_by("due_date", "-created_at")[:6]

        due_soon_tasks = (
            tasks.exclude(status__in=[dm.Task.Status.DONE, dm.Task.Status.CANCELLED])
            .filter(due_date__isnull=False, due_date__lte=today + timezone.timedelta(days=7))
            .order_by("due_date", "-priority")[:6]
        )

        recent_projects = projects.order_by("-updated_at")[:6]

        team_velocity = list(
            current_workspace.teams.filter(is_archived=False)
            .annotate(
                projects_count=Count("projects", filter=Q(projects__is_archived=False), distinct=True),
                sprints_count=Count("sprints", filter=Q(sprints__is_archived=False), distinct=True),
            )
            .order_by("-velocity_current", "name")[:6]
        )

        team_load = list(
            memberships.order_by("-current_load_percent", "user__username")[:6]
        )

        top_cost_members = list(
            user_profiles.order_by("-cost_per_day", "-billable_rate_per_day")[:6]
        )

        upcoming_milestones = list(
            current_workspace.milestones
            .filter(is_archived=False)
            .exclude(status=dm.Milestone.Status.DONE)
            .select_related("project", "owner")
            .order_by("due_date")[:6]
        )

        overloaded_members_count = memberships.filter(current_load_percent__gte=85).count()

        kanban_columns = []
        active_sprint_task_stats = {
            "todo": 0,
            "in_progress": 0,
            "review": 0,
            "done": 0,
        }

        if active_sprint:
            custom_columns = list(active_sprint.project.board_columns.order_by("position"))

            if custom_columns:
                for col in custom_columns:
                    col_tasks_qs = sprint_tasks.filter(
                        status=col.mapped_status) if col.mapped_status else sprint_tasks.none()
                    col_tasks = list(col_tasks_qs[:10])
                    kanban_columns.append({
                        "id": f"col-{col.pk}",
                        "name": col.name,
                        "color": col.color or "#7C6FF7",
                        "count": col_tasks_qs.count(),
                        "tasks": col_tasks,
                        "is_done": col.is_done_column,
                        "mapped_status": col.mapped_status,
                    })
            else:
                default_map = [
                    ("todo", "À faire", "#94A3B8", dm.Task.Status.TODO, False),
                    ("in_progress", "En cours", "#7C6FF7", dm.Task.Status.IN_PROGRESS, False),
                    ("review", "Review", "#E8950A", dm.Task.Status.REVIEW, False),
                    ("done", "Terminé", "#34A853", dm.Task.Status.DONE, True),
                ]
                for key, label, color, status_value, is_done in default_map:
                    col_tasks_qs = sprint_tasks.filter(status=status_value)
                    col_tasks = list(col_tasks_qs[:10])
                    kanban_columns.append({
                        "id": f"col-{key}",
                        "name": label,
                        "color": color,
                        "count": col_tasks_qs.count(),
                        "tasks": col_tasks,
                        "is_done": is_done,
                        "mapped_status": status_value,
                    })

            active_sprint_task_stats = {
                "todo": sprint_tasks.filter(status=dm.Task.Status.TODO).count(),
                "in_progress": sprint_tasks.filter(status=dm.Task.Status.IN_PROGRESS).count(),
                "review": sprint_tasks.filter(status=dm.Task.Status.REVIEW).count(),
                "done": sprint_tasks.filter(status=dm.Task.Status.DONE).count(),
            }

        analysis_cards = self.build_analysis_cards(
            stats=stats,
            finance=finance,
            active_sprint=active_sprint,
            delayed_projects=project_stats["delayed_projects_count"],
            blocked_tasks=task_stats["blocked_tasks_count"],
            critical_risks=risk_stats["critical_risks_count"],
            overloaded_members_count=overloaded_members_count,
        )

        ctx.update({
            "workspaces": workspaces,
            "current_workspace": current_workspace,
            "workspace": current_workspace,

            "projects": projects[:8],
            "recent_projects": recent_projects,
            "tasks": tasks[:12],

            "active_sprint": active_sprint,
            "sprint_tasks": sprint_tasks,
            "kanban_columns": kanban_columns,
            "active_sprint_task_stats": active_sprint_task_stats,

            "late_projects": late_projects,
            "blocked_tasks": blocked_tasks,
            "due_soon_tasks": due_soon_tasks,
            "high_risk_projects": high_risk_projects,
            "upcoming_milestones": upcoming_milestones,

            "done_tasks_count": stats["done_tasks_count"],
            "pending_tasks_count": stats["pending_tasks_count"],
            "active_projects_count": stats["active_projects_count"],
            "active_members_count": stats["active_members_count"],
            "remote_members_count": stats["remote_members_count"],

            "ai_insights": ai_insights,
            "notifications": notifications,
            "activities": activities,

            "stats": stats,
            "finance": finance,
            "task_distribution": task_distribution,
            "task_distribution_pct": task_distribution_pct,

            "team_velocity": team_velocity,
            "team_load": team_load,
            "top_cost_members": top_cost_members,

            "analysis_cards": analysis_cards,

            "debug_info": {
                "workspace_found": True,
                "workspace_name": current_workspace.name,
                "projects_count": projects.count(),
                "tasks_count": tasks.count(),
                "memberships_count": memberships.count(),
                "profiles_count": user_profiles.count(),
                "sprints_count": sprints.count(),
                "risks_count": risks.count(),
                "notifications_count": notifications.count(),
                "activities_count": activities.count(),
            },
        })
        return ctx


# =============================================================================
# WORKSPACES
# =============================================================================
class WorkspaceListView(DevflowListView):
    model = dm.Workspace
    form_class = WorkspaceForm
    template_name = "project/workspace/list.html"
    section = "workspace"
    page_title = "Workspaces"
    search_fields = ("name", "slug", "description", "quarter_label")

    def get_queryset(self):
        return (
            dm.Workspace.objects
            .select_related("owner")
            .annotate(
                teams_count=Count("teams", distinct=True),
                projects_count=Count("projects", distinct=True),
                memberships_count=Count("memberships", distinct=True),
            )
            .order_by("name")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["total_workspaces"] = self.get_queryset().count()
        ctx["active_workspaces"] = self.get_queryset().filter(is_active=True, is_archived=False).count()
        ctx["archived_workspaces"] = self.get_queryset().filter(is_archived=True).count()
        return ctx


class WorkspaceDetailView(DevflowDetailView):
    model = dm.Workspace
    template_name = "project/workspace/detail.html"
    section = "workspace"
    page_title = "Détail workspace"

    def get_queryset(self):
        return (
            dm.Workspace.objects
            .select_related("owner")
            .prefetch_related(
                Prefetch(
                    "teams",
                    queryset=dm.Team.objects.filter(is_archived=False)
                             .select_related("workspace")
                             .order_by("name")[:10],
                    to_attr="prefetched_teams",
                ),
                Prefetch(
                    "projects",
                    queryset=dm.Project.objects.filter(is_archived=False)
                             .select_related("team", "owner", "product_manager", "workspace")
                             .prefetch_related(
                        Prefetch(
                            "members",
                            queryset=dm.ProjectMember.objects.select_related("user", "team"),
                        )
                    )
                             .order_by("name")[:10],
                    to_attr="prefetched_projects",
                ),
                Prefetch(
                    "user_profiles",
                    queryset=dm.UserProfile.objects.select_related("user", "workspace")
                             .order_by("user__username")[:10],
                    to_attr="prefetched_user_profiles",
                ),
            )
            .annotate(
                teams_count=Count("teams", distinct=True),
                projects_count=Count("projects", distinct=True),
                members_count=Count("user_profiles", distinct=True),
            )
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        workspace = self.object

        ctx.update({
            "teams": getattr(workspace, "prefetched_teams", []),
            "projects": getattr(workspace, "prefetched_projects", []),
            "members": getattr(workspace, "prefetched_user_profiles", []),
            "settings_obj": getattr(workspace, "settings", None),
            "snapshots": getattr(workspace, "dashboard_snapshots", []).all()[:10] if hasattr(workspace,
                                                                                             "dashboard_snapshots") else [],
            "integrations": getattr(workspace, "integrations", []).all()[:10] if hasattr(workspace,
                                                                                         "integrations") else [],
        })
        return ctx


class WorkspaceCreateView(DevflowCreateView):
    model = dm.Workspace
    form_class = WorkspaceForm
    template_name = "project/workspace/form.html"
    section = "workspace"
    page_title = "Créer workspace"
    success_list_url_name = "workspace_list"

    def form_valid(self, form):
        if not form.instance.owner_id:
            form.instance.owner = self.request.user
        messages.success(self.request, "Workspace créé avec succès.")
        return super().form_valid(form)


class WorkspaceUpdateView(DevflowUpdateView):
    model = dm.Workspace
    form_class = WorkspaceForm
    template_name = "project/workspace/form.html"
    section = "workspace"
    page_title = "Modifier workspace"
    success_list_url_name = "workspace_list"

    def form_valid(self, form):
        messages.success(self.request, "Workspace mis à jour avec succès.")
        return super().form_valid(form)


class WorkspaceDeleteView(DevflowDeleteView):
    model = dm.Workspace
    template_name = "project/workspace/confirm_delete.html"
    section = "workspace"
    page_title = "Supprimer workspace"
    success_list_url_name = "workspace_list"

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Workspace supprimé avec succès.")
        return super().delete(request, *args, **kwargs)


class WorkspaceArchiveView(ArchiveObjectView):
    model = dm.Workspace
    success_list_url_name = "workspace_list"

    def post(self, request, *args, **kwargs):
        messages.success(request, "Workspace archivé avec succès.")
        return super().post(request, *args, **kwargs)


# =============================================================================
# TEAMS
# =============================================================================
class TeamListView(DevflowListView):
    model = dm.Team
    template_name = "project/team/list.html"
    section = "team"
    page_title = "Équipes"
    search_placeholder = "Rechercher par nom, slug ou description…"
    search_fields = ("name", "slug", "description", "team_type")
    filter_fields = [
        {"name": "team_type", "label": "Type", "type": "choices",
         "choices": dm.Team.TeamType.choices},
        {"name": "is_active", "label": "Active", "type": "bool"},
    ]


class TeamDetailView(DevflowDetailView):
    model = dm.Team
    template_name = "project/team/detail.html"
    section = "team"
    page_title = "Détail équipe"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        team = self.object
        ctx["memberships"] = team.memberships.select_related("user")
        ctx["projects"] = team.projects.filter(is_archived=False)
        ctx["sprints"] = team.sprints.filter(is_archived=False)[:10]
        return ctx


class TeamCreateView(DevflowCreateView):
    model = dm.Team
    form_class = TeamForm
    template_name = "project/team/form.html"
    section = "team"
    page_title = "Créer équipe"
    success_list_url_name = "team_list"


class TeamUpdateView(DevflowUpdateView):
    model = dm.Team
    form_class = TeamForm
    template_name = "project/team/form.html"
    section = "team"
    page_title = "Modifier équipe"
    success_list_url_name = "team_list"


class TeamDeleteView(DevflowDeleteView):
    model = dm.Team
    section = "team"
    page_title = "Supprimer équipe"
    success_list_url_name = "team_list"


class TeamArchiveView(ArchiveObjectView):
    model = dm.Team
    success_list_url_name = "team_list"


class TeamMembershipListView(DevflowListView):
    model = dm.TeamMembership
    template_name = "project/team_membership/list.html"
    section = "team"
    page_title = "Membres d'équipe"
    search_placeholder = "Rechercher par nom, email ou job title…"
    search_fields = (
        "job_title", "role", "status",
        "user__username", "user__first_name", "user__last_name", "user__email",
    )
    filter_fields = [
        {"name": "team", "label": "Équipe", "type": "model",
         "queryset": "_teams_qs", "lookup": "team_id"},
        {"name": "role", "label": "Rôle", "type": "choices",
         "choices": dm.TeamMembership.Role.choices},
        {"name": "status", "label": "Statut", "type": "choices",
         "choices": dm.TeamMembership.Status.choices},
    ]

    def _teams_qs(self):
        ws = self.get_current_workspace()
        return dm.Team.objects.filter(workspace=ws, is_archived=False).order_by("name") if ws else dm.Team.objects.none()


class TeamMembershipDetailView(DevflowDetailView):
    model = dm.TeamMembership
    template_name = "project/team_membership/detail.html"
    section = "team"
    page_title = "Détail appartenance équipe"


class TeamMembershipCreateView(DevflowCreateView):
    model = dm.TeamMembership
    form_class = TeamMembershipForm
    template_name = "project/team_membership/form.html"
    section = "team"
    page_title = "Créer appartenance équipe"
    success_list_url_name = "team_membership_list"

    def get_initial(self):
        initial = super().get_initial()
        team_id = self.request.GET.get("team")
        user_id = self.request.GET.get("user")
        if team_id:
            initial["team"] = team_id
        if user_id:
            initial["user"] = user_id
        return initial


class TeamMembershipUpdateView(DevflowUpdateView):
    model = dm.TeamMembership
    form_class = TeamMembershipForm
    template_name = "project/team_membership/form.html"
    section = "team"
    page_title = "Modifier appartenance équipe"
    success_list_url_name = "team_membership_list"


class TeamMembershipDeleteView(DevflowDeleteView):
    model = dm.TeamMembership
    section = "team"
    page_title = "Supprimer appartenance équipe"
    success_list_url_name = "team_membership_list"


# =============================================================================
# PROJECTS
# =============================================================================

class ProjectListView(DevflowListView):
    model = dm.Project
    template_name = "project/list.html"
    section = "project"
    page_title = "Projets"
    context_object_name = "items"
    paginate_by = 12
    search_placeholder = "Rechercher par nom, code, description ou stack…"
    search_fields = ("name", "slug", "code", "description", "tech_stack", "ai_risk_label")
    filter_fields = [
        {"name": "status", "label": "Statut", "type": "choices",
         "choices": dm.Project.Status.choices},
        {"name": "priority", "label": "Priorité", "type": "choices",
         "choices": dm.Project.Priority.choices},
        {"name": "health_status", "label": "Santé", "type": "choices",
         "choices": dm.Project.HealthStatus.choices},
        {"name": "team", "label": "Équipe", "type": "model",
         "queryset": "_teams_qs", "lookup": "team_id"},
    ]

    def _teams_qs(self):
        ws = self.get_current_workspace()
        return dm.Team.objects.filter(workspace=ws, is_archived=False).order_by("name") if ws else dm.Team.objects.none()
    PROJECT_TIME_FIELDS = {
        "estimated": ("estimated_hours", "planned_hours", "forecast_hours"),
        "spent": ("spent_hours", "actual_hours", "logged_hours", "consumed_hours"),
    }

    TASK_TIME_FIELDS = {
        "estimated": ("estimated_hours", "planned_hours", "forecast_hours"),
        "spent": ("spent_hours", "actual_hours", "logged_hours", "consumed_hours"),
    }
    PROJECT_FINANCE_FIELDS = {
        "budget": ("budget", "estimated_budget", "planned_budget"),
        "billed": ("billed_amount", "invoiced_amount", "amount_billed", "revenue_amount"),
        "cost": ("cost_amount", "actual_cost", "internal_cost", "cost_total"),

    }

    def _to_decimal(self, value):

        if value in (None, ""):

            return Decimal("0")

        try:

            return Decimal(str(value))

        except Exception:

            return Decimal("0")

    def _first_attr_value(self, obj, field_names):

        for field_name in field_names:

            if hasattr(obj, field_name):

                value = getattr(obj, field_name, None)

                if value not in (None, ""):

                    return value

        return None

    def _sum_task_field(self, project, field_names):

        total = Decimal("0")

        tasks = getattr(project, "prefetched_tasks", None)

        if tasks is None and hasattr(project, "tasks"):

            tasks = project.tasks.all()

        elif tasks is None:

            tasks = []

        for task in tasks:

            value = self._first_attr_value(task, field_names)

            total += self._to_decimal(value)

        return total

    def _project_metrics(self, project):

        budget = self._to_decimal(

            self._first_attr_value(project, self.PROJECT_FINANCE_FIELDS["budget"])

        )

        billed = self._to_decimal(

            self._first_attr_value(project, self.PROJECT_FINANCE_FIELDS["billed"])

        )

        cost = self._to_decimal(

            self._first_attr_value(project, self.PROJECT_FINANCE_FIELDS["cost"])

        )

        estimated_time = self._to_decimal(

            self._first_attr_value(project, self.PROJECT_TIME_FIELDS["estimated"])

        )

        spent_time = self._to_decimal(

            self._first_attr_value(project, self.PROJECT_TIME_FIELDS["spent"])

        )

        if estimated_time == 0:

            estimated_time = self._sum_task_field(project, self.TASK_TIME_FIELDS["estimated"])

        if spent_time == 0:

            spent_time = self._sum_task_field(project, self.TASK_TIME_FIELDS["spent"])

        if billed == 0:

            category = getattr(project, "category", None)

            if category and getattr(category, "is_billable", False):

                billed = budget

        pnl = billed - cost

        task_count = getattr(project, "tasks_count", None)

        if task_count is None:

            if hasattr(project, "tasks"):

                task_count = project.tasks.count()

            else:

                task_count = 0

        return {

            "budget": budget,

            "billed": billed,

            "cost": cost,

            "estimated_time": estimated_time,

            "spent_time": spent_time,

            "task_count": task_count,

            "pnl": pnl,

        }

    def _build_category_sections(self, projects):

        grouped = OrderedDict()

        for project in projects:

            category = getattr(project, "category", None)

            category_key = category.pk if category else "uncategorized"

            category_label = category.name if category else "Sans catégorie"

            if category_key not in grouped:

                grouped[category_key] = {

                    "category": category,

                    "label": category_label,

                    "color": getattr(category, "color", "") if category else "",

                    "is_billable": getattr(category, "is_billable", False) if category else False,

                    "budget_type": getattr(category, "get_budget_type_display", lambda: "—")() if category else "—",

                    "projects": [],

                    "summary": {

                        "projects_count": 0,

                        "task_count": 0,

                        "budget": Decimal("0"),

                        "billed": Decimal("0"),

                        "cost": Decimal("0"),

                        "estimated_time": Decimal("0"),

                        "spent_time": Decimal("0"),

                        "pnl": Decimal("0"),

                    },

                }

            metrics = self._project_metrics(project)

            grouped[category_key]["projects"].append(project)

            grouped[category_key]["summary"]["projects_count"] += 1

            grouped[category_key]["summary"]["task_count"] += metrics["task_count"]

            grouped[category_key]["summary"]["budget"] += metrics["budget"]

            grouped[category_key]["summary"]["billed"] += metrics["billed"]

            grouped[category_key]["summary"]["cost"] += metrics["cost"]

            grouped[category_key]["summary"]["estimated_time"] += metrics["estimated_time"]

            grouped[category_key]["summary"]["spent_time"] += metrics["spent_time"]

            grouped[category_key]["summary"]["pnl"] += metrics["pnl"]

        return list(grouped.values())

    def get_queryset(self):

        current_workspace = self.get_current_workspace()

        queryset = (

            super()

            .get_queryset()

            .filter(is_archived=False)

            .select_related(

                "workspace",

                "team",

                "owner",

                "product_manager",

                "category",

            )

            .prefetch_related(

                Prefetch(

                    "tasks",

                    queryset=dm.Task.objects.filter(is_archived=False) if hasattr(dm, "Task") else None,

                    to_attr="prefetched_tasks",

                )

            )

            .annotate(tasks_count=Count("tasks", distinct=True))

        )

        if current_workspace:

            queryset = queryset.filter(workspace=current_workspace)

        else:

            queryset = queryset.none()

        status_filter = self.request.GET.get("status")

        priority_filter = self.request.GET.get("priority")

        health_filter = self.request.GET.get("health")

        category_filter = self.request.GET.get("category")

        ordering = self.request.GET.get("ordering")

        if status_filter:

            queryset = queryset.filter(status=status_filter)

        if priority_filter:

            queryset = queryset.filter(priority=priority_filter)

        if health_filter:

            queryset = queryset.filter(health_status=health_filter)

        if category_filter:

            queryset = queryset.filter(category_id=category_filter)

        ordering_map = {

            "name": "name",

            "-name": "-name",

            "progress": "progress_percent",

            "-progress": "-progress_percent",

            "target_date": "target_date",

            "-target_date": "-target_date",

            "created_at": "created_at",

            "-created_at": "-created_at",

            "category": "category__name",

            "-category": "-category__name",

        }

        if ordering in ordering_map:

            queryset = queryset.order_by(ordering_map[ordering], "-created_at")

        else:

            queryset = queryset.order_by("category__name", "name", "-created_at")

        return queryset

    def get_context_data(self, **kwargs):

        ctx = super().get_context_data(**kwargs)

        today = timezone.localdate()

        filtered_queryset = self.get_queryset()

        category_sections = self._build_category_sections(filtered_queryset)

        current_workspace = self.get_current_workspace()

        base_qs = (

            dm.Project.objects

            .filter(is_archived=False)

            .select_related("team", "workspace", "owner", "product_manager", "category")

        )

        if current_workspace:

            base_qs = base_qs.filter(workspace=current_workspace)

        else:

            base_qs = base_qs.none()

        ctx.update({

            "status_choices": dm.Project.Status.choices,

            "priority_choices": dm.Project.Priority.choices,

            "health_choices": dm.Project.HealthStatus.choices,

            "category_choices": dm.ProjectCategory.objects.order_by("name"),

            "current_status": self.request.GET.get("status", ""),

            "current_priority": self.request.GET.get("priority", ""),

            "current_health": self.request.GET.get("health", ""),

            "current_category": self.request.GET.get("category", ""),

            "current_ordering": self.request.GET.get("ordering", ""),

            "current_search": self.request.GET.get("q", ""),

            "category_sections": category_sections,

            "stats": {

                "total": base_qs.count(),

                "in_progress": base_qs.filter(status=dm.Project.Status.IN_PROGRESS).count(),

                "done": base_qs.filter(status=dm.Project.Status.DONE).count(),

                "blocked": base_qs.filter(status=dm.Project.Status.BLOCKED).count(),

                "delayed": base_qs.filter(

                    Q(status=dm.Project.Status.DELAYED) |

                    Q(

                        target_date__lt=today,

                        status__in=[

                            dm.Project.Status.PLANNED,

                            dm.Project.Status.IN_PROGRESS,

                            dm.Project.Status.IN_DELIVERY,

                            dm.Project.Status.BLOCKED,

                            dm.Project.Status.ON_HOLD,

                        ],

                    )

                ).count(),

                "critical": base_qs.filter(priority=dm.Project.Priority.CRITICAL).count(),

            }

        })

        return ctx


class ProjectBudgetExportExcelView(LoginRequiredMixin, View):

    def get(self, request, pk):

        project = get_object_or_404(

            dm.Project.objects.select_related("workspace", "budgetestimatif"),

            pk=pk,

        )

        estimate_lines = (

            project.estimate_lines.select_related("category")

            .order_by("budget_stage", "label", "id")

        )

        budget = getattr(project, "budgetestimatif", None)

        wb = Workbook()

        ws = wb.active

        ws.title = "Budget estimatif"

        header_fill = PatternFill("solid", fgColor="1F2937")

        header_font = Font(color="FFFFFF", bold=True)

        thin = Side(style="thin", color="D1D5DB")

        currency = budget.currency if budget else "XOF"

        def style_header(cell):

            cell.fill = header_fill

            cell.font = header_font

            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            cell.alignment = Alignment(horizontal="center", vertical="center")

        def style_money(cell):

            cell.number_format = '#,##0.00'

        # En-tête méta

        ws["A1"] = "Projet"

        ws["B1"] = project.name

        ws["A2"] = "Code"

        ws["B2"] = project.code or "-"

        ws["A3"] = "Workspace"

        ws["B3"] = project.workspace.name

        ws["A4"] = "Date export"

        ws["B4"] = timezone.localtime().strftime("%d/%m/%Y %H:%M")

        ws["A5"] = "Devise"

        ws["B5"] = currency

        # Tableau principal

        row = 7

        headers = [

            "Libellé",

            "Catégorie",

            "Stage budget",

            "Quantité",

            "Coût unitaire",

            "Coût total",

            "Vente unitaire",

            "Vente totale",

            "Marge",

        ]

        for col, value in enumerate(headers, start=1):

            cell = ws.cell(row=row, column=col, value=value)

            style_header(cell)

        row += 1

        total_cost = Decimal("0.00")

        total_sale = Decimal("0.00")

        total_margin = Decimal("0.00")

        for line in estimate_lines:

            cost_amount = line.cost_amount or Decimal("0.00")

            sale_amount = line.sale_amount or Decimal("0.00")

            margin_amount = line.margin_amount or Decimal("0.00")

            ws.cell(row=row, column=1, value=line.label or "")

            ws.cell(row=row, column=2, value=line.category.name if line.category else "")

            ws.cell(

                row=row,

                column=3,

                value=line.get_budget_stage_display() if hasattr(line, "get_budget_stage_display") else (line.budget_stage or ""),

            )

            ws.cell(row=row, column=4, value=float(line.quantity or 0))

            ws.cell(row=row, column=5, value=float(line.cost_unit_amount or 0))

            ws.cell(row=row, column=6, value=float(cost_amount))

            ws.cell(row=row, column=7, value=float(line.sale_unit_amount or 0))

            ws.cell(row=row, column=8, value=float(sale_amount))

            ws.cell(row=row, column=9, value=float(margin_amount))

            for money_col in [5, 6, 7, 8, 9]:

                style_money(ws.cell(row=row, column=money_col))

            total_cost += cost_amount

            total_sale += sale_amount

            total_margin += margin_amount

            row += 1

        # Ligne total

        ws.cell(row=row, column=5, value="TOTAL")

        ws.cell(row=row, column=6, value=float(total_cost))

        ws.cell(row=row, column=8, value=float(total_sale))

        ws.cell(row=row, column=9, value=float(total_margin))

        for col in [5, 6, 8, 9]:

            ws.cell(row=row, column=col).font = Font(bold=True)

            if col != 5:

                style_money(ws.cell(row=row, column=col))

        row += 2

        # Bloc synthèse budget

        if budget:

            summary_rows = [

                ("Statut budget", budget.get_status_display() if hasattr(budget, "get_status_display") else budget.status),

                ("Version", budget.version_number),

                ("Coût main d'œuvre estimé", float(budget.estimated_labor_cost or 0)),

                ("Coût logiciels estimé", float(budget.estimated_software_cost or 0)),

                ("Coût infrastructure estimé", float(budget.estimated_infra_cost or 0)),

                ("Coût sous-traitance estimé", float(budget.estimated_subcontract_cost or 0)),

                ("Autres coûts estimés", float(budget.estimated_other_cost or 0)),

                ("Contingence", float(budget.contingency_amount or 0)),

                ("Réserve management", float(budget.management_reserve_amount or 0)),

                ("Coût total estimé", float(budget.total_estimated_cost or 0)),

                ("Markup (%)", float(budget.markup_percent or 0)),

                ("Revenu planifié", float(budget.planned_revenue or 0)),

                ("Revenu attendu", float(budget.expected_revenue_amount or 0)),

                ("Marge estimée", float(budget.estimated_margin_amount or 0)),

                ("Marge estimée (%)", float(budget.estimated_margin_percent or 0)),

                ("Budget approuvé", float(budget.approved_budget or 0)),

                ("Taux alerte (%)", float(budget.alert_threshold_percent or 0)),

            ]

            ws.cell(row=row, column=1, value="Synthèse budget")

            ws.cell(row=row, column=1).font = Font(bold=True, size=12)

            row += 1

            money_labels = {

                "Coût main d'œuvre estimé",

                "Coût logiciels estimé",

                "Coût infrastructure estimé",

                "Coût sous-traitance estimé",

                "Autres coûts estimés",

                "Contingence",

                "Réserve management",

                "Coût total estimé",

                "Revenu planifié",

                "Revenu attendu",

                "Marge estimée",

                "Budget approuvé",

            }

            percent_labels = {

                "Markup (%)",

                "Marge estimée (%)",

                "Taux alerte (%)",

            }

            for label, value in summary_rows:

                ws.cell(row=row, column=1, value=label)

                ws.cell(row=row, column=2, value=value)

                if label in money_labels:

                    style_money(ws.cell(row=row, column=2))

                elif label in percent_labels:

                    ws.cell(row=row, column=2).number_format = '0.00'

                row += 1

        # Largeurs colonnes

        widths = {

            "A": 34,

            "B": 22,

            "C": 20,

            "D": 12,

            "E": 16,

            "F": 16,

            "G": 16,

            "H": 16,

            "I": 16,

        }

        for col, width in widths.items():

            ws.column_dimensions[col].width = width

        stream = BytesIO()

        wb.save(stream)

        stream.seek(0)

        filename = f"budget_estimatif_{project.slug or project.pk}.xlsx"

        response = HttpResponse(

            stream.getvalue(),

            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        )

        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response




class ProjectDetailView(DevflowDetailView):
    model = dm.Project
    template_name = "project/detail.html"
    section = "project"
    page_title = "Détail projet"

    def get_queryset(self):
        return (
            dm.Project.objects.select_related(
                "workspace",
                "team",
                "owner",
                "product_manager",
                "category",
            ).prefetch_related(
                Prefetch(
                    "members",
                    queryset=dm.ProjectMember.objects.select_related(
                        "user",
                        "user__profile",
                        "team",
                    ).order_by("user__username"),
                ),
                Prefetch(
                    "labels",
                    queryset=dm.ProjectLabel.objects.select_related("label").order_by("label__name"),
                ),
                Prefetch(
                    "board_columns",
                    queryset=dm.BoardColumn.objects.order_by("position", "id"),
                ),
                Prefetch(
                    "tasks",
                    queryset=(
                        dm.Task.objects.filter(is_archived=False)
                        .select_related("assignee", "reporter", "sprint", "backlog_item", "parent", "workspace")
                        .prefetch_related(
                            "labels__label",
                            "assignments__user",
                            "checklists__items",
                            "attachments",
                            "comments__author",
                            "incoming_dependencies__from_task",
                            "outgoing_dependencies__to_task",
                            "releases",
                            "milestones",
                        )
                        .order_by("-created_at")
                    ),
                ),
                Prefetch(
                    "sprints",
                    queryset=(
                        dm.Sprint.objects.filter(is_archived=False)
                        .select_related("team", "workspace", "project")
                        .prefetch_related(
                            "metrics",
                            "financial_snapshots",
                        )
                        .order_by("-start_date", "-created_at")
                    ),
                ),
                Prefetch(
                    "backlog_items",
                    queryset=(
                        dm.BacklogItem.objects.filter(is_archived=False)
                        .select_related("sprint", "parent", "reporter", "workspace", "project")
                        .prefetch_related(
                            "children",
                            "tasks",
                            "financial_snapshots",
                        )
                        .order_by("rank", "-created_at")
                    ),
                ),
                Prefetch(
                    "milestones",
                    queryset=(
                        dm.Milestone.objects.filter(is_archived=False)
                        .select_related("owner", "workspace", "project")
                        .prefetch_related(
                            "milestone_tasks__task",
                            "roadmap_items",
                        )
                        .order_by("due_date", "created_at")
                    ),
                ),
                Prefetch(
                    "releases",
                    queryset=(
                        dm.Release.objects.filter(is_archived=False)
                        .prefetch_related("tasks", "sprints")
                        .order_by("-release_date", "-created_at")
                    ),
                ),
                Prefetch(
                    "risks",
                    queryset=(
                        dm.Risk.objects.filter(is_archived=False)
                        .select_related("owner", "task", "workspace", "project")
                        .order_by("-created_at")
                    ),
                ),
                Prefetch(
                    "pull_requests",
                    queryset=dm.PullRequest.objects.select_related("author", "task").order_by("-created_at"),
                ),
                Prefetch(
                    "ai_insights",
                    queryset=dm.AInsight.objects.select_related("task", "sprint").order_by("-detected_at"),
                ),
                Prefetch(
                    "activity_logs",
                    queryset=dm.ActivityLog.objects.select_related("actor", "task", "sprint").order_by("-created_at"),
                ),
                Prefetch(
                    "timesheet_entries",
                    queryset=(
                        dm.TimesheetEntry.objects.select_related(
                            "user",
                            "task",
                            "approved_by",
                            "cost_snapshot",
                        ).order_by("-entry_date", "-created_at")
                    ),
                ),
                Prefetch(
                    "estimate_lines",
                    queryset=(
                        dm.ProjectEstimateLine.objects.select_related(
                            "category",
                            "task",
                            "sprint",
                            "milestone",
                            "created_by",
                        ).order_by("budget_stage", "label", "id")
                    ),
                ),
                Prefetch(
                    "expenses",
                    queryset=(
                        dm.ProjectExpense.objects.select_related(
                            "category",
                            "task",
                            "sprint",
                            "milestone",
                            "created_by",
                            "validated_by",
                            "level1_approved_by",
                            "level2_approved_by",
                            "rejected_by",
                        ).order_by("-expense_date", "-created_at")
                    ),
                ),
                Prefetch(
                    "revenues",
                    queryset=dm.ProjectRevenue.objects.order_by("expected_date", "title", "id"),
                ),
                Prefetch(
                    "source_imports",
                    queryset=dm.ProjectDocumentImport.objects.select_related("uploaded_by").order_by("-created_at"),
                ),
                Prefetch(
                    "kpis",
                    queryset=dm.ProjectKPI.objects.order_by("module_name", "name", "id"),
                ),
                Prefetch(
                    "module_rois",
                    queryset=dm.ProjectModuleROI.objects.order_by("module_name", "id"),
                ),
                Prefetch(
                    "roadmap_items",
                    queryset=(
                        dm.RoadmapItem.objects.select_related("roadmap", "milestone", "project")
                        .order_by("start_date", "row", "title", "id")
                    ),
                ),
            )
        )

    def can_view_financial_data(self, project):
        user = self.request.user

        if not user.is_authenticated:
            return False

        if user.is_superuser:
            return True

        if user.has_perm("project.view_financial_data") or user.has_perm("project.view_projectexpense_financial"):
            return True

        membership = project.workspace.memberships.filter(user=user).first()
        if not membership:
            return False

        return membership.role in [
            dm.TeamMembership.Role.ADMIN,
            dm.TeamMembership.Role.CTO,
            dm.TeamMembership.Role.PM,
            dm.TeamMembership.Role.PRODUCT_OWNER,
            dm.TeamMembership.Role.TECH_LEAD,
        ]

    def _get_project_working_days(self, project):
        """
        Calcule le nombre de jours ouvrés du projet.
        Approche simple : lundi-vendredi.
        """
        if not project.start_date or not project.target_date:
            return 0

        if project.target_date < project.start_date:
            return 0

        current = project.start_date
        total = 0
        while current <= project.target_date:
            if current.weekday() < 5:  # 0=lundi ... 4=vendredi
                total += 1
            current += timezone.timedelta(days=1)
        return total

    def _build_member_cost_summary(self, members_qs, project):
        """
        Synthèse RH calculée à partir :
        - BillingRate.get_user_daily_cost(user)
        - BillingRate.get_user_sale_daily_rate(user)
        - allocation_percent de ProjectMember
        """
        member_cost_rows = []

        daily_cost_total = Decimal("0.00")
        daily_sale_total = Decimal("0.00")
        monthly_cost_total = Decimal("0.00")
        monthly_sale_total = Decimal("0.00")
        project_cost_total = Decimal("0.00")
        project_sale_total = Decimal("0.00")

        project_working_days = self._get_project_working_days(project)
        standard_month_working_days = Decimal("22")

        for member in members_qs:
            user = member.user
            allocation_percent = Decimal(member.allocation_percent or 0)
            allocation_ratio = allocation_percent / Decimal("100")

            daily_cost = dm.BillingRate.get_user_daily_cost(user) or Decimal("0.00")
            daily_sale = dm.BillingRate.get_user_sale_daily_rate(user) or Decimal("0.00")

            allocated_daily_cost = daily_cost * allocation_ratio
            allocated_daily_sale = daily_sale * allocation_ratio

            allocated_monthly_cost = allocated_daily_cost * standard_month_working_days
            allocated_monthly_sale = allocated_daily_sale * standard_month_working_days

            allocated_project_cost = allocated_daily_cost * Decimal(project_working_days)
            allocated_project_sale = allocated_daily_sale * Decimal(project_working_days)

            daily_cost_total += allocated_daily_cost
            daily_sale_total += allocated_daily_sale
            monthly_cost_total += allocated_monthly_cost
            monthly_sale_total += allocated_monthly_sale
            project_cost_total += allocated_project_cost
            project_sale_total += allocated_project_sale

            profile = getattr(user, "profile", None)

            member_cost_rows.append({
                "member": member,
                "user": user,
                "team": member.team,
                "role": member.role,
                "allocation_percent": allocation_percent,
                "profile_cost_per_day": getattr(profile, "cost_per_day", Decimal("0.00")) if profile else Decimal("0.00"),
                "profile_billable_rate_per_day": getattr(profile, "billable_rate_per_day", Decimal("0.00")) if profile else Decimal("0.00"),
                "daily_cost": daily_cost,
                "daily_sale": daily_sale,
                "allocated_daily_cost": allocated_daily_cost,
                "allocated_daily_sale": allocated_daily_sale,
                "allocated_monthly_cost": allocated_monthly_cost,
                "allocated_monthly_sale": allocated_monthly_sale,
                "allocated_project_cost": allocated_project_cost,
                "allocated_project_sale": allocated_project_sale,
            })

        summary = {
            "count": len(member_cost_rows),
            "project_working_days": project_working_days,
            "daily_cost_total": daily_cost_total,
            "daily_sale_total": daily_sale_total,
            "monthly_cost_total": monthly_cost_total,
            "monthly_sale_total": monthly_sale_total,
            "project_cost_total": project_cost_total,
            "project_sale_total": project_sale_total,
            "project_margin_total": project_sale_total - project_cost_total,
        }

        return member_cost_rows, summary

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        project = self.object
        today = timezone.localdate()
        can_view_financials = self.can_view_financial_data(project)

        money_field = DecimalField(max_digits=14, decimal_places=2)
        hours_field = DecimalField(max_digits=12, decimal_places=2)

        members_qs = project.members.all()
        tasks_qs = project.tasks.filter(is_archived=False)
        sprints_qs = project.sprints.filter(is_archived=False)
        backlog_qs = project.backlog_items.filter(is_archived=False)
        milestones_qs = project.milestones.filter(is_archived=False)
        releases_qs = project.releases.filter(is_archived=False)
        risks_qs = project.risks.filter(is_archived=False)
        prs_qs = project.pull_requests.all()
        ai_insights_qs = project.ai_insights.all()
        activity_qs = project.activity_logs.all()
        board_columns_qs = project.board_columns.all()
        roadmap_items_qs = project.roadmap_items.all()
        imports_qs = project.source_imports.all()
        kpis_qs = project.kpis.all()
        module_rois_qs = project.module_rois.all()
        timesheets_qs = project.timesheet_entries.all()
        budget = getattr(project, "budgetestimatif", None)

        dependency_qs = (
            dm.TaskDependency.objects.filter(from_task__project=project)
            .select_related("from_task", "to_task", "created_by")
            .order_by("-created_at")
        )

        dependency_summary = {
            "count": dependency_qs.count(),
            "blocking": dependency_qs.filter(
                dependency_type=dm.TaskDependency.DependencyType.BLOCKS
            ).count(),
            "related": dependency_qs.filter(
                dependency_type=dm.TaskDependency.DependencyType.RELATES_TO
            ).count(),
        }

        task_stats = tasks_qs.aggregate(
            total=Count("id"),
            todo=Count("id", filter=Q(status=dm.Task.Status.TODO)),
            in_progress=Count("id", filter=Q(status=dm.Task.Status.IN_PROGRESS)),
            review=Count("id", filter=Q(status=dm.Task.Status.REVIEW)),
            done=Count("id", filter=Q(status=dm.Task.Status.DONE)),
            blocked=Count("id", filter=Q(status=dm.Task.Status.BLOCKED)),
            cancelled=Count("id", filter=Q(status=dm.Task.Status.CANCELLED)),
            overdue=Count(
                "id",
                filter=Q(due_date__lt=today) & ~Q(status__in=[dm.Task.Status.DONE, dm.Task.Status.CANCELLED]),
            ),
            flagged=Count("id", filter=Q(is_flagged=True)),
            total_estimate_hours=Coalesce(
                Sum("estimate_hours"),
                Value(Decimal("0.00")),
                output_field=hours_field,
            ),
            total_spent_hours=Coalesce(
                Sum("spent_hours"),
                Value(Decimal("0.00")),
                output_field=hours_field,
            ),
        )

        planning_stats = {
            "milestones_count": milestones_qs.count(),
            "releases_count": releases_qs.count(),
            "sprints_count": sprints_qs.count(),
            "backlog_count": backlog_qs.count(),
            "roadmap_items_count": roadmap_items_qs.count(),
            "planning_items_count": (
                milestones_qs.count()
                + releases_qs.count()
                + sprints_qs.count()
                + roadmap_items_qs.count()
            ),
        }

        sprint_summary = sprints_qs.annotate(
            tasks_count=Count("tasks", distinct=True),
            backlog_count=Count("backlog_items", distinct=True),
            metrics_count=Count("metrics", distinct=True),
        )

        team_summary = (
            members_qs.values("team__id", "team__name")
            .annotate(
                members_count=Count("id"),
                primary_count=Count("id", filter=Q(is_primary=True)),
            )
            .order_by("team__name")
        )

        evolution_summary = {
            "activities_count": activity_qs.count(),
            "pull_requests_count": prs_qs.count(),
            "risks_count": risks_qs.count(),
            "insights_count": ai_insights_qs.filter(is_dismissed=False).count(),
        }

        ai_summary = ai_insights_qs.aggregate(
            total=Count("id"),
            unread=Count("id", filter=Q(is_read=False)),
            dismissed=Count("id", filter=Q(is_dismissed=True)),
            critical=Count("id", filter=Q(severity=dm.AInsight.Severity.CRITICAL)),
            high=Count("id", filter=Q(severity=dm.AInsight.Severity.HIGH)),
        )

        timesheet_stats = timesheets_qs.aggregate(
            total_entries=Count("id"),
            total_hours=Coalesce(Sum("hours"), Value(Decimal("0.00")), output_field=hours_field),
            approved_hours=Coalesce(
                Sum("hours", filter=Q(approval_status=dm.TimesheetEntry.ApprovalStatus.APPROVED)),
                Value(Decimal("0.00")),
                output_field=hours_field,
            ),
            submitted_hours=Coalesce(
                Sum("hours", filter=Q(approval_status=dm.TimesheetEntry.ApprovalStatus.SUBMITTED)),
                Value(Decimal("0.00")),
                output_field=hours_field,
            ),
        )

        kpi_summary = {
            "count": kpis_qs.count(),
            "modules_count": kpis_qs.exclude(module_name="").values("module_name").distinct().count(),
        }

        roi_summary = module_rois_qs.aggregate(
            count=Count("id"),
            total_cost=Coalesce(Sum("estimated_cost"), Value(Decimal("0.00")), output_field=money_field),
            total_revenue=Coalesce(Sum("estimated_revenue"), Value(Decimal("0.00")), output_field=money_field),
        )
        roi_summary["global_roi_percent"] = (
            ((roi_summary["total_revenue"] - roi_summary["total_cost"]) / roi_summary["total_cost"]) * Decimal("100")
            if roi_summary["total_cost"] > 0
            else Decimal("0.00")
        )

        import_summary = {
            "count": imports_qs.count(),
            "completed": imports_qs.filter(status=dm.ProjectDocumentImport.ImportStatus.COMPLETED).count(),
            "processing": imports_qs.filter(status=dm.ProjectDocumentImport.ImportStatus.PROCESSING).count(),
            "failed": imports_qs.filter(status=dm.ProjectDocumentImport.ImportStatus.FAILED).count(),
        }

        member_cost_rows, member_cost_summary = self._build_member_cost_summary(members_qs, project)

        expenses_qs = dm.ProjectExpense.objects.none()
        estimate_lines_qs = dm.ProjectEstimateLine.objects.none()
        revenues_qs = dm.ProjectRevenue.objects.none()

        expense_stats = {
            "total_expenses": Decimal("0.00"),
            "estimated_expenses": Decimal("0.00"),
            "forecast_expenses": Decimal("0.00"),
            "committed_expenses": Decimal("0.00"),
            "accrued_expenses": Decimal("0.00"),
            "paid_expenses": Decimal("0.00"),
            "rejected_expenses": Decimal("0.00"),
            "labor_cost": Decimal("0.00"),
            "direct_cost": Decimal("0.00"),
            "other_cost": Decimal("0.00"),
            "validated_expenses": Decimal("0.00"),
            "draft_expenses": Decimal("0.00"),
        }

        estimate_stats = {
            "total_estimate_lines": 0,
            "total_estimated_cost": Decimal("0.00"),
            "total_estimated_sale": Decimal("0.00"),
            "total_estimated_margin": Decimal("0.00"),
            "estimated_cost": Decimal("0.00"),
            "baseline_cost": Decimal("0.00"),
            "forecast_cost": Decimal("0.00"),
            "raf_cost": Decimal("0.00"),
            "labor_cost": Decimal("0.00"),
            "direct_cost": Decimal("0.00"),
            "other_cost": Decimal("0.00"),
            "calculated_member_labor_cost": member_cost_summary["project_cost_total"],
            "calculated_member_labor_sale": member_cost_summary["project_sale_total"],
        }

        revenue_stats = {
            "total_revenue": Decimal("0.00"),
            "planned_revenue": Decimal("0.00"),
            "invoiced_revenue": Decimal("0.00"),
            "received_revenue": Decimal("0.00"),
            "pending_revenue": Decimal("0.00"),
            "remaining_to_invoice": Decimal("0.00"),
            "remaining_to_collect": Decimal("0.00"),
        }

        client_billing_summary = {
            "planned_revenue": Decimal("0.00"),
            "invoiced_revenue": Decimal("0.00"),
            "received_revenue": Decimal("0.00"),
            "remaining_to_invoice": Decimal("0.00"),
            "remaining_to_collect": Decimal("0.00"),
            "collection_rate_percent": 0,
            "invoicing_rate_percent": 0,
        }

        overview = {
            "approved_budget": Decimal("0.00"),
            "planned_revenue": Decimal("0.00"),
            "estimated_cost": Decimal("0.00"),
            "baseline_cost": Decimal("0.00"),
            "forecast_cost": Decimal("0.00"),
            "committed_cost": Decimal("0.00"),
            "accrued_cost": Decimal("0.00"),
            "actual_cost": Decimal("0.00"),
            "raf_cost": Decimal("0.00"),
            "forecast_final_cost": Decimal("0.00"),
            "received_revenue": Decimal("0.00"),
            "invoiced_revenue": Decimal("0.00"),
            "remaining_budget": Decimal("0.00"),
            "remaining_budget_forecast": Decimal("0.00"),
            "forecast_margin": Decimal("0.00"),
            "real_margin": Decimal("0.00"),
            "gross_margin": Decimal("0.00"),
            "operating_margin": Decimal("0.00"),
            "net_profit": Decimal("0.00"),
            "profit_margin_percent": 0,
            "expense_ratio_percent": 0,
            "forecast_consumption_percent": 0,
            "currency": getattr(budget, "currency", "XOF") if budget else "XOF",
        }

        budget_variance = Decimal("0.00")
        forecast_margin_amount = Decimal("0.00")
        budget_remaining = Decimal("0.00")

        sprint_financial_snapshots_qs = dm.SprintFinancialSnapshot.objects.filter(
            sprint__project=project
        ).select_related("sprint")

        feature_financial_snapshots_qs = dm.FeatureFinancialSnapshot.objects.filter(
            backlog_item__project=project
        ).select_related("backlog_item")

        sprint_financial_summary = sprint_financial_snapshots_qs.aggregate(
            count=Count("id"),
            total_estimated_cost=Coalesce(
                Sum("estimated_cost"),
                Value(Decimal("0.00")),
                output_field=money_field,
            ),
        )

        feature_financial_summary = feature_financial_snapshots_qs.aggregate(
            count=Count("id"),
            total_estimated_cost=Coalesce(
                Sum("estimated_cost"),
                Value(Decimal("0.00")),
                output_field=money_field,
            ),
            total_estimated_revenue=Coalesce(
                Sum("estimated_revenue"),
                Value(Decimal("0.00")),
                output_field=money_field,
            ),
        )
        feature_financial_summary["global_roi_percent"] = (
            (
                (feature_financial_summary["total_estimated_revenue"] - feature_financial_summary["total_estimated_cost"])
                / feature_financial_summary["total_estimated_cost"]
            ) * Decimal("100")
            if feature_financial_summary["total_estimated_cost"] > 0
            else Decimal("0.00")
        )

        if can_view_financials:
            expenses_qs = project.expenses.all()
            estimate_lines_qs = project.estimate_lines.all()
            revenues_qs = project.revenues.all()

            expense_stats = expenses_qs.aggregate(
                total_expenses=Coalesce(
                    Sum("amount", filter=~Q(status=dm.ProjectExpense.ExpenseStatus.REJECTED)),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
                estimated_expenses=Coalesce(
                    Sum("amount", filter=Q(status=dm.ProjectExpense.ExpenseStatus.ESTIMATED)),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
                forecast_expenses=Coalesce(
                    Sum(
                        "amount",
                        filter=Q(status__in=[
                            dm.ProjectExpense.ExpenseStatus.FORECAST,
                            dm.ProjectExpense.ExpenseStatus.DRAFT,
                        ]),
                    ),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
                committed_expenses=Coalesce(
                    Sum("amount", filter=Q(status=dm.ProjectExpense.ExpenseStatus.COMMITTED)),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
                accrued_expenses=Coalesce(
                    Sum("amount", filter=Q(status=dm.ProjectExpense.ExpenseStatus.ACCRUED)),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
                paid_expenses=Coalesce(
                    Sum("amount", filter=Q(status=dm.ProjectExpense.ExpenseStatus.PAID)),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
                rejected_expenses=Coalesce(
                    Sum("amount", filter=Q(status=dm.ProjectExpense.ExpenseStatus.REJECTED)),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
                validated_expenses=Coalesce(
                    Sum("amount", filter=Q(status=dm.ProjectExpense.ExpenseStatus.VALIDATED)),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
                draft_expenses=Coalesce(
                    Sum("amount", filter=Q(status=dm.ProjectExpense.ExpenseStatus.DRAFT)),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
            )

            labor_cost = Decimal("0.00")
            direct_cost = Decimal("0.00")
            other_cost = Decimal("0.00")

            for expense in expenses_qs.exclude(status=dm.ProjectExpense.ExpenseStatus.REJECTED):
                amount = expense.amount or Decimal("0.00")
                if expense.is_labor_cost:
                    labor_cost += amount
                elif expense.is_direct_cost:
                    direct_cost += amount
                else:
                    other_cost += amount

            expense_stats["labor_cost"] = labor_cost
            expense_stats["direct_cost"] = direct_cost
            expense_stats["other_cost"] = other_cost

            estimate_margin_expr = ExpressionWrapper(
                F("sale_amount") - F("cost_amount"),
                output_field=money_field,
            )

            estimate_agg = estimate_lines_qs.aggregate(
                total_estimate_lines=Count("id"),
                total_estimated_cost=Coalesce(
                    Sum("cost_amount"),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
                total_estimated_sale=Coalesce(
                    Sum("sale_amount"),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
                total_estimated_margin=Coalesce(
                    Sum(estimate_margin_expr),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
                estimated_cost=Coalesce(
                    Sum("cost_amount", filter=Q(budget_stage=dm.ProjectEstimateLine.BudgetStage.ESTIMATED)),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
                baseline_cost=Coalesce(
                    Sum("cost_amount", filter=Q(budget_stage=dm.ProjectEstimateLine.BudgetStage.BASELINE)),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
                forecast_cost=Coalesce(
                    Sum("cost_amount", filter=Q(budget_stage=dm.ProjectEstimateLine.BudgetStage.FORECAST)),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
                raf_cost=Coalesce(
                    Sum("cost_amount", filter=Q(budget_stage=dm.ProjectEstimateLine.BudgetStage.RAF)),
                    Value(Decimal("0.00")),
                    output_field=money_field,
                ),
            )
            estimate_stats = {
                **estimate_agg,
                "calculated_member_labor_cost": member_cost_summary["project_cost_total"],
                "calculated_member_labor_sale": member_cost_summary["project_sale_total"],
            }

            estimate_labor = Decimal("0.00")
            estimate_direct = Decimal("0.00")
            estimate_other = Decimal("0.00")

            for line in estimate_lines_qs:
                amount = line.cost_amount or Decimal("0.00")
                category = getattr(line, "category", None)
                if category and getattr(category, "is_labor_category", False):
                    estimate_labor += amount
                elif category and getattr(category, "is_direct_cost_category", False):
                    estimate_direct += amount
                else:
                    estimate_other += amount

            # Si aucune ligne RH n'existe, on injecte le coût RH calculé depuis l'équipe projet
            if estimate_labor <= 0 and member_cost_summary["project_cost_total"] > 0:
                estimate_labor = member_cost_summary["project_cost_total"]
                estimate_stats["total_estimated_cost"] += member_cost_summary["project_cost_total"]
                estimate_stats["total_estimated_sale"] += member_cost_summary["project_sale_total"]
                estimate_stats["total_estimated_margin"] += (
                    member_cost_summary["project_sale_total"] - member_cost_summary["project_cost_total"]
                )
                estimate_stats["estimated_cost"] += member_cost_summary["project_cost_total"]

            estimate_stats["labor_cost"] = estimate_labor
            estimate_stats["direct_cost"] = estimate_direct
            estimate_stats["other_cost"] = estimate_other

            revenue_stats = revenues_qs.aggregate(
                total_revenue=Coalesce(Sum("amount"), Value(Decimal("0.00")), output_field=money_field),
                planned_revenue=Coalesce(Sum("amount"), Value(Decimal("0.00")), output_field=money_field),
                invoiced_revenue=Coalesce(Sum("invoiced_amount"), Value(Decimal("0.00")), output_field=money_field),
                received_revenue=Coalesce(Sum("received_amount"), Value(Decimal("0.00")), output_field=money_field),
            )
            revenue_stats["pending_revenue"] = revenue_stats["planned_revenue"] - revenue_stats["received_revenue"]
            revenue_stats["remaining_to_invoice"] = max(
                revenue_stats["planned_revenue"] - revenue_stats["invoiced_revenue"],
                Decimal("0.00"),
            )
            revenue_stats["remaining_to_collect"] = max(
                revenue_stats["invoiced_revenue"] - revenue_stats["received_revenue"],
                Decimal("0.00"),
            )

            invoicing_rate_percent = 0
            collection_rate_percent = 0

            if revenue_stats["planned_revenue"] > 0:
                invoicing_rate_percent = int(
                    (revenue_stats["invoiced_revenue"] / revenue_stats["planned_revenue"]) * Decimal("100")
                )

            if revenue_stats["invoiced_revenue"] > 0:
                collection_rate_percent = int(
                    (revenue_stats["received_revenue"] / revenue_stats["invoiced_revenue"]) * Decimal("100")
                )

            client_billing_summary = {
                "planned_revenue": revenue_stats["planned_revenue"],
                "invoiced_revenue": revenue_stats["invoiced_revenue"],
                "received_revenue": revenue_stats["received_revenue"],
                "remaining_to_invoice": revenue_stats["remaining_to_invoice"],
                "remaining_to_collect": revenue_stats["remaining_to_collect"],
                "collection_rate_percent": collection_rate_percent,
                "invoicing_rate_percent": invoicing_rate_percent,
            }

            approved_budget = budget.approved_budget if budget else Decimal("0.00")
            planned_revenue = budget.expected_revenue_amount if budget else revenue_stats["planned_revenue"]

            estimated_cost = estimate_stats["total_estimated_cost"]
            baseline_cost = estimate_stats["baseline_cost"] or estimate_stats["total_estimated_cost"]
            forecast_cost = estimate_stats["forecast_cost"]
            committed_cost = expense_stats["committed_expenses"]
            accrued_cost = expense_stats["accrued_expenses"]
            actual_cost = expense_stats["paid_expenses"] + expense_stats["validated_expenses"]
            raf_cost = estimate_stats["raf_cost"]

            forecast_final_cost = actual_cost + committed_cost + accrued_cost + raf_cost

            remaining_budget = approved_budget - actual_cost
            remaining_budget_forecast = approved_budget - forecast_final_cost
            forecast_margin = planned_revenue - forecast_final_cost
            real_margin = revenue_stats["received_revenue"] - actual_cost
            gross_margin = revenue_stats["received_revenue"] - expense_stats["direct_cost"]
            operating_margin = gross_margin - expense_stats["labor_cost"]
            net_profit = operating_margin

            expense_ratio_percent = 0
            forecast_consumption_percent = 0
            profit_margin_percent = 0

            if approved_budget > 0:
                expense_ratio_percent = int((actual_cost / approved_budget) * Decimal("100"))
                forecast_consumption_percent = int((forecast_final_cost / approved_budget) * Decimal("100"))

            if revenue_stats["received_revenue"] > 0:
                profit_margin_percent = int((net_profit / revenue_stats["received_revenue"]) * Decimal("100"))

            overview = {
                "approved_budget": approved_budget,
                "planned_revenue": planned_revenue,
                "estimated_cost": estimated_cost,
                "baseline_cost": baseline_cost,
                "forecast_cost": forecast_cost,
                "committed_cost": committed_cost,
                "accrued_cost": accrued_cost,
                "actual_cost": actual_cost,
                "raf_cost": raf_cost,
                "forecast_final_cost": forecast_final_cost,
                "received_revenue": revenue_stats["received_revenue"],
                "invoiced_revenue": revenue_stats["invoiced_revenue"],
                "remaining_budget": remaining_budget,
                "remaining_budget_forecast": remaining_budget_forecast,
                "forecast_margin": forecast_margin,
                "real_margin": real_margin,
                "gross_margin": gross_margin,
                "operating_margin": operating_margin,
                "net_profit": net_profit,
                "profit_margin_percent": profit_margin_percent,
                "expense_ratio_percent": expense_ratio_percent,
                "forecast_consumption_percent": forecast_consumption_percent,
                "currency": getattr(budget, "currency", "XOF") if budget else "XOF",
            }

            budget_variance = approved_budget - actual_cost
            forecast_margin_amount = forecast_margin
            budget_remaining = remaining_budget

        quick_actions = {
            "planification": [
                {"label": "Nouveau sprint", "url": f"/sprints/create/?project={project.pk}", "style": "primary", "icon": "calendar"},
                {"label": "Nouveau jalon", "url": f"/milestones/create/?project={project.pk}", "style": "soft", "icon": "flag"},
                {"label": "Nouvelle release", "url": f"/releases/create/?project={project.pk}", "style": "soft", "icon": "rocket"},
                {"label": "Ajouter backlog item", "url": f"/backlog-items/create/?project={project.pk}", "style": "soft", "icon": "list"},
            ],
            "budget_estimatif": [],
            "budget_previsionnel": [],
            "facturation_client": [
                {"label": "Générer une facture", "url": reverse_lazy("invoice_generate_from_project", args=[project.pk]), "style": "primary", "icon": "sparkles"},
                {"label": "Facture manuelle", "url": f"{reverse_lazy('invoice_create')}?project={project.pk}", "style": "soft", "icon": "file"},
                {"label": "Voir toutes les factures", "url": f"{reverse_lazy('invoice_list')}?project={project.pk}", "style": "soft", "icon": "list"},
                {"label": "Ajouter prévision revenu", "url": f"/project-revenues/create/?project={project.pk}", "style": "soft", "icon": "trending-up"},
            ],
            "depenses": [],
            "equipes": [
                {"label": "Ajouter membre projet", "url": f"/project-members/create/?project={project.pk}", "style": "primary", "icon": "users"},
                {"label": "Associer une équipe", "url": f"/projects/{project.pk}/update/", "style": "soft", "icon": "layers"},
            ],
            "evolution": [
                {"label": "Ajouter activité", "url": f"/activity-logs/create/?project={project.pk}", "style": "primary", "icon": "pulse"},
                {"label": "Nouveau risque", "url": f"/risks/create/?project={project.pk}", "style": "soft", "icon": "alert"},
                {"label": "Nouvel insight IA", "url": f"/ai-insights/create/?project={project.pk}", "style": "soft", "icon": "sparkles"},
            ],
            "taches": [
                {"label": "Nouvelle tâche", "url": f"/tasks/create/?project={project.pk}", "style": "primary", "icon": "check-square"},
                {"label": "Voir board", "url": f"/boards/?project={project.pk}", "style": "soft", "icon": "columns"},
                {"label": "Voir toutes les tâches", "url": f"/tasks/?project={project.pk}", "style": "soft", "icon": "list"},
            ],
            "ia_import": [
                {"label": "Importer un document", "url": f"/project-imports/create/?project={project.pk}", "style": "primary", "icon": "upload"},
            ],
            "kpi_roi": [
                {"label": "Ajouter KPI", "url": f"/project-kpis/create/?project={project.pk}", "style": "primary", "icon": "chart-line"},
                {"label": "Ajouter ROI module", "url": f"/project-module-rois/create/?project={project.pk}", "style": "soft", "icon": "percent"},
            ],
        }

        if can_view_financials:
            quick_actions["budget_estimatif"] = [
                {"label": "Créer / éditer budget estimatif", "url": f"/project-budgets/create/?project={project.pk}", "style": "primary", "icon": "calculator"},
                {"label": "Ajouter ligne d'estimation", "url": f"/project-estimate-lines/create/?project={project.pk}", "style": "soft", "icon": "plus"},
            ]
            quick_actions["budget_previsionnel"] = [
                {"label": "Ajouter prévision de revenu", "url": f"/project-revenues/create/?project={project.pk}", "style": "primary", "icon": "banknote"},
                {"label": "Réviser budget", "url": f"/project-budgets/create/?project={project.pk}", "style": "soft", "icon": "refresh"},
            ]
            quick_actions["facturation_client"] = [
                {"label": "Ajouter revenu client", "url": f"/project-revenues/create/?project={project.pk}", "style": "primary", "icon": "banknote"},
                {"label": "Voir les revenus client", "url": f"/project-revenues/?project={project.pk}", "style": "soft", "icon": "wallet"},
            ]
            quick_actions["depenses"] = [
                {"label": "Nouvelle dépense", "url": f"/project-expenses/create/?project={project.pk}", "style": "primary", "icon": "receipt"},
                {"label": "Voir toutes les dépenses", "url": f"/project-expenses/?project={project.pk}", "style": "soft", "icon": "table"},
            ]

        tabs = [
            {"key": "planification", "label": "Plan", "count": planning_stats["planning_items_count"]},
            {"key": "equipes", "label": "Équipes", "count": members_qs.count()},
            {"key": "evolution", "label": "Évolution", "count": evolution_summary["activities_count"]},
            {"key": "taches", "label": "Tâches", "count": task_stats["total"]},
            {"key": "kpi_roi", "label": "KPI & ROI", "count": kpi_summary["count"] + roi_summary["count"]},
        ]

        if can_view_financials:
            tabs.insert(1, {"key": "budget_estimatif", "label": "Bu.estimatif", "count": estimate_stats["total_estimate_lines"]})
            tabs.insert(2, {"key": "budget_previsionnel", "label": "Bu.prévisionnel", "count": revenues_qs.count()})
            tabs.insert(3, {"key": "facturation_client", "label": "Facturation", "count": revenues_qs.count()})
            tabs.insert(4, {"key": "depenses", "label": "Dépenses", "count": expenses_qs.count()})

        active_tab = self.request.GET.get("tab", "planification")
        if active_tab in {"budget_estimatif", "budget_previsionnel", "facturation_client", "depenses"} and not can_view_financials:
            active_tab = "planification"

        valid_tab_keys = {tab["key"] for tab in tabs}
        if active_tab not in valid_tab_keys:
            active_tab = "planification"

        planning_ai_insights = ai_insights_qs.filter(
            insight_type__in=[
                dm.AInsight.InsightType.DELIVERY,
                dm.AInsight.InsightType.RISK,
                dm.AInsight.InsightType.WORKLOAD,
                dm.AInsight.InsightType.ALERT,
            ],
            is_dismissed=False,
        ).order_by("-detected_at")

        ctx.update({
            "project_obj": project,
            "workspace_obj": project.workspace,
            "planning_ai_insights": planning_ai_insights[:10],

            "members": members_qs,
            "sprints": sprint_summary,
            "all_sprints": sprints_qs,
            "tasks": tasks_qs[:50],
            "all_tasks": tasks_qs,
            "backlog_items": backlog_qs[:50],
            "all_backlog_items": backlog_qs,
            "pull_requests": prs_qs[:15],
            "all_pull_requests": prs_qs,
            "risks": risks_qs[:15],
            "all_risks": risks_qs,
            "milestones": milestones_qs,
            "releases": releases_qs,
            "roadmap_items": roadmap_items_qs,
            "board_columns": board_columns_qs,
            "labels": project.labels.all(),
            "activity_logs": activity_qs[:30],
            "all_activity_logs": activity_qs,
            "ai_insights": ai_insights_qs[:20],
            "all_ai_insights": ai_insights_qs,
            "timesheets": timesheets_qs[:50],
            "all_timesheets": timesheets_qs,

            "project_imports": imports_qs,
            "project_kpis": kpis_qs,
            "project_module_rois": module_rois_qs,
            "sprint_financial_snapshots": sprint_financial_snapshots_qs,
            "feature_financial_snapshots": feature_financial_snapshots_qs,

            "expenses": expenses_qs[:50] if can_view_financials else [],
            "all_expenses": expenses_qs if can_view_financials else [],
            "estimate_lines": estimate_lines_qs if can_view_financials else [],
            "revenues": revenues_qs if can_view_financials else [],
            "all_revenues": revenues_qs if can_view_financials else [],

            "budget_obj": budget if can_view_financials else None,
            "financial_overview": overview,
            "sprint_financial_summary": sprint_financial_summary,
            "feature_financial_summary": feature_financial_summary,

            "task_stats": task_stats,
            "planning_stats": planning_stats,
            "team_summary": team_summary,
            "evolution_summary": evolution_summary,
            "ai_summary": ai_summary,
            "timesheet_stats": timesheet_stats,
            "import_summary": import_summary,
            "kpi_summary": kpi_summary,
            "roi_summary": roi_summary,

            "expense_stats": expense_stats,
            "estimate_stats": estimate_stats,
            "revenue_stats": revenue_stats,
            "client_billing_summary": client_billing_summary,

            "member_cost_rows": member_cost_rows,
            "member_cost_summary": member_cost_summary,

            "overdue_tasks": task_stats["overdue"],
            "planning_items_count": planning_stats["planning_items_count"],
            "budget_variance": budget_variance,
            "forecast_margin_amount": forecast_margin_amount,
            "budget_remaining": budget_remaining,

            "tabs": tabs,
            "quick_actions": quick_actions,
            "active_tab": active_tab,
            "can_view_financials": can_view_financials,
            "task_dependencies": dependency_qs,
            "dependency_summary": dependency_summary,
        })

        # ─── Module Facturation : factures rattachées au projet ────────
        try:
            invoices_qs = (
                project.invoices.filter(is_archived=False)
                .select_related("client", "issued_by")
                .order_by("-issue_date")
            )
            invoice_agg = invoices_qs.aggregate(
                total_ttc=Sum("total_ttc"),
                total_paid=Sum("paid_amount"),
                count_total=Count("id"),
                count_overdue=Count("id", filter=Q(status=dm.Invoice.Status.OVERDUE)),
                count_draft=Count("id", filter=Q(status=dm.Invoice.Status.DRAFT)),
                count_paid=Count("id", filter=Q(status=dm.Invoice.Status.PAID)),
            )
            ctx["project_invoices"] = invoices_qs[:10]
            ctx["all_project_invoices"] = invoices_qs
            ctx["invoice_summary"] = {
                "total_ttc": invoice_agg["total_ttc"] or Decimal("0"),
                "total_paid": invoice_agg["total_paid"] or Decimal("0"),
                "total_due": (invoice_agg["total_ttc"] or Decimal("0"))
                              - (invoice_agg["total_paid"] or Decimal("0")),
                "count_total": invoice_agg["count_total"] or 0,
                "count_overdue": invoice_agg["count_overdue"] or 0,
                "count_draft": invoice_agg["count_draft"] or 0,
                "count_paid": invoice_agg["count_paid"] or 0,
            }
        except Exception:
            ctx["project_invoices"] = []
            ctx["all_project_invoices"] = []
            ctx["invoice_summary"] = {
                "total_ttc": Decimal("0"), "total_paid": Decimal("0"),
                "total_due": Decimal("0"), "count_total": 0,
                "count_overdue": 0, "count_draft": 0, "count_paid": 0,
            }

        ctx["active_quick_actions"] = quick_actions.get(active_tab, [])
        return ctx

class ProjectCreateView(DevflowCreateView):
    model = dm.Project
    form_class = ProjectForm
    template_name = "project/create.html"
    section = "project"
    page_title = "Créer projet"
    success_list_url_name = "project_list"
    success_message = "Projet créé avec succès."

    def get_initial(self):
        initial = super().get_initial()
        initial.setdefault("status", dm.Project.Status.PLANNED)
        initial.setdefault("priority", dm.Project.Priority.MEDIUM)
        initial.setdefault("health_status", dm.Project.HealthStatus.GRAY)
        initial.setdefault("progress_percent", 0)
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({
            "section": self.section,
            "page_title": self.page_title,
            "form_title": "Créer un nouveau projet",
            "form_subtitle": "Renseignez les informations de cadrage, de pilotage et de planification du projet.",
            "submit_label": "Enregistrer le projet",
            "cancel_url": reverse_lazy(self.success_list_url_name),
        })
        return ctx

    def get_workspace(self):
        workspace = getattr(self.request, "workspace", None)

        if not workspace:
            workspace = getattr(getattr(self.request.user, "profile", None), "workspace", None)

        return workspace

    def form_valid(self, form):
        workspace = form.cleaned_data.get("workspace") or self.get_workspace()

        if not workspace:
            form.add_error("workspace", "Veuillez sélectionner un workspace.")
            form.add_error(None, "Aucun workspace actif n'est associé à cette création de projet.")
            return self.form_invalid(form)

        try:
            with transaction.atomic():
                obj = form.save(commit=False)

                obj.workspace = workspace

                if not getattr(obj, "owner_id", None) and self.request.user.is_authenticated:
                    obj.owner = self.request.user

                if obj.progress_percent is None:
                    obj.progress_percent = 0

                if hasattr(obj, "risk_score") and obj.risk_score is None:
                    obj.risk_score = 0

                obj.save()
                form.save_m2m()

                if hasattr(dm, "ProjectBudget"):
                    dm.ProjectBudget.objects.get_or_create(
                        project=obj,
                        defaults={
                            "status": dm.ProjectBudget.Status.DRAFT,
                            "currency": "XOF",
                        },
                    )

                self.object = obj

        except ValidationError as e:
            form.add_error(None, "; ".join(e.messages))
            return self.form_invalid(form)

        except Exception as e:
            form.add_error(None, f"Erreur lors de l'enregistrement : {e}")
            return self.form_invalid(form)

        messages.success(self.request, self.success_message)
        return redirect(reverse_lazy(self.success_list_url_name))



class ProjectDocumentImportListView(DevflowListView):
    model = dm.ProjectDocumentImport
    template_name = "project/document_import/list.html"
    context_object_name = "imports"
    paginate_by = 20
    section = "project"
    page_title = "Documents projet"

    def get_queryset(self):
        qs = (
            dm.ProjectDocumentImport.objects.select_related(
                "workspace",
                "uploaded_by",
                "project",
                "project__team",
                "project__owner",
                "project__product_manager",
                "project__category",
            )
            .order_by("-created_at", "-id")
        )

        workspace_id = self.request.GET.get("workspace")
        project_id = self.request.GET.get("project")
        status = self.request.GET.get("status")
        q = (self.request.GET.get("q") or "").strip()

        if workspace_id:
            qs = qs.filter(workspace_id=workspace_id)

        if project_id:
            qs = qs.filter(project_id=project_id)

        if status:
            qs = qs.filter(status=status)

        if q:
            qs = qs.filter(
                Q(file__icontains=q)
                | Q(extracted_text__icontains=q)
                | Q(error_message__icontains=q)
                | Q(project__name__icontains=q)
                | Q(project__code__icontains=q)
                | Q(workspace__name__icontains=q)
                | Q(uploaded_by__username__icontains=q)
                | Q(uploaded_by__first_name__icontains=q)
                | Q(uploaded_by__last_name__icontains=q)
            )

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = self.object_list

        workspace_id = self.request.GET.get("workspace")
        project_id = self.request.GET.get("project")
        status = self.request.GET.get("status")
        q = (self.request.GET.get("q") or "").strip()

        stats = qs.aggregate(
            total=Count("id"),
            uploaded=Count("id", filter=Q(status=dm.ProjectDocumentImport.ImportStatus.UPLOADED)),
            processing=Count("id", filter=Q(status=dm.ProjectDocumentImport.ImportStatus.PROCESSING)),
            completed=Count("id", filter=Q(status=dm.ProjectDocumentImport.ImportStatus.COMPLETED)),
            failed=Count("id", filter=Q(status=dm.ProjectDocumentImport.ImportStatus.FAILED)),
        )

        ctx.update(
            {
                "stats": stats,
                "status_choices": dm.ProjectDocumentImport.ImportStatus.choices,
                "workspace_list": dm.Workspace.objects.filter(is_archived=False).order_by("name"),
                "project_list": (
                    dm.Project.objects.filter(is_archived=False)
                    .select_related("workspace")
                    .order_by("name")
                ),
                "filters": {
                    "workspace": workspace_id or "",
                    "project": project_id or "",
                    "status": status or "",
                    "q": q,
                },
            }
        )
        return ctx


class ProjectDocumentImportDetailView(DevflowDetailView):
    model = dm.ProjectDocumentImport
    template_name = "project/document_import/detail.html"
    context_object_name = "import_obj"
    section = "project"
    page_title = "Détail document projet"

    def get_queryset(self):
        return (
            dm.ProjectDocumentImport.objects.select_related(
                "workspace",
                "uploaded_by",
                "project",
                "project__workspace",
                "project__team",
                "project__owner",
                "project__product_manager",
                "project__category",
            )
            .order_by("-created_at", "-id")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        import_obj = self.object

        file_name = ""
        file_url = ""
        file_ext = ""

        if import_obj.file:
            file_name = import_obj.file.name.split("/")[-1]
            file_url = import_obj.file.url
            if "." in file_name:
                file_ext = file_name.rsplit(".", 1)[-1].lower()

        related_imports = (
            dm.ProjectDocumentImport.objects.filter(project=import_obj.project)
            .exclude(pk=import_obj.pk)
            .select_related("uploaded_by", "project")
            .order_by("-created_at")[:10]
            if import_obj.project_id
            else dm.ProjectDocumentImport.objects.none()
        )

        ctx.update(
            {
                "file_name": file_name,
                "file_url": file_url,
                "file_ext": file_ext,
                "related_imports": related_imports,
                "has_extracted_text": bool(import_obj.extracted_text),
                "has_ai_payload": bool(import_obj.ai_payload),
                "has_error": bool(import_obj.error_message),
            }
        )
        return ctx


class ProjectDocumentImportCreateView(DevflowCreateView):
    model = dm.ProjectDocumentImport
    form_class = ProjectDocumentImportForm
    template_name = "project/document_import/form.html"
    section = "project"
    page_title = "Importer un document projet"

    def get_project(self):
        project_id = self.request.GET.get("project") or self.request.POST.get("project")
        if not project_id:
            return None
        return get_object_or_404(
            dm.Project.objects.select_related("workspace"),
            pk=project_id,
            is_archived=False,
        )

    def get_workspace(self):
        project = self.get_project()
        if project:
            return project.workspace

        workspace_id = self.request.GET.get("workspace") or self.request.POST.get("workspace")
        if workspace_id:
            return get_object_or_404(dm.Workspace, pk=workspace_id, is_archived=False)

        return None

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["workspace"] = self.get_workspace()
        kwargs["project"] = self.get_project()
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        project = self.get_project()

        if project:
            initial["project"] = project
        initial.setdefault("status", dm.ProjectDocumentImport.ImportStatus.UPLOADED)
        return initial

    def form_valid(self, form):
        project = form.cleaned_data.get("project") or self.get_project()
        workspace = self.get_workspace()

        if project and workspace is None:
            workspace = project.workspace

        if workspace is None:
            form.add_error("project", "Le projet ou le workspace est requis pour l'import.")
            return self.form_invalid(form)

        obj = form.save(commit=False)
        obj.workspace = workspace
        obj.project = project
        obj.uploaded_by = self.request.user if self.request.user.is_authenticated else None

        if not obj.status:
            obj.status = dm.ProjectDocumentImport.ImportStatus.UPLOADED

        obj.save()
        self.object = obj

        messages.success(self.request, "Le document a été importé avec succès.")
        return super().form_valid(form)

    def get_success_url(self):
        if self.object.project_id:
            return reverse("project_document_import_detail", kwargs={"pk": self.object.pk})
        return reverse("project_document_import_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(
            {
                "project_obj": self.get_project(),
                "workspace_obj": self.get_workspace(),
            }
        )
        return ctx
# class ProjectUpdateView(DevflowUpdateView):
#     model = dm.Project
#     form_class = ProjectForm
#     section = "project"
#     page_title = "Modifier projet"
#     success_list_url_name = "project_list"
class ProjectUpdateView(DevflowUpdateView):
    model = dm.Project
    form_class = ProjectForm
    template_name = "project/update.html"
    section = "project"
    page_title = "Modifier projet"
    success_list_url_name = "project_list"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({
            "section": self.section,
            "page_title": self.page_title,
            "form_title": f"Modifier le projet : {self.object.name}",
            "form_subtitle": "Mettez à jour les informations de cadrage, de pilotage et de planification du projet.",
            "submit_label": "Enregistrer les modifications",
            "cancel_url": reverse_lazy("project_detail", kwargs={"pk": self.object.pk}),
            "item": self.object,
        })
        return ctx


class ProjectDeleteView(DevflowDeleteView):
    model = dm.Project
    section = "project"
    page_title = "Supprimer projet"
    success_list_url_name = "project_list"


class ProjectArchiveView(ArchiveObjectView):
    model = dm.Project
    success_list_url_name = "project_list"


class ProjectMemberListView(DevflowListView):
    model = dm.ProjectMember
    template_name = "project/project_member/list.html"
    section = "project"
    page_title = "Membres projet"
    search_placeholder = "Rechercher par nom, email, rôle ou projet…"
    search_fields = (
        "role", "project__name",
        "user__username", "user__first_name", "user__last_name", "user__email",
    )
    filter_fields = [
        {"name": "project", "label": "Projet", "type": "model",
         "queryset": "_projects_qs", "lookup": "project_id"},
        {"name": "team", "label": "Équipe", "type": "model",
         "queryset": "_teams_qs", "lookup": "team_id"},
        {"name": "is_primary", "label": "Primaire", "type": "bool"},
    ]

    def _projects_qs(self):
        ws = self.get_current_workspace()
        return dm.Project.objects.filter(workspace=ws, is_archived=False).order_by("name") if ws else dm.Project.objects.none()

    def _teams_qs(self):
        ws = self.get_current_workspace()
        return dm.Team.objects.filter(workspace=ws, is_archived=False).order_by("name") if ws else dm.Team.objects.none()


class ProjectMemberDetailView(DevflowDetailView):
    model = dm.ProjectMember
    template_name = "project/project_member/detail.html"
    section = "project"
    page_title = "Détail membre projet"


class ProjectMemberCreateView(DevflowCreateView):
    model = dm.ProjectMember
    form_class = ProjectMemberForm
    template_name = "project/project_member/form.html"
    section = "project"
    page_title = "Ajouter membre projet"
    success_list_url_name = "project_member_list"

    def get_initial(self):
        initial = super().get_initial()
        project_id = self.request.GET.get("project")
        if project_id:
            initial["project"] = project_id
        return initial


class ProjectMemberUpdateView(DevflowUpdateView):
    model = dm.ProjectMember
    form_class = ProjectMemberForm
    template_name = "project/project_member/form.html"
    section = "project"
    page_title = "Modifier membre projet"
    success_list_url_name = "project_member_list"


class ProjectMemberDeleteView(DevflowDeleteView):
    model = dm.ProjectMember
    section = "project"
    page_title = "Supprimer membre projet"
    success_list_url_name = "project_member_list"


# =============================================================================
# SPRINTS
# =============================================================================
@login_required
@require_POST
def sprint_status_update(request):
    try:
        payload = json.loads(request.body.decode("utf-8"))
        sprint_id = payload.get("sprint_id")
        status = payload.get("status")

        allowed_statuses = {
            dm.Sprint.Status.PLANNED,
            dm.Sprint.Status.ACTIVE,
            dm.Sprint.Status.REVIEW,
            dm.Sprint.Status.DONE,
            dm.Sprint.Status.CANCELLED,
        }

        if status not in allowed_statuses:
            return JsonResponse(
                {"success": False, "message": "Statut invalide."},
                status=400
            )

        sprint = get_object_or_404(dm.Sprint, pk=sprint_id)
        sprint.status = status
        sprint.save(update_fields=["status", "updated_at"])

        return JsonResponse({
            "success": True,
            "sprint_id": sprint.pk,
            "status": sprint.status,
            "status_label": sprint.get_status_display(),
        })
    except Exception as exc:
        return JsonResponse(
            {"success": False, "message": str(exc)},
            status=400
        )


@login_required
@require_POST
def task_status_update(request):
    try:
        payload = json.loads(request.body.decode("utf-8"))
        task_id = payload.get("task_id")
        status = payload.get("status")

        allowed_statuses = {
            dm.Task.Status.TODO,
            dm.Task.Status.IN_PROGRESS,
            dm.Task.Status.REVIEW,
            dm.Task.Status.DONE,
            dm.Task.Status.BLOCKED,
            dm.Task.Status.EXPIRED,
        }

        if status not in allowed_statuses:
            return JsonResponse({"success": False, "message": "Statut invalide."}, status=400)

        task = get_object_or_404(dm.Task, pk=task_id)
        task.status = status
        task.save(update_fields=["status", "updated_at"])

        return JsonResponse({"success": True, "status": task.status})
    except Exception as exc:
        return JsonResponse({"success": False, "message": str(exc)}, status=400)


class SprintListView(DevflowListView):
    model = dm.Sprint
    template_name = "project/sprint/list.html"
    section = "sprint"
    page_title = "Sprints"
    search_placeholder = "Rechercher par nom, goal, projet ou équipe…"
    search_fields = ("name", "goal", "status", "project__name", "team__name")
    filter_fields = [
        {"name": "status", "label": "Statut", "type": "choices",
         "choices": dm.Sprint.Status.choices},
        {"name": "project", "label": "Projet", "type": "model",
         "queryset": "_projects_qs", "lookup": "project_id"},
        {"name": "team", "label": "Équipe", "type": "model",
         "queryset": "_teams_qs", "lookup": "team_id"},
    ]

    def _projects_qs(self):
        ws = self.get_current_workspace()
        return dm.Project.objects.filter(workspace=ws, is_archived=False).order_by("name") if ws else dm.Project.objects.none()

    def _teams_qs(self):
        ws = self.get_current_workspace()
        return dm.Team.objects.filter(workspace=ws, is_archived=False).order_by("name") if ws else dm.Team.objects.none()

    def get_queryset(self):
        open_tasks_preview_qs = (
            dm.Task.objects.filter(is_archived=False)
            .exclude(status=dm.Task.Status.DONE)
            .select_related("assignee", "reporter", "backlog_item")
            .order_by("due_date", "priority", "created_at")
        )

        return (
            dm.Sprint.objects
            .select_related("workspace", "project", "team")
            .prefetch_related(
                Prefetch("tasks", queryset=open_tasks_preview_qs, to_attr="open_tasks_preview")
            )
            .annotate(
                tasks_count=Count("tasks", filter=Q(tasks__is_archived=False), distinct=True),
                backlog_count=Count("backlog_items", filter=Q(backlog_items__is_archived=False), distinct=True),

                todo_tasks_count=Count(
                    "tasks",
                    filter=Q(tasks__is_archived=False, tasks__status=dm.Task.Status.TODO),
                    distinct=True,
                ),
                in_progress_tasks_count=Count(
                    "tasks",
                    filter=Q(tasks__is_archived=False, tasks__status=dm.Task.Status.IN_PROGRESS),
                    distinct=True,
                ),
                review_tasks_count=Count(
                    "tasks",
                    filter=Q(tasks__is_archived=False, tasks__status=dm.Task.Status.REVIEW),
                    distinct=True,
                ),
                blocked_tasks_count=Count(
                    "tasks",
                    filter=Q(tasks__is_archived=False, tasks__status=dm.Task.Status.BLOCKED),
                    distinct=True,
                ),
                done_tasks_count=Count(
                    "tasks",
                    filter=Q(tasks__is_archived=False, tasks__status=dm.Task.Status.DONE),
                    distinct=True,
                ),
            )
            .order_by("-start_date", "-id")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = ctx["object_list"]
        today = timezone.localdate()

        ctx.update({
            "stats": qs.aggregate(
                total=Count("id"),
                planned=Count("id", filter=Q(status=dm.Sprint.Status.PLANNED)),
                active=Count("id", filter=Q(status=dm.Sprint.Status.ACTIVE)),
                review=Count("id", filter=Q(status=dm.Sprint.Status.REVIEW)),
                done=Count("id", filter=Q(status=dm.Sprint.Status.DONE)),
                cancelled=Count("id", filter=Q(status=dm.Sprint.Status.CANCELLED)),
                total_velocity_target=Coalesce(Sum("velocity_target"), 0),
                total_velocity_completed=Coalesce(Sum("velocity_completed"), 0),
            ),
            "today": today,
        })
        return ctx


class SprintDetailView(DevflowDetailView):
    model = dm.Sprint
    template_name = "project/sprint/detail.html"
    section = "sprint"
    page_title = "Détail sprint"

    def get_queryset(self):
        return (
            dm.Sprint.objects.select_related("project", "team", "workspace")
            .prefetch_related("metrics", "backlog_items", "ai_insights")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        sprint = self.object

        sprint_tasks_qs = (
            sprint.tasks.filter(is_archived=False)
            .select_related("assignee", "reporter", "backlog_item", "sprint")
            .order_by("-created_at")
        )

        sprint_backlog_qs = (
            sprint.backlog_items.filter(is_archived=False)
            .select_related("parent", "reporter")
            .order_by("rank", "title")
        )

        sprint_task_stats = sprint_tasks_qs.aggregate(
            total=Count("id"),
            todo=Count("id", filter=Q(status=dm.Task.Status.TODO)),
            in_progress=Count("id", filter=Q(status=dm.Task.Status.IN_PROGRESS)),
            review=Count("id", filter=Q(status=dm.Task.Status.REVIEW)),
            done=Count("id", filter=Q(status=dm.Task.Status.DONE)),
            blocked=Count("id", filter=Q(status=dm.Task.Status.BLOCKED)),
        )

        today = timezone.localdate()
        sprint_overdue_tasks = (
            sprint_tasks_qs.filter(due_date__lt=today)
            .exclude(status=dm.Task.Status.DONE)
            .count()
        )

        ctx.update({
            "sprint_obj": sprint,
            "sprint_metrics": sprint.metrics.all(),
            "sprint_tasks": sprint_tasks_qs,
            "sprint_backlog_items": sprint_backlog_qs,
            "sprint_task_stats": sprint_task_stats,
            "sprint_overdue_tasks": sprint_overdue_tasks,
            "review_obj": getattr(sprint, "review", None),
            "retrospective_obj": getattr(sprint, "retrospective", None),
            "sprint_insights": sprint.ai_insights.filter(is_dismissed=False),
        })
        return ctx


class SprintCreateView(DevflowCreateView):
    model = dm.Sprint
    form_class = SprintForm
    template_name = "project/sprint/form.html"
    section = "sprint"
    page_title = "Créer sprint"
    success_list_url_name = "sprint_list"

    def get_initial(self):
        initial = super().get_initial()
        project_id = self.request.GET.get("project")
        if project_id:
            initial["project"] = project_id
        return initial


class SprintUpdateView(DevflowUpdateView):
    model = dm.Sprint
    form_class = SprintForm
    template_name = "project/sprint/form.html"
    section = "sprint"
    page_title = "Modifier sprint"
    success_list_url_name = "sprint_list"


class SprintDeleteView(DevflowDeleteView):
    model = dm.Sprint
    section = "sprint"
    page_title = "Supprimer sprint"
    success_list_url_name = "sprint_list"


class SprintArchiveView(ArchiveObjectView):
    model = dm.Sprint
    success_list_url_name = "sprint_list"


class SprintMetricListView(DevflowListView):
    model = dm.SprintMetric
    template_name = "project/sprint_metric/list.html"
    section = "sprint"
    page_title = "Métriques sprint"
    search_fields = ()


class SprintMetricDetailView(DevflowDetailView):
    model = dm.SprintMetric
    template_name = "project/sprint_metric/detail.html"
    section = "sprint"
    page_title = "Détail métrique sprint"


class SprintMetricCreateView(DevflowCreateView):
    model = dm.SprintMetric
    form_class = SprintMetricForm
    section = "sprint"
    page_title = "Créer métrique sprint"
    success_list_url_name = "sprint_metric_list"


class SprintMetricUpdateView(DevflowUpdateView):
    model = dm.SprintMetric
    form_class = SprintMetricForm
    section = "sprint"
    page_title = "Modifier métrique sprint"
    success_list_url_name = "sprint_metric_list"


class SprintMetricDeleteView(DevflowDeleteView):
    model = dm.SprintMetric
    section = "sprint"
    page_title = "Supprimer métrique sprint"
    success_list_url_name = "sprint_metric_list"


# =============================================================================
# BACKLOG / TASKS
# =============================================================================
class BacklogItemListView(DevflowListView):
    model = dm.BacklogItem
    template_name = "project/backlog_item/list.html"
    section = "backlog"
    page_title = "Backlog"
    search_fields = ("title", "description", "acceptance_criteria", "item_type")


class BacklogItemDetailView(DevflowDetailView):
    model = dm.BacklogItem
    template_name = "project/backlog_item/detail.html"
    section = "backlog"
    page_title = "Détail backlog"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        item = self.object
        ctx["children"] = item.children.filter(is_archived=False)
        ctx["tasks"] = item.tasks.filter(is_archived=False)
        return ctx


class BacklogItemCreateView(DevflowCreateView):
    model = dm.BacklogItem
    form_class = BacklogItemForm
    section = "backlog"
    page_title = "Créer backlog item"
    success_list_url_name = "backlog_item_list"


class BacklogItemUpdateView(DevflowUpdateView):
    model = dm.BacklogItem
    form_class = BacklogItemForm
    section = "backlog"
    page_title = "Modifier backlog item"
    success_list_url_name = "backlog_item_list"


class BacklogItemDeleteView(DevflowDeleteView):
    model = dm.BacklogItem
    section = "backlog"
    page_title = "Supprimer backlog item"
    success_list_url_name = "backlog_item_list"


class BacklogItemArchiveView(ArchiveObjectView):
    model = dm.BacklogItem
    success_list_url_name = "backlog_item_list"


class TaskQuickAssignView(DevflowBaseMixin, View):
    """
    Assignation/désassignation rapide d'une tâche.
    Accepte indifféremment 'assignee' ou 'user' dans le POST pour rester
    rétro-compatible avec les anciens templates.
    Utilise Task.assign() pour garantir la cohérence FK + TaskAssignment.
    DevflowBaseMixin hérite déjà de LoginRequiredMixin.
    """
    def post(self, request, pk):
        task = self.filter_by_workspace(
            dm.Task.objects.all()
        ).select_related("project", "workspace").get(pk=pk)

        assignee_id = request.POST.get("assignee") or request.POST.get("user")

        if assignee_id:
            assignee = get_object_or_404(User, pk=assignee_id, is_active=True)
            task.assign(assignee, assigned_by=request.user)
            messages.success(request, f"Tâche affectée à {assignee}.")
        else:
            task.unassign(actor=request.user)
            messages.success(request, "Assignation supprimée.")

        next_url = request.POST.get("next")
        return redirect(next_url or request.META.get("HTTP_REFERER") or "task_list")


class TaskQuickStatusView(DevflowBaseMixin, View):
    def post(self, request, pk):
        task = self.filter_by_workspace(dm.Task.objects.all()).select_related("project", "workspace").get(pk=pk)
        status = request.POST.get("status")

        allowed_statuses = {choice[0] for choice in dm.Task.Status.choices}
        if status in allowed_statuses:
            task.status = status

            if status == dm.Task.Status.IN_PROGRESS and not task.started_at:
                task.started_at = timezone.now()

            if status == dm.Task.Status.DONE and not task.completed_at:
                task.completed_at = timezone.now()

            if status != dm.Task.Status.DONE:
                task.completed_at = None

            task.save(update_fields=["status", "started_at", "completed_at", "updated_at"])

            dm.ActivityLog.objects.create(
                workspace=task.workspace,
                actor=request.user,
                project=task.project,
                task=task,
                activity_type=dm.ActivityLog.ActivityType.TASK_MOVED,
                title=f"Statut changé vers {task.get_status_display()}",
                description=f"{request.user} a mis à jour le statut de « {task.title} ».",
            )

            messages.success(request, "Statut mis à jour.")

        next_url = request.POST.get("next")
        return redirect(next_url or "task_list")


class TaskQuickCommentView(DevflowBaseMixin, View):
    def post(self, request, pk):
        task = self.filter_by_workspace(dm.Task.objects.all()).select_related("project", "workspace").get(pk=pk)
        body = (request.POST.get("body") or "").strip()
        is_internal = request.POST.get("is_internal") == "1"

        if body:
            dm.TaskComment.objects.create(
                task=task,
                author=request.user,
                body=body,
                is_internal=is_internal,
            )

            dm.Task.objects.filter(pk=task.pk).update(comments_count=(task.comments_count or 0) + 1)

            dm.ActivityLog.objects.create(
                workspace=task.workspace,
                actor=request.user,
                project=task.project,
                task=task,
                activity_type=dm.ActivityLog.ActivityType.COMMENT_ADDED,
                title="Commentaire ajouté",
                description=f"{request.user} a commenté la tâche « {task.title} ».",
            )

            messages.success(request, "Commentaire ajouté.")
        else:
            messages.warning(request, "Le commentaire est vide.")

        next_url = request.POST.get("next")
        return redirect(next_url or "task_detail")


class TaskToggleFlagView(DevflowBaseMixin, View):
    def post(self, request, pk):
        task = self.filter_by_workspace(dm.Task.objects.all()).select_related("project", "workspace").get(pk=pk)
        task.is_flagged = not task.is_flagged
        task.save(update_fields=["is_flagged", "updated_at"])

        messages.success(
            request,
            "Tâche marquée comme sensible." if task.is_flagged else "Marquage retiré."
        )

        next_url = request.POST.get("next")
        return redirect(next_url or "task_list")

# ──────────────────────────────────────────────────────────────────────────
# NOTE: Les classes TaskQuickAssignView et TaskQuickCommentView en doublon
# ont été supprimées (audit 2026-04-29). Les versions canoniques utilisant
# DevflowBaseMixin (filter_by_workspace + ActivityLog) sont conservées plus
# haut dans ce fichier.
# ──────────────────────────────────────────────────────────────────────────


class TaskQuickAttachmentView(LoginRequiredMixin, View):
    def post(self, request, pk):
        task = get_object_or_404(dm.Task, pk=pk, is_archived=False)
        uploaded = request.FILES.get("file")

        if not uploaded:
            messages.error(request, "Aucun fichier sélectionné.")
            return redirect(request.META.get("HTTP_REFERER", "task_list"))

        dm.TaskAttachment.objects.create(
            task=task,
            uploaded_by=request.user,
            file=uploaded,
            name=uploaded.name,
            mime_type=getattr(uploaded, "content_type", "") or "",
            size=getattr(uploaded, "size", 0) or 0,
        )

        task.attachments_count = task.attachments.count()
        task.save(update_fields=["attachments_count", "updated_at"])

        messages.success(request, "Pièce jointe ajoutée.")
        return redirect(request.META.get("HTTP_REFERER", "task_list"))
class TaskKanbanMoveView(LoginRequiredMixin, View):
    def post(self, request, pk):
        task = get_object_or_404(dm.Task, pk=pk, is_archived=False)
        new_status = request.POST.get("status")
        new_position = request.POST.get("position")

        allowed_statuses = {choice[0] for choice in dm.Task.Status.choices}
        if new_status not in allowed_statuses:
            return JsonResponse({"ok": False, "error": "Statut invalide."}, status=400)

        try:
            new_position = int(new_position or 0)
        except (TypeError, ValueError):
            new_position = 0

        task.status = new_status
        task.position = max(new_position, 0)

        if new_status == dm.Task.Status.IN_PROGRESS and not task.started_at:
            task.started_at = timezone.now()

        if new_status == dm.Task.Status.DONE and not task.completed_at:
            task.completed_at = timezone.now()

        task.save()

        return JsonResponse({
            "ok": True,
            "task_id": task.pk,
            "status": task.status,
            "position": task.position,
        })
class TaskListView(DevflowListView):
    model = dm.Task
    template_name = "project/task/list.html"
    section = "task"
    page_title = "Tâches"
    search_placeholder = "Rechercher par titre, description ou projet…"
    search_fields = (
        "title", "description", "status", "priority",
        "project__name", "sprint__name",
        "assignee__username", "assignee__first_name", "assignee__last_name",
        "reporter__username",
    )
    filter_fields = [
        {"name": "status", "label": "Statut", "type": "choices",
         "choices": dm.Task.Status.choices},
        {"name": "priority", "label": "Priorité", "type": "choices",
         "choices": dm.Task.Priority.choices},
        {"name": "project", "label": "Projet", "type": "model",
         "queryset": "_projects_qs", "lookup": "project_id"},
        {"name": "assignee", "label": "Assigné à", "type": "model",
         "queryset": "_users_qs", "lookup": "assignee_id"},
    ]

    def _projects_qs(self):
        ws = self.get_current_workspace()
        return dm.Project.objects.filter(workspace=ws, is_archived=False).order_by("name") if ws else dm.Project.objects.none()

    def _users_qs(self):
        ws = self.get_current_workspace()
        if not ws:
            return User.objects.none()
        return User.objects.filter(
            is_active=True, devflow_memberships__workspace=ws
        ).distinct().order_by("last_name", "first_name", "username")

    def get_queryset(self):
        queryset = (
            super()
            .get_queryset()
            .select_related(
                "workspace",
                "project",
                "sprint",
                "assignee",
                "reporter",
                "parent",
                "backlog_item",
            )
            .prefetch_related(
                Prefetch(
                    "assignments",
                    queryset=dm.TaskAssignment.objects.select_related("user", "assigned_by").order_by("-is_active", "user__username"),
                ),
                Prefetch(
                    "comments",
                    queryset=dm.TaskComment.objects.select_related("author").order_by("-created_at"),
                ),
                Prefetch(
                    "attachments",
                    queryset=dm.TaskAttachment.objects.select_related("uploaded_by").order_by("-created_at"),
                ),
                Prefetch(
                    "labels",
                    queryset=dm.TaskLabel.objects.select_related("label").order_by("label__name"),
                ),
                Prefetch(
                    "subtasks",
                    queryset=dm.Task.objects.filter(is_archived=False).select_related("assignee").order_by("position", "-created_at"),
                ),
                Prefetch(
                    "incoming_dependencies",
                    queryset=dm.TaskDependency.objects.select_related("from_task").order_by("-created_at"),
                ),
                Prefetch(
                    "outgoing_dependencies",
                    queryset=dm.TaskDependency.objects.select_related("to_task").order_by("-created_at"),
                ),
            )
            .filter(is_archived=False)
            .annotate(
                subtasks_count=Count("subtasks", filter=Q(subtasks__is_archived=False), distinct=True),
                active_dependencies_count=Count("incoming_dependencies", distinct=True),
                comments_total=Count("comments", distinct=True),
                attachments_total=Count("attachments", distinct=True),
            )
        )

        status_filter = self.request.GET.get("status")
        priority_filter = self.request.GET.get("priority")
        project_filter = self.request.GET.get("project")
        sprint_filter = self.request.GET.get("sprint")
        assignee_filter = self.request.GET.get("assignee")
        flagged_filter = self.request.GET.get("flagged")

        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if priority_filter:
            queryset = queryset.filter(priority=priority_filter)
        if project_filter:
            queryset = queryset.filter(project_id=project_filter)
        if sprint_filter:
            queryset = queryset.filter(sprint_id=sprint_filter)
        if assignee_filter:
            queryset = queryset.filter(assignee_id=assignee_filter)
        if flagged_filter == "1":
            queryset = queryset.filter(is_flagged=True)

        return queryset.order_by("project__name", "position", "-priority", "-created_at")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = ctx["object_list"]
        today = timezone.localdate()

        project_id = self.request.GET.get("project")
        current_project = None
        board_columns = []
        kanban_columns = []

        if project_id:
            current_project = (
                dm.Project.objects.select_related("workspace", "team", "owner", "product_manager")
                .filter(pk=project_id)
                .first()
            )
            if current_project:
                board_columns = list(
                    current_project.board_columns.all().order_by("position")
                )

        if not board_columns:
            board_columns = [
                type("Col", (), {"name": "À faire", "mapped_status": dm.Task.Status.TODO, "color": "#94A3B8", "wip_limit": None, "is_done_column": False}),
                type("Col", (), {"name": "En cours", "mapped_status": dm.Task.Status.IN_PROGRESS, "color": "#0EA5C9", "wip_limit": None, "is_done_column": False}),
                type("Col", (), {"name": "Review", "mapped_status": dm.Task.Status.REVIEW, "color": "#F59E0B", "wip_limit": None, "is_done_column": False}),
                type("Col", (), {"name": "Terminé", "mapped_status": dm.Task.Status.DONE, "color": "#10B981", "wip_limit": None, "is_done_column": True}),
                type("Col", (), {"name": "Bloqué", "mapped_status": dm.Task.Status.BLOCKED, "color": "#EF4444", "wip_limit": None, "is_done_column": False}),
                type("Col", (), {"name": "Expirée", "mapped_status": dm.Task.Status.EXPIRED, "color": "#7F1D1D", "wip_limit": None, "is_done_column": True}),
            ]

        for col in board_columns:
            tasks = [task for task in qs if task.status == col.mapped_status]
            kanban_columns.append({
                "name": col.name,
                "mapped_status": col.mapped_status,
                "color": getattr(col, "color", "#7C6FF7"),
                "wip_limit": getattr(col, "wip_limit", None),
                "is_done_column": getattr(col, "is_done_column", False),
                "tasks": tasks,
                "count": len(tasks),
            })

        ctx.update({
            "today": today,
            "current_project_obj": current_project,
            "stats": qs.aggregate(
                total=Count("id"),
                todo=Count("id", filter=Q(status=dm.Task.Status.TODO)),
                in_progress=Count("id", filter=Q(status=dm.Task.Status.IN_PROGRESS)),
                review=Count("id", filter=Q(status=dm.Task.Status.REVIEW)),
                done=Count("id", filter=Q(status=dm.Task.Status.DONE)),
                blocked=Count("id", filter=Q(status=dm.Task.Status.BLOCKED)),
                cancelled=Count("id", filter=Q(status=dm.Task.Status.CANCELLED)),
                critical=Count("id", filter=Q(priority=dm.Task.Priority.CRITICAL)),
                flagged=Count("id", filter=Q(is_flagged=True)),
                overdue=Count("id", filter=Q(due_date__lt=today) & ~Q(status=dm.Task.Status.DONE)),
                total_estimate=Coalesce(
                    Sum("estimate_hours"),
                    Value(Decimal("0.00")),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                ),
                total_spent=Coalesce(
                    Sum("spent_hours"),
                    Value(Decimal("0.00")),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                ),
            ),
            "projects_filter": dm.Project.objects.filter(is_archived=False).order_by("name").only("id", "name"),
            "sprints_filter": dm.Sprint.objects.filter(is_archived=False).order_by("-start_date").only("id", "name"),
            "assignees_filter": (
                dm.Task.objects.filter(assignee__isnull=False)
                .select_related("assignee")
                .values("assignee__id", "assignee__username")
                .distinct()
                .order_by("assignee__username")
            ),
            "assignable_users": (
                get_user_model().objects.filter(is_active=True)
                .order_by("username")
                .only("id", "username", "first_name", "last_name")
            ),
            "current_status": self.request.GET.get("status", ""),
            "current_priority": self.request.GET.get("priority", ""),
            "current_project": self.request.GET.get("project", ""),
            "current_sprint": self.request.GET.get("sprint", ""),
            "current_assignee": self.request.GET.get("assignee", ""),
            "current_flagged": self.request.GET.get("flagged", ""),
            "kanban_columns": kanban_columns,
            "board_columns": board_columns,
            "is_kanban_mode": True,
        })
        ctx["columns"] = [
            ("TODO", "À faire", {
                "title_class": "text-[var(--text2)]",
                "badge_class": "bg-[var(--bg2)] text-[var(--text2)]",
            }),
            ("IN_PROGRESS", "En cours", {
                "title_class": "text-[var(--accent)]",
                "badge_class": "bg-[var(--accent-bg)] text-[var(--accent)]",
            }),
            ("REVIEW", "Review", {
                "title_class": "text-[var(--amber)]",
                "badge_class": "bg-[var(--amber-bg)] text-[var(--amber)]",
            }),
            ("DONE", "Terminé", {
                "title_class": "text-[var(--green)]",
                "badge_class": "bg-[var(--green-bg)] text-[var(--green)]",
            }),
            ("BLOCKED", "Bloqué", {
                "title_class": "text-[var(--red)]",
                "badge_class": "bg-[var(--red-bg)] text-[var(--red)]",
            }),
            ("CANCELLED", "Annulé", {
                "title_class": "text-[var(--text3)]",
                "badge_class": "bg-[var(--bg2)] text-[var(--text3)]",
            }),
        ]
        return ctx


class TaskDetailView(DevflowDetailView):
    model = dm.Task
    template_name = "project/task/detail.html"
    section = "task"
    page_title = "Détail tâche"

    def get_queryset(self):
        return (
            dm.Task.objects
            .select_related(
                "workspace",
                "project",
                "sprint",
                "backlog_item",
                "parent",
                "reporter",
                "assignee",
            )
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        task = self.object
        today = timezone.localdate()

        comments = task.comments.select_related("author")
        attachments = task.attachments.all()
        assignments = task.assignments.select_related("user", "assigned_by")
        pull_requests = task.pull_requests.select_related("author")
        risks = task.risks.filter(is_archived=False).select_related("owner")
        checklists = task.checklists.prefetch_related("items")
        outgoing_dependencies = task.outgoing_dependencies.select_related("to_task")
        incoming_dependencies = task.incoming_dependencies.select_related("from_task")
        labels = task.labels.select_related("label")
        subtasks = task.subtasks.filter(is_archived=False).select_related("assignee", "reporter")
        insights = task.ai_insights.filter(is_dismissed=False).order_by("-detected_at")

        checklist_items_total = 0
        checklist_items_done = 0
        for checklist in checklists:
            items = list(checklist.items.all())
            checklist.items_cached = items
            checklist_items_total += len(items)
            checklist_items_done += sum(1 for i in items if i.is_checked)

        related_tasks = (
            dm.Task.objects.filter(project=task.project, is_archived=False)
            .exclude(pk=task.pk)
            .select_related("assignee", "sprint")
            .order_by("-updated_at")[:6]
        )

        is_overdue = bool(
            task.due_date and task.due_date < today
            and task.status not in (
                dm.Task.Status.DONE,
                dm.Task.Status.CANCELLED,
                dm.Task.Status.EXPIRED,
            )
        )

        ctx.update({
            "comments": comments,
            "attachments": attachments,
            "assignments": assignments,
            "pull_requests": pull_requests,
            "risks": risks,
            "checklists": checklists,
            "outgoing_dependencies": outgoing_dependencies,
            "incoming_dependencies": incoming_dependencies,
            "labels": labels,
            "subtasks": subtasks,
            "insights": insights,
            "related_tasks": related_tasks,
            "is_overdue": is_overdue,
            "checklist_items_total": checklist_items_total,
            "checklist_items_done": checklist_items_done,
            "checklist_progress": int(
                (checklist_items_done / checklist_items_total) * 100) if checklist_items_total else 0,
        })
        return ctx


class TaskCreateView(DevflowCreateView):
    model = dm.Task
    form_class = TaskForm
    template_name = "project/task/form.html"
    section = "task"
    page_title = "Créer tâche"
    success_list_url_name = "task_list"

    def get_initial(self):
        initial = super().get_initial()
        project_id = self.request.GET.get("project")
        sprint_id = self.request.GET.get("sprint")
        parent_id = self.request.GET.get("parent")
        assignee_id = self.request.GET.get("assignee")
        backlog_id = self.request.GET.get("backlog_item")

        if project_id:
            initial["project"] = project_id
        if sprint_id:
            initial["sprint"] = sprint_id
        if parent_id:
            initial["parent"] = parent_id
        if assignee_id:
            initial["assignee"] = assignee_id
        if backlog_id:
            initial["backlog_item"] = backlog_id

        # Initiation status par défaut sur "TODO" pour éviter le piège UX
        # où le formulaire reste vide et bloque sur le clean.
        initial.setdefault("status", dm.Task.Status.TODO)
        initial.setdefault("priority", dm.Task.Priority.MEDIUM)
        return initial

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def form_valid(self, form):
        obj = form.save(commit=False)

        if obj.project_id and not obj.workspace_id:
            obj.workspace = obj.project.workspace

        if self.request.user.is_authenticated and not obj.reporter_id:
            obj.reporter = self.request.user

        obj.comments_count = obj.comments_count or 0
        obj.attachments_count = obj.attachments_count or 0
        obj.position = obj.position or 0

        obj.save()
        form.save_m2m()

        # Si l'utilisateur a aussi rempli l'assignee, on crée le
        # TaskAssignment correspondant pour garder la cohérence (sinon il
        # n'y a que le FK assignee, pas la M2M assignments).
        if obj.assignee_id and not dm.TaskAssignment.objects.filter(
            task=obj, user=obj.assignee
        ).exists():
            dm.TaskAssignment.objects.create(
                task=obj,
                user=obj.assignee,
                assigned_by=self.request.user if self.request.user.is_authenticated else None,
                allocation_percent=100,
            )

        messages.success(self.request, "Tâche créée avec succès.")
        self.object = obj
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))

class TaskUpdateView(DevflowUpdateView):
    model = dm.Task
    form_class = TaskForm
    template_name = "project/task/form.html"
    section = "task"
    page_title = "Modifier tâche"
    success_list_url_name = "task_list"

    def form_valid(self, form):
        messages.success(self.request, "Tâche modifiée avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class TaskDeleteView(DevflowDeleteView):
    model = dm.Task
    section = "task"
    page_title = "Supprimer tâche"
    success_list_url_name = "task_list"


class TaskArchiveView(ArchiveObjectView):
    model = dm.Task
    success_list_url_name = "task_list"


class TaskMoveView(DevflowBaseMixin, View):
    def post(self, request, pk):
        task = self.filter_by_workspace(dm.Task.objects.all()).get(pk=pk)
        status = request.POST.get("status")
        position = request.POST.get("position")
        if status in dict(dm.Task.Status.choices):
            task.status = status
        if position:
            task.position = int(position)
        task.save()
        messages.success(request, "Tâche déplacée avec succès.")
        return redirect("task_detail", pk=task.pk)


class TaskExtendView(WorkspaceSecurityMixin, DevflowBaseMixin, View):
    """
    Reconduire une tâche en retard : nouveau couple (start_date, due_date)
    posé par le PM. Retour au statut TODO si elle était EXPIRED.
    """
    template_name = "project/task/extend.html"

    def _get_task_or_403(self, request, pk):
        task = self.filter_by_workspace(
            dm.Task.objects.all()
        ).select_related("project", "workspace").filter(pk=pk).first()
        if not task:
            return None
        # Seul le PM ou l'owner peut arbitrer
        pm = task.project.product_manager_id if task.project_id else None
        owner = task.project.owner_id if task.project_id else None
        if request.user.id not in (pm, owner) and not request.user.is_staff:
            return False
        return task

    def get(self, request, pk):
        task = self._get_task_or_403(request, pk)
        if task is None:
            messages.error(request, "Tâche introuvable.")
            return redirect("task_list")
        if task is False:
            messages.error(request, "Action réservée au chef de projet ou propriétaire.")
            return redirect("task_detail", pk=pk)
        return render(request, self.template_name, {"task": task})

    def post(self, request, pk):
        from datetime import date as _date
        task = self._get_task_or_403(request, pk)
        if task is None:
            messages.error(request, "Tâche introuvable.")
            return redirect("task_list")
        if task is False:
            messages.error(request, "Action réservée au chef de projet ou propriétaire.")
            return redirect("task_detail", pk=pk)

        new_start = request.POST.get("start_date") or ""
        new_due = request.POST.get("due_date") or ""
        try:
            new_due_d = _date.fromisoformat(new_due) if new_due else None
        except ValueError:
            new_due_d = None
        try:
            new_start_d = _date.fromisoformat(new_start) if new_start else None
        except ValueError:
            new_start_d = None
        if not new_due_d:
            messages.error(request, "Une nouvelle échéance est requise.")
            return render(request, self.template_name, {"task": task})
        if new_due_d < timezone.localdate():
            messages.error(request, "L'échéance doit être dans le futur.")
            return render(request, self.template_name, {"task": task})

        task.start_date = new_start_d or task.start_date
        task.due_date = new_due_d
        task.expired_at = None
        task.pm_overdue_notified_at = None
        if task.status == dm.Task.Status.EXPIRED:
            task.status = dm.Task.Status.TODO
        task.save(update_fields=[
            "start_date", "due_date", "expired_at",
            "pm_overdue_notified_at", "status", "updated_at",
        ])

        try:
            dm.ActivityLog.objects.create(
                workspace=task.workspace, actor=request.user,
                project=task.project, task=task,
                activity_type=dm.ActivityLog.ActivityType.TASK_MOVED,
                title="Tâche reconduite",
                description=(
                    f"{request.user} a reconduit la tâche « {task.title} » "
                    f"jusqu'au {new_due_d:%d/%m/%Y}."
                ),
            )
        except Exception:
            pass
        messages.success(
            request,
            f"Tâche reconduite jusqu'au {new_due_d:%d/%m/%Y}."
        )
        return redirect("task_detail", pk=task.pk)


class TaskMarkExpiredView(WorkspaceSecurityMixin, DevflowBaseMixin, View):
    """Marque la tâche comme EXPIRED — statut définitif côté planning."""

    def post(self, request, pk):
        task = self.filter_by_workspace(
            dm.Task.objects.all()
        ).select_related("project", "workspace").filter(pk=pk).first()
        if not task:
            messages.error(request, "Tâche introuvable.")
            return redirect("task_list")
        pm = task.project.product_manager_id if task.project_id else None
        owner = task.project.owner_id if task.project_id else None
        if request.user.id not in (pm, owner) and not request.user.is_staff:
            messages.error(request, "Action réservée au chef de projet ou propriétaire.")
            return redirect("task_detail", pk=pk)

        task.status = dm.Task.Status.EXPIRED
        task.expired_at = timezone.localdate()
        task.save(update_fields=["status", "expired_at", "updated_at"])
        try:
            dm.ActivityLog.objects.create(
                workspace=task.workspace, actor=request.user,
                project=task.project, task=task,
                activity_type=dm.ActivityLog.ActivityType.TASK_MOVED,
                title="Tâche marquée expirée",
                description=f"{request.user} a marqué la tâche « {task.title} » comme expirée non traitée.",
            )
        except Exception:
            pass
        messages.success(request, "Tâche marquée comme expirée non traitée.")
        return redirect("task_detail", pk=task.pk)


class TaskMarkDoneView(DevflowBaseMixin, View):
    def post(self, request, pk):
        task = self.filter_by_workspace(dm.Task.objects.all()).get(pk=pk)
        task.status = dm.Task.Status.DONE
        task.progress_percent = 100
        task.completed_at = timezone.now()
        task.save()
        messages.success(request, "Tâche marquée comme terminée.")
        return redirect("task_detail", pk=task.pk)


class TaskAssignmentListView(DevflowListView):
    model = dm.TaskAssignment
    template_name = "project/task_assignment/list.html"
    section = "task"
    page_title = "Affectations de tâches"
    search_placeholder = "Rechercher par utilisateur ou tâche…"
    search_fields = (
        "user__username", "user__first_name", "user__last_name",
        "assigned_by__username", "task__title",
    )
    filter_fields = [
        {"name": "user", "label": "Utilisateur", "type": "model",
         "queryset": "_users_qs", "lookup": "user_id"},
        {"name": "is_active", "label": "Active", "type": "bool"},
    ]

    def _users_qs(self):
        ws = self.get_current_workspace()
        if not ws:
            return User.objects.none()
        return User.objects.filter(
            is_active=True, devflow_memberships__workspace=ws
        ).distinct().order_by("last_name", "first_name", "username")


class TaskAssignmentDetailView(DevflowDetailView):
    model = dm.TaskAssignment
    template_name = "project/task_assignment/detail.html"
    section = "task"
    page_title = "Détail affectation"


class TaskAssignmentCreateView(DevflowCreateView):
    model = dm.TaskAssignment
    form_class = TaskAssignmentForm
    template_name = "project/task_assignment/form.html"
    section = "task"
    page_title = "Créer affectation tâche"
    success_list_url_name = "task_assignment_list"

    def get_initial(self):
        initial = super().get_initial()
        task_id = self.request.GET.get("task")
        user_id = self.request.GET.get("user")
        if task_id:
            initial["task"] = task_id
        if user_id:
            initial["user"] = user_id
        return initial


class TaskAssignmentUpdateView(DevflowUpdateView):
    model = dm.TaskAssignment
    form_class = TaskAssignmentForm
    template_name = "project/task_assignment/form.html"
    section = "task"
    page_title = "Modifier affectation tâche"
    success_list_url_name = "task_assignment_list"


class TaskAssignmentDeleteView(DevflowDeleteView):
    model = dm.TaskAssignment
    section = "task"
    page_title = "Supprimer affectation tâche"
    success_list_url_name = "task_assignment_list"


class TaskCommentListView(DevflowListView):
    model = dm.TaskComment
    template_name = "project/task_comment/list.html"
    context_object_name = "items"
    section = "task"
    page_title = "Commentaires de tâches"
    search_fields = ("body", "author__username", "task__title", "task__project__name")
    paginate_by = 30

    def get_queryset(self):
        qs = (
            super()
            .get_queryset()
            .select_related(
                "task",
                "task__project",
                "author",
            )
            .order_by("-created_at")
        )

        task_id = self.request.GET.get("task")
        project_id = self.request.GET.get("project")
        internal = self.request.GET.get("internal")

        if task_id:
            qs = qs.filter(task_id=task_id)

        if project_id:
            qs = qs.filter(task__project_id=project_id)

        if internal == "1":
            qs = qs.filter(is_internal=True)
        elif internal == "0":
            qs = qs.filter(is_internal=False)

        return qs

    def get_task(self):
        task_id = self.request.GET.get("task")
        if not task_id:
            return None
        return get_object_or_404(
            dm.Task.objects.select_related("project", "assignee", "reporter"),
            pk=task_id,
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        current_task = self.get_task()

        ctx["current_task"] = current_task
        ctx["current_project"] = current_task.project if current_task else None
        ctx["quick_form"] = TaskCommentQuickForm(task=current_task)
        ctx["tasks"] = dm.Task.objects.select_related("project").order_by("-created_at")[:100]

        return ctx

    def post(self, request, *args, **kwargs):
        current_task = self.get_task()
        form = TaskCommentQuickForm(request.POST, task=current_task)

        if form.is_valid():
            comment = form.save(commit=False)
            comment.author = request.user
            comment.save()

            if hasattr(comment.task, "comments_count"):
                dm.Task.objects.filter(pk=comment.task_id).update(
                    comments_count=comment.task.comments.count()
                )

            messages.success(request, "Commentaire ajouté avec succès.")

            redirect_url = reverse("task_comment_list")
            return redirect(f"{redirect_url}?task={comment.task_id}")

        self.object_list = self.get_queryset()
        context = self.get_context_data()
        context["quick_form"] = form
        return self.render_to_response(context)


class TaskCommentDetailView(DevflowDetailView):
    model = dm.TaskComment
    template_name = "project/task_comment/detail.html"
    section = "task"
    page_title = "Détail commentaire"


class TaskCommentCreateView(DevflowCreateView):
    model = dm.TaskComment
    form_class = TaskCommentForm
    section = "task"
    page_title = "Créer commentaire"
    success_list_url_name = "task_comment_list"


class TaskCommentUpdateView(DevflowUpdateView):
    model = dm.TaskComment
    form_class = TaskCommentForm
    section = "task"
    page_title = "Modifier commentaire"
    success_list_url_name = "task_comment_list"


class TaskCommentDeleteView(DevflowDeleteView):
    model = dm.TaskComment
    section = "task"
    page_title = "Supprimer commentaire"
    success_list_url_name = "task_comment_list"


class TaskAttachmentListView(DevflowListView):
    model = dm.TaskAttachment
    template_name = "project/task_attachment/list.html"
    section = "task"
    page_title = "Pièces jointes tâches"
    search_fields = ("name", "mime_type")


class TaskAttachmentDetailView(DevflowDetailView):
    model = dm.TaskAttachment
    template_name = "project/task_attachment/detail.html"
    section = "task"
    page_title = "Détail pièce jointe tâche"


class TaskAttachmentCreateView(DevflowCreateView):
    model = dm.TaskAttachment
    form_class = TaskAttachmentForm
    section = "task"
    page_title = "Créer pièce jointe tâche"
    success_list_url_name = "task_attachment_list"


class TaskAttachmentUpdateView(DevflowUpdateView):
    model = dm.TaskAttachment
    form_class = TaskAttachmentForm
    section = "task"
    page_title = "Modifier pièce jointe tâche"
    success_list_url_name = "task_attachment_list"


class TaskAttachmentDeleteView(DevflowDeleteView):
    model = dm.TaskAttachment
    section = "task"
    page_title = "Supprimer pièce jointe tâche"
    success_list_url_name = "task_attachment_list"


# =============================================================================
# ENGINEERING / RISK / AI
# =============================================================================
class PullRequestListView(DevflowListView):
    model = dm.PullRequest
    template_name = "project/pull_request/list.html"
    section = "project"
    page_title = "Pull requests"
    search_fields = ("title", "repository", "branch_name", "external_id")


class PullRequestDetailView(DevflowDetailView):
    model = dm.PullRequest
    template_name = "project/pull_request/detail.html"
    section = "project"
    page_title = "Détail pull request"


class PullRequestCreateView(DevflowCreateView):
    model = dm.PullRequest
    form_class = PullRequestForm
    section = "project"
    page_title = "Créer pull request"
    success_list_url_name = "pull_request_list"


class PullRequestUpdateView(DevflowUpdateView):
    model = dm.PullRequest
    form_class = PullRequestForm
    section = "project"
    page_title = "Modifier pull request"
    success_list_url_name = "pull_request_list"


class PullRequestDeleteView(DevflowDeleteView):
    model = dm.PullRequest
    section = "project"
    page_title = "Supprimer pull request"
    success_list_url_name = "pull_request_list"


class RiskListView(DevflowListView):
    model = dm.Risk
    template_name = "project/risk/list.html"
    section = "analytics"
    page_title = "Risques"
    search_fields = (
        "title",
        "description",
        "mitigation_plan",
        "severity",
        "status",
        "project__name",
        "task__title",
        "owner__username",
    )

    def get_queryset(self):
        return (
            dm.Risk.objects
            .select_related("workspace", "project", "task", "owner")
            .filter(is_archived=False)
            .order_by("-created_at")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = ctx["object_list"]
        today = timezone.localdate()

        ctx.update({
            "today": today,
            "stats": qs.aggregate(
                total=Count("id"),
                open_count=Count("id", filter=Q(status=dm.Risk.Status.OPEN)),
                mitigated_count=Count("id", filter=Q(status=dm.Risk.Status.MITIGATED)),
                escalated_count=Count("id", filter=Q(status=dm.Risk.Status.ESCALATED)),
                closed_count=Count("id", filter=Q(status=dm.Risk.Status.CLOSED)),
                low_count=Count("id", filter=Q(severity=dm.Risk.Severity.LOW)),
                medium_count=Count("id", filter=Q(severity=dm.Risk.Severity.MEDIUM)),
                high_count=Count("id", filter=Q(severity=dm.Risk.Severity.HIGH)),
                critical_count=Count("id", filter=Q(severity=dm.Risk.Severity.CRITICAL)),
                overdue_count=Count(
                    "id",
                    filter=Q(due_date__lt=today) & ~Q(status=dm.Risk.Status.CLOSED),
                ),
            ),
        })
        return ctx


class RiskDetailView(DevflowDetailView):
    model = dm.Risk
    template_name = "project/risk/detail.html"
    section = "analytics"
    page_title = "Détail risque"

    def get_queryset(self):
        return (
            dm.Risk.objects
            .select_related("workspace", "project", "task", "owner")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        risk = self.object
        today = timezone.localdate()

        related_risks = (
            dm.Risk.objects
            .filter(project=risk.project, is_archived=False)
            .exclude(pk=risk.pk)
            .select_related("owner", "task")
            .order_by("-created_at")[:5]
        )

        ctx.update({
            "today": today,
            "related_risks": related_risks,
            "is_overdue": bool(
                risk.due_date and risk.due_date < today and risk.status != dm.Risk.Status.CLOSED
            ),
        })
        return ctx


class RiskCreateView(DevflowCreateView):
    model = dm.Risk
    form_class = RiskForm
    template_name = "project/risk/form.html"
    section = "analytics"
    page_title = "Créer risque"
    success_list_url_name = "risk_list"

    def form_valid(self, form):
        messages.success(self.request, "Risque créé avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class RiskUpdateView(DevflowUpdateView):
    model = dm.Risk
    form_class = RiskForm
    template_name = "project/risk/form.html"
    section = "analytics"
    page_title = "Modifier risque"
    success_list_url_name = "risk_list"

    def form_valid(self, form):
        messages.success(self.request, "Risque modifié avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class RiskDeleteView(DevflowDeleteView):
    model = dm.Risk
    section = "analytics"
    page_title = "Supprimer risque"
    success_list_url_name = "risk_list"


class RiskArchiveView(ArchiveObjectView):
    model = dm.Risk
    success_list_url_name = "risk_list"


class AInsightDashboardView(DevflowBaseMixin, TemplateView):
    template_name = "project/ai_insight/dashboard.html"
    section = "analytics"
    page_title = "Dashboard IA"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        insights = (
            self.filter_by_workspace(dm.AInsight.objects.all())
            .select_related("project", "sprint", "task")
            .order_by("-detected_at", "-created_at")
        )

        active_insights = insights.filter(is_dismissed=False)
        hidden_insights = insights.filter(is_dismissed=True)

        stats = active_insights.aggregate(
            total=Count("id"),
            unread_count=Count("id", filter=Q(is_read=False)),
            read_count=Count("id", filter=Q(is_read=True)),
            critical_count=Count("id", filter=Q(severity=dm.AInsight.Severity.CRITICAL)),
            high_count=Count("id", filter=Q(severity=dm.AInsight.Severity.HIGH)),
            medium_count=Count("id", filter=Q(severity=dm.AInsight.Severity.MEDIUM)),
            low_count=Count("id", filter=Q(severity=dm.AInsight.Severity.LOW)),
            info_count=Count("id", filter=Q(severity=dm.AInsight.Severity.INFO)),
            avg_score=Avg("score"),
        )

        type_breakdown = [
            {
                "key": choice[0],
                "label": choice[1],
                "count": active_insights.filter(insight_type=choice[0]).count(),
            }
            for choice in dm.AInsight.InsightType.choices
        ]

        severity_breakdown = [
            {
                "key": choice[0],
                "label": choice[1],
                "count": active_insights.filter(severity=choice[0]).count(),
            }
            for choice in dm.AInsight.Severity.choices
        ]

        insights_by_project = (
            active_insights
            .values("project__id", "project__name")
            .annotate(
                total=Count("id"),
                unread=Count("id", filter=Q(is_read=False)),
                critical=Count("id", filter=Q(severity=dm.AInsight.Severity.CRITICAL)),
                avg_score=Avg("score"),
            )
            .order_by("-total", "-critical", "project__name")
        )

        critical_unread = (
            active_insights
            .filter(
                is_read=False,
                severity__in=[
                    dm.AInsight.Severity.CRITICAL,
                    dm.AInsight.Severity.HIGH,
                ],
            )
            .select_related("project", "sprint", "task")
            .order_by("-score", "-detected_at")[:8]
        )

        priority_recommendations = (
            active_insights
            .exclude(recommendation="")
            .filter(
                severity__in=[
                    dm.AInsight.Severity.CRITICAL,
                    dm.AInsight.Severity.HIGH,
                    dm.AInsight.Severity.MEDIUM,
                ]
            )
            .select_related("project", "sprint", "task")
            .order_by("-score", "-detected_at")[:6]
        )

        recent_active = active_insights.select_related("project", "sprint", "task")[:10]
        recent_hidden = hidden_insights.select_related("project", "sprint", "task")[:10]

        ctx.update({
            "stats": stats,
            "active_count": active_insights.count(),
            "hidden_count": hidden_insights.count(),
            "type_breakdown": type_breakdown,
            "severity_breakdown": severity_breakdown,
            "insights_by_project": insights_by_project,
            "critical_unread": critical_unread,
            "priority_recommendations": priority_recommendations,
            "recent_active": recent_active,
            "recent_hidden": recent_hidden,
        })
        return ctx


class AInsightListView(DevflowListView):
    model = dm.AInsight
    template_name = "project/ai_insight/list.html"
    section = "analytics"
    page_title = "Insights IA"
    search_fields = (
        "title",
        "summary",
        "recommendation",
        "insight_type",
        "severity",
        "project__name",
        "sprint__name",
        "task__title",
    )

    def get_queryset(self):
        return (
            dm.AInsight.objects
            .select_related("workspace", "project", "sprint", "task")
            .filter(is_dismissed=False)
            .order_by("-detected_at", "-created_at")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = ctx["object_list"]
        today = timezone.now()

        ctx.update({
            "today": today,
            "stats": qs.aggregate(
                total=Count("id"),
                unread_count=Count("id", filter=Q(is_read=False)),
                read_count=Count("id", filter=Q(is_read=True)),
                info_count=Count("id", filter=Q(severity=dm.AInsight.Severity.INFO)),
                low_count=Count("id", filter=Q(severity=dm.AInsight.Severity.LOW)),
                medium_count=Count("id", filter=Q(severity=dm.AInsight.Severity.MEDIUM)),
                high_count=Count("id", filter=Q(severity=dm.AInsight.Severity.HIGH)),
                critical_count=Count("id", filter=Q(severity=dm.AInsight.Severity.CRITICAL)),
                avg_score=Avg("score"),
                risk_count=Count("id", filter=Q(insight_type=dm.AInsight.InsightType.RISK)),
                workload_count=Count("id", filter=Q(insight_type=dm.AInsight.InsightType.WORKLOAD)),
                delivery_count=Count("id", filter=Q(insight_type=dm.AInsight.InsightType.DELIVERY)),
                suggestion_count=Count("id", filter=Q(insight_type=dm.AInsight.InsightType.SUGGESTION)),
            ),
        })
        return ctx


class AInsightDetailView(DevflowDetailView):
    model = dm.AInsight
    template_name = "project/ai_insight/detail.html"
    section = "analytics"
    page_title = "Détail insight IA"

    def get_queryset(self):
        return (
            dm.AInsight.objects
            .select_related("workspace", "project", "sprint", "task")
        )

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        if not obj.is_read:
            obj.is_read = True
            obj.save(update_fields=["is_read", "updated_at"])
        return obj

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        insight = self.object

        related_insights = (
            dm.AInsight.objects
            .filter(is_dismissed=False)
            .exclude(pk=insight.pk)
            .select_related("project", "sprint", "task")
            .filter(
                Q(project=insight.project) |
                Q(sprint=insight.sprint) |
                Q(task=insight.task) |
                Q(insight_type=insight.insight_type)
            )
            .distinct()
            .order_by("-detected_at")[:6]
        )

        ctx.update({
            "related_insights": related_insights,
        })
        return ctx


class AInsightCreateView(DevflowCreateView):
    model = dm.AInsight
    form_class = AInsightForm
    template_name = "project/ai_insight/form.html"
    section = "analytics"
    page_title = "Créer insight IA"
    success_list_url_name = "ai_insight_list"

    def get_initial(self):
        initial = super().get_initial()
        initial.setdefault("detected_at", timezone.now())
        initial.setdefault("score", 50)
        initial.setdefault("severity", dm.AInsight.Severity.INFO)
        return initial

    def form_valid(self, form):
        messages.success(self.request, "Insight IA créé avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class AInsightUpdateView(DevflowUpdateView):
    model = dm.AInsight
    form_class = AInsightForm
    template_name = "project/ai_insight/form.html"
    section = "analytics"
    page_title = "Modifier insight IA"
    success_list_url_name = "ai_insight_list"

    def form_valid(self, form):
        messages.success(self.request, "Insight IA modifié avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class AInsightDeleteView(DevflowDeleteView):
    model = dm.AInsight
    section = "analytics"
    page_title = "Supprimer insight IA"
    success_list_url_name = "ai_insight_list"


class AInsightDismissView(DevflowBaseMixin, View):
    def post(self, request, pk):
        insight = self.filter_by_workspace(dm.AInsight.objects.all()).get(pk=pk)
        insight.is_dismissed = True
        insight.save(update_fields=["is_dismissed", "updated_at"])
        messages.success(request, "Insight masqué.")
        return redirect("ai_insight_list")


# =============================================================================
# NOTIFICATIONS / ACTIVITY
# =============================================================================
class NotificationListView(DevflowListView):
    model = dm.Notification
    template_name = "project/notification/list.html"
    section = "notification"
    page_title = "Notifications"
    search_fields = ("title", "body", "notification_type")

    def get_queryset(self):
        qs = (
            super()
            .get_queryset()
            .select_related("workspace", "recipient")
        )

        if self.request.user.is_authenticated:
            qs = qs.filter(recipient=self.request.user)
        else:
            qs = qs.none()

        return qs.order_by("-created_at")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        base_qs = self.get_queryset()

        ctx["unread_count"] = base_qs.filter(is_read=False).count()
        ctx["total_count"] = base_qs.count()
        ctx["read_count"] = base_qs.filter(is_read=True).count()

        return ctx


class NotificationDetailView(DevflowDetailView):
    model = dm.Notification
    template_name = "project/notification/detail.html"
    section = "notification"
    page_title = "Détail notification"

    def get_queryset(self):
        return super().get_queryset().filter(recipient=self.request.user)


class NotificationCreateView(DevflowCreateView):
    model = dm.Notification
    form_class = NotificationForm
    section = "notification"
    page_title = "Créer notification"
    success_list_url_name = "notification_list"


class NotificationUpdateView(DevflowUpdateView):
    model = dm.Notification
    form_class = NotificationForm
    section = "notification"
    page_title = "Modifier notification"
    success_list_url_name = "notification_list"


class NotificationDeleteView(DevflowDeleteView):
    model = dm.Notification
    section = "notification"
    page_title = "Supprimer notification"
    success_list_url_name = "notification_list"


class NotificationMarkReadView(DevflowBaseMixin, View):
    def post(self, request, pk):
        notification = self.filter_by_workspace(dm.Notification.objects.filter(recipient=request.user)).get(pk=pk)
        notification.is_read = True
        notification.read_at = timezone.now()
        notification.save(update_fields=["is_read", "read_at", "updated_at"])
        messages.success(request, "Notification marquée comme lue.")
        return redirect("notification_list")


class NotificationMarkAllReadView(DevflowBaseMixin, View):
    def post(self, request):
        queryset = self.filter_by_workspace(dm.Notification.objects.filter(recipient=request.user, is_read=False))
        queryset.update(is_read=True, read_at=timezone.now())
        messages.success(request, "Toutes les notifications ont été marquées comme lues.")
        return redirect("notification_list")


class ActivityLogListView(DevflowListView):
    model = dm.ActivityLog
    template_name = "project/activity_log/list.html"
    section = "analytics"
    page_title = "Activité"
    search_fields = ("title", "description", "activity_type")


class ActivityLogDetailView(DevflowDetailView):
    model = dm.ActivityLog
    template_name = "project/activity_log/detail.html"
    section = "analytics"
    page_title = "Détail activité"


class ActivityLogCreateView(DevflowCreateView):
    model = dm.ActivityLog
    form_class = ActivityLogForm
    section = "analytics"
    page_title = "Créer activité"
    success_list_url_name = "activity_log_list"


class ActivityLogUpdateView(DevflowUpdateView):
    model = dm.ActivityLog
    form_class = ActivityLogForm
    section = "analytics"
    page_title = "Modifier activité"
    success_list_url_name = "activity_log_list"


class ActivityLogDeleteView(DevflowDeleteView):
    model = dm.ActivityLog
    section = "analytics"
    page_title = "Supprimer activité"
    success_list_url_name = "activity_log_list"


# =============================================================================
# CHAT / CHANNELS / MESSAGES
# =============================================================================
class DirectChannelListView(DevflowListView):
    model = dm.DirectChannel
    template_name = "project/direct_channel/list.html"
    section = "messages"
    page_title = "Canaux"
    search_fields = ("name",)


class DirectChannelDetailView(DevflowDetailView):
    model = dm.DirectChannel
    template_name = "project/chat/channel_detail.html"
    section = "messages"
    page_title = "Détail canal"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        channel = self.object
        ctx["memberships"] = channel.memberships.select_related("user")
        ctx["messages"] = channel.messages.select_related("author", "parent")
        return ctx


class DirectChannelCreateView(DevflowCreateView):
    model = dm.DirectChannel
    form_class = DirectChannelForm
    section = "messages"
    page_title = "Créer canal"
    success_list_url_name = "direct_channel_list"


class DirectChannelUpdateView(DevflowUpdateView):
    model = dm.DirectChannel
    form_class = DirectChannelForm
    section = "messages"
    page_title = "Modifier canal"
    success_list_url_name = "direct_channel_list"


class DirectChannelDeleteView(DevflowDeleteView):
    model = dm.DirectChannel
    section = "messages"
    page_title = "Supprimer canal"
    success_list_url_name = "direct_channel_list"


class ChannelMembershipListView(DevflowListView):
    model = dm.ChannelMembership
    template_name = "project/channel_membership/list.html"
    section = "messages"
    page_title = "Membres des canaux"
    search_fields = ("user__username",)


class ChannelMembershipDetailView(DevflowDetailView):
    model = dm.ChannelMembership
    template_name = "project/channel_membership/detail.html"
    section = "messages"
    page_title = "Détail membre canal"


class ChannelMembershipCreateView(DevflowCreateView):
    model = dm.ChannelMembership
    form_class = ChannelMembershipForm
    section = "messages"
    page_title = "Ajouter membre canal"
    success_list_url_name = "channel_membership_list"


class ChannelMembershipUpdateView(DevflowUpdateView):
    model = dm.ChannelMembership
    form_class = ChannelMembershipForm
    section = "messages"
    page_title = "Modifier membre canal"
    success_list_url_name = "channel_membership_list"


class ChannelMembershipDeleteView(DevflowDeleteView):
    model = dm.ChannelMembership
    section = "messages"
    page_title = "Supprimer membre canal"
    success_list_url_name = "channel_membership_list"


class MessageListView(DevflowListView):
    model = dm.Message
    template_name = "project/message/list.html"
    section = "messages"
    page_title = "Messages"
    search_fields = ("body", "author__username")


class MessageDetailView(DevflowDetailView):
    model = dm.Message
    template_name = "project/message/detail.html"
    section = "messages"
    page_title = "Détail message"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        message = self.object
        ctx["replies"] = message.replies.select_related("author")
        ctx["attachments"] = message.attachments.all()
        ctx["reactions"] = message.reactions.select_related("user")
        return ctx


class MessageCreateView(DevflowCreateView):
    model = dm.Message
    form_class = MessageForm
    section = "messages"
    page_title = "Créer message"
    success_list_url_name = "message_list"


class MessageUpdateView(DevflowUpdateView):
    model = dm.Message
    form_class = MessageForm
    section = "messages"
    page_title = "Modifier message"
    success_list_url_name = "message_list"


class MessageDeleteView(DevflowDeleteView):
    model = dm.Message
    section = "messages"
    page_title = "Supprimer message"
    success_list_url_name = "message_list"


# =============================================================================
# TIMESHEET / SNAPSHOT / PREFERENCES
# =============================================================================
class TimesheetEntryListView(DevflowListView):
    model = dm.TimesheetEntry
    template_name = "project/timesheet_entry/list.html"
    section = "timesheet"
    page_title = "Timesheets"
    search_fields = ("description", "user__username", "user__first_name", "user__last_name", "project__name")
    paginate_by = 50

    def get_context_data(self, **kwargs):
        from datetime import timedelta as _td
        from collections import defaultdict
        from django.db.models import Sum, Count

        ctx = super().get_context_data(**kwargs)
        qs = self.get_queryset()
        ws = self.get_current_workspace()

        groups_dict = {}
        for entry in qs.select_related("user").order_by("-entry_date", "user_id"):
            monday = entry.entry_date - _td(days=entry.entry_date.weekday())
            key = (entry.user_id, monday)
            if key not in groups_dict:
                groups_dict[key] = {
                    "user": entry.user,
                    "monday": monday,
                    "sunday": monday + _td(days=6),
                    "iso_week": monday.isocalendar()[1],
                    "year": monday.year,
                    "total_hours": 0.0,
                    "entry_count": 0,
                    "statuses": set(),
                }
            grp = groups_dict[key]
            grp["total_hours"] += float(entry.hours or 0)
            grp["entry_count"] += 1
            grp["statuses"].add(entry.approval_status)

        # Calcul du statut consolidé
        for grp in groups_dict.values():
            statuses = grp["statuses"]
            if statuses == {dm.TimesheetEntry.ApprovalStatus.APPROVED}:
                grp["status"] = dm.TimesheetEntry.ApprovalStatus.APPROVED
            elif dm.TimesheetEntry.ApprovalStatus.REJECTED in statuses:
                grp["status"] = dm.TimesheetEntry.ApprovalStatus.REJECTED
            elif dm.TimesheetEntry.ApprovalStatus.SUBMITTED in statuses:
                grp["status"] = dm.TimesheetEntry.ApprovalStatus.SUBMITTED
            else:
                grp["status"] = dm.TimesheetEntry.ApprovalStatus.DRAFT
            grp["status_label"] = dict(dm.TimesheetEntry.ApprovalStatus.choices).get(
                grp["status"], "—"
            )
            del grp["statuses"]

        groups = sorted(groups_dict.values(), key=lambda g: (g["monday"], g["user"].username), reverse=True)
        ctx["weekly_groups"] = groups
        ctx["status_choices"] = dm.TimesheetEntry.ApprovalStatus.choices
        return ctx


class TimesheetWeekValidateView(WorkspaceSecurityMixin, DevflowBaseMixin, View):
    """
    POST /timesheets/week/validate/  — valide (APPROVED) toutes les entries
    de la semaine d'un utilisateur.
    POST avec params : user (pk), monday (YYYY-MM-DD), action (approve|reject|reopen|submit).
    """

    def post(self, request, *args, **kwargs):
        from datetime import date as _date, timedelta as _td

        user_id = request.POST.get("user")
        monday_raw = request.POST.get("monday")
        action = (request.POST.get("action") or "approve").strip()

        if not (user_id and monday_raw):
            messages.error(request, "Paramètres manquants.")
            return redirect("timesheet_entry_list")
        try:
            monday = _date.fromisoformat(monday_raw)
        except ValueError:
            messages.error(request, "Date invalide.")
            return redirect("timesheet_entry_list")

        target_user = get_object_or_404(User, pk=user_id)
        sunday = monday + _td(days=6)
        qs = self.filter_by_workspace(
            dm.TimesheetEntry.objects.filter(
                user=target_user,
                entry_date__gte=monday,
                entry_date__lte=sunday,
            )
        )

        if not qs.exists():
            messages.warning(request, "Aucune entrée pour cette semaine.")
            return redirect("timesheet_entry_list")

        action_map = {
            "submit": dm.TimesheetEntry.ApprovalStatus.SUBMITTED,
            "approve": dm.TimesheetEntry.ApprovalStatus.APPROVED,
            "reject": dm.TimesheetEntry.ApprovalStatus.REJECTED,
            "reopen": dm.TimesheetEntry.ApprovalStatus.DRAFT,
        }
        if action not in action_map:
            messages.error(request, "Action inconnue.")
            return redirect("timesheet_entry_list")

        new_status = action_map[action]
        update_fields = {"approval_status": new_status}
        if action in ("approve", "reject"):
            update_fields["approved_by"] = request.user
            update_fields["approved_at"] = timezone.now()
        elif action == "reopen":
            update_fields["approved_by"] = None
            update_fields["approved_at"] = None
        qs.update(**update_fields)

        labels = {
            "submit": "Semaine soumise pour validation",
            "approve": f"Semaine du {monday:%d/%m/%Y} validée pour {target_user}",
            "reject": "Semaine rejetée — l'utilisateur peut la modifier puis re-soumettre",
            "reopen": "Semaine rouverte pour modifications",
        }
        messages.success(request, labels[action])
        return redirect(request.META.get("HTTP_REFERER") or "timesheet_entry_list")


class TimesheetEntryDetailView(DevflowDetailView):
    model = dm.TimesheetEntry
    template_name = "project/timesheet_entry/detail.html"
    section = "timesheet"
    page_title = "Détail timesheet"


class TimesheetEntryCreateView(DevflowCreateView):
    model = dm.TimesheetEntry
    form_class = TimesheetEntryForm
    section = "timesheet"
    page_title = "Créer entrée timesheet"
    success_list_url_name = "timesheet_entry_list"


class TimesheetEntryUpdateView(DevflowUpdateView):
    model = dm.TimesheetEntry
    form_class = TimesheetEntryForm
    section = "timesheet"
    page_title = "Modifier entrée timesheet"
    success_list_url_name = "timesheet_entry_list"


class TimesheetEntryDeleteView(DevflowDeleteView):
    model = dm.TimesheetEntry
    section = "timesheet"
    page_title = "Supprimer entrée timesheet"
    success_list_url_name = "timesheet_entry_list"


# =========================================================================
# CALENDRIER TIMESHEETS — saisie hebdomadaire style Microsoft Teams
# =========================================================================
class TimesheetCalendarView(WorkspaceSecurityMixin, DevflowBaseMixin, View):
    """
    Vue calendrier hebdomadaire :
      - lignes = tâches actives assignées au user
      - colonnes = 7 jours de la semaine sélectionnée
      - cellules = total d'heures saisies pour (task, day)
      - saisie inline : POST AJAX vers timesheet_calendar_save
    """

    template_name = "project/timesheet_entry/calendar.html"

    def _resolve_week(self, request):
        from datetime import date as _date, timedelta as _td

        raw = request.GET.get("date")
        try:
            cur = _date.fromisoformat(raw) if raw else timezone.localdate()
        except ValueError:
            cur = timezone.localdate()
        # Monday of the week
        monday = cur - _td(days=cur.weekday())
        days = [monday + _td(days=i) for i in range(7)]
        return monday, days

    def get(self, request, *args, **kwargs):
        from datetime import timedelta as _td
        from django.db.models import Sum

        ws = self.get_current_workspace()
        monday, days = self._resolve_week(request)
        labels = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
        # Liste de couples (label, date) pour itération directe en template
        days_meta = [{"label": labels[i], "date": d} for i, d in enumerate(days)]

        # ── 1. Tâches actives ───────────────────────────────────────
        # Une tâche apparaît sur la semaine si :
        #   - elle est assignée à l'utilisateur (FK ou TaskAssignment actif), ET
        #   - sa période [start, end] intersecte la semaine [monday, sunday]
        # Périmètre période :
        #   - début = task.start_date  (sinon task.created_at, sinon -∞)
        #   - fin   = task.due_date    (sinon +∞ ; si tâche DONE, fin = completed_at)
        # Exception : on garde la tâche si l'utilisateur a déjà saisi
        # des heures dessus dans la semaine (pour ne pas perdre une saisie).
        sunday_date = days[-1]
        all_assigned_ids = set(
            dm.TaskAssignment.objects.filter(
                user=request.user, is_active=True,
                task__is_archived=False,
            ).values_list("task_id", flat=True)
        )
        all_assigned_ids |= set(
            dm.Task.objects.filter(
                assignee=request.user, is_archived=False,
            ).exclude(status__in=[
                dm.Task.Status.DONE,
                dm.Task.Status.CANCELLED,
                dm.Task.Status.EXPIRED,
            ]).values_list("id", flat=True)
        )

        candidate_tasks = (
            dm.Task.objects.filter(pk__in=all_assigned_ids)
            .select_related("project")
        )

        def _task_visible_on_week(t):
            # Borne gauche
            left = t.start_date or (t.created_at.date() if t.created_at else None)
            if left and left > sunday_date:
                return False
            # Borne droite : si la tâche a été terminée AVANT le lundi, on cache.
            if t.status == dm.Task.Status.DONE and t.completed_at:
                if t.completed_at.date() < monday:
                    return False
            # Si due_date est strictement < monday ET la tâche n'est pas complétée,
            # on cache aussi (la tâche est "hors période") sauf si EXPIRED → cachée
            # car le PM a tranché.
            if t.due_date and t.due_date < monday and t.status not in (
                dm.Task.Status.DONE,
            ):
                # Tâche hors période : on cache, sauf si l'utilisateur a saisi
                # des heures dessus cette semaine (cas géré ensuite).
                return False
            return True

        tasks_qs = sorted(
            (t for t in candidate_tasks if _task_visible_on_week(t)),
            key=lambda t: (t.project.name if t.project_id else "", t.title),
        )

        entry_qs = (
            dm.TimesheetEntry.objects.filter(
                user=request.user,
                entry_date__gte=monday,
                entry_date__lte=days[-1],
            )
            .select_related("task", "project")
        )

        existing_task_ids = {t.id for t in tasks_qs}
        # Tâches qui ont des entries cette semaine mais qu'on aurait masquées :
        # on les ré-injecte pour ne jamais perdre une saisie.
        extra_ids = set(
            entry_qs.exclude(task__isnull=True)
            .exclude(task_id__in=existing_task_ids)
            .values_list("task_id", flat=True)
        )
        tasks_list = list(tasks_qs)
        if extra_ids:
            tasks_list += list(
                dm.Task.objects.filter(pk__in=extra_ids).select_related("project")
            )

        # ── 2. Statut semaine (verrou si APPROVED) ──────────────────
        week_status = self._compute_week_status(entry_qs)
        is_locked = week_status == dm.TimesheetEntry.ApprovalStatus.APPROVED

        # Map (task_id, day) -> hours pour les lignes "tâches"
        task_cell_map = {}
        for e in entry_qs.filter(task__isnull=False):
            key = (e.task_id, e.entry_date)
            task_cell_map[key] = (task_cell_map.get(key) or 0) + float(e.hours or 0)

        # Construit la grille des tâches
        grid = []
        day_totals = {d: 0.0 for d in days}
        for t in tasks_list:
            row = {
                "kind": "task",
                "task": t,
                "project": t.project,
                "label": t.title,
                "row_id": f"task-{t.id}",
                "cells": [],
                "total": 0.0,
            }
            for d in days:
                hours = float(task_cell_map.get((t.id, d)) or 0)
                row["cells"].append({"date": d, "hours": hours})
                row["total"] += hours
                day_totals[d] += hours
            grid.append(row)

        # ── 3. Activités libres (entries sans tâche) ────────────────
        # Regroupement par (project_id, description normalisée).
        free_groups = {}  # key -> dict
        for e in entry_qs.filter(task__isnull=True):
            desc = (e.description or "Activité").strip()
            key = (e.project_id, desc.lower())
            if key not in free_groups:
                free_groups[key] = {
                    "kind": "activity",
                    "project": e.project,
                    "label": desc,
                    "row_id": f"activity-{e.project_id or 0}-{abs(hash(desc.lower())) % 10**8}",
                    "description": desc,
                    "project_id": e.project_id,
                    "cells_map": {},
                    "total": 0.0,
                }
            grp = free_groups[key]
            grp["cells_map"][e.entry_date] = float(e.hours or 0)
            grp["total"] += float(e.hours or 0)
            day_totals[e.entry_date] = day_totals.get(e.entry_date, 0.0) + float(e.hours or 0)

        for grp in free_groups.values():
            cells = []
            for d in days:
                cells.append({"date": d, "hours": grp["cells_map"].get(d, 0.0)})
            grp["cells"] = cells
            del grp["cells_map"]
            grid.append(grp)

        week_total = sum(day_totals.values())

        # Projets disponibles pour la modale "Nouvelle activité"
        if ws:
            projects_qs = dm.Project.objects.filter(
                workspace=ws, is_archived=False
            ).order_by("name")
        else:
            projects_qs = dm.Project.objects.none()

        prev_week = monday - _td(days=7)
        next_week = monday + _td(days=7)
        today = timezone.localdate()

        ctx = {
            "section": "timesheet",
            "page_title": "Calendrier des heures",
            "current_workspace": ws,
            "monday": monday,
            "days": days,
            "days_meta": days_meta,
            "today": today,
            "grid": grid,
            "day_totals": day_totals,
            "week_total": week_total,
            "prev_week": prev_week,
            "next_week": next_week,
            "iso_week": monday.isocalendar()[1],
            "year": monday.year,
            "week_status": week_status,
            "week_status_label": dict(dm.TimesheetEntry.ApprovalStatus.choices).get(
                week_status, "Aucune saisie"
            ),
            "is_locked": is_locked,
            "projects": projects_qs,
        }
        return render(request, self.template_name, ctx)

    @staticmethod
    def _compute_week_status(entry_qs):
        """
        Retourne :
          APPROVED si toutes approved
          REJECTED si au moins une rejetée
          SUBMITTED si au moins une soumise
          DRAFT sinon
          "" si aucune entrée
        """
        statuses = set(entry_qs.values_list("approval_status", flat=True))
        if not statuses:
            return ""
        if statuses == {dm.TimesheetEntry.ApprovalStatus.APPROVED}:
            return dm.TimesheetEntry.ApprovalStatus.APPROVED
        if dm.TimesheetEntry.ApprovalStatus.REJECTED in statuses:
            return dm.TimesheetEntry.ApprovalStatus.REJECTED
        if dm.TimesheetEntry.ApprovalStatus.SUBMITTED in statuses:
            return dm.TimesheetEntry.ApprovalStatus.SUBMITTED
        return dm.TimesheetEntry.ApprovalStatus.DRAFT


class TimesheetCalendarSaveView(WorkspaceSecurityMixin, DevflowBaseMixin, View):
    """
    Endpoint AJAX appelé par le calendrier.
    POST { task_id, date (YYYY-MM-DD), hours (decimal) }
    Update or create la TimesheetEntry. Si hours == 0, supprime.
    Retourne JSON { success, hours, day_total, week_total }.
    """

    def post(self, request, *args, **kwargs):
        from datetime import date as _date, timedelta as _td
        from decimal import Decimal, InvalidOperation
        from django.db.models import Sum
        from django.http import JsonResponse

        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except Exception:
            payload = request.POST

        task_id = payload.get("task_id") or payload.get("task")
        project_id = payload.get("project_id") or payload.get("project")
        description = (payload.get("description") or "").strip()
        raw_date = payload.get("date")
        raw_hours = payload.get("hours")

        if not raw_date or (not task_id and not project_id):
            return JsonResponse({"success": False, "error": "Paramètres manquants."}, status=400)

        try:
            target_date = _date.fromisoformat(raw_date)
        except (TypeError, ValueError):
            return JsonResponse({"success": False, "error": "Date invalide."}, status=400)

        try:
            hours = Decimal(str(raw_hours or "0").replace(",", "."))
        except (InvalidOperation, TypeError):
            return JsonResponse({"success": False, "error": "Heures invalides."}, status=400)
        if hours < 0 or hours > 24:
            return JsonResponse({"success": False, "error": "Heures hors plage 0–24."}, status=400)

        # Verrou : si la semaine est APPROVED, on refuse toute modification
        monday = target_date - _td(days=target_date.weekday())
        sunday = monday + _td(days=6)
        week_qs = dm.TimesheetEntry.objects.filter(
            user=request.user, entry_date__gte=monday, entry_date__lte=sunday,
        )
        week_status_set = set(week_qs.values_list("approval_status", flat=True))
        if (
            week_status_set
            and week_status_set <= {dm.TimesheetEntry.ApprovalStatus.APPROVED}
        ):
            return JsonResponse(
                {"success": False, "error": "Cette semaine est validée et verrouillée."},
                status=403,
            )

        task = None
        project = None
        if task_id:
            task = (
                self.filter_by_workspace(dm.Task.objects.all())
                .select_related("project", "workspace")
                .filter(pk=task_id).first()
            )
            if not task:
                return JsonResponse({"success": False, "error": "Tâche introuvable."}, status=404)
            is_assigned = (
                task.assignee_id == request.user.id
                or dm.TaskAssignment.objects.filter(task=task, user=request.user).exists()
            )
            if not is_assigned:
                return JsonResponse(
                    {"success": False, "error": "Vous n'êtes pas affecté à cette tâche."},
                    status=403,
                )
            project = task.project
        else:
            project = (
                self.filter_by_workspace(dm.Project.objects.all())
                .filter(pk=project_id, is_archived=False).first()
            )
            if not project:
                return JsonResponse({"success": False, "error": "Projet introuvable."}, status=404)
            if not description:
                return JsonResponse(
                    {"success": False, "error": "Un libellé est requis pour une activité libre."},
                    status=400,
                )

        ws = (task.workspace if task else project.workspace) if (task or project) else None

        if hours == 0:
            qs_del = dm.TimesheetEntry.objects.filter(
                user=request.user, entry_date=target_date,
            )
            if task:
                qs_del = qs_del.filter(task=task)
            else:
                qs_del = qs_del.filter(task__isnull=True, project=project, description__iexact=description)
            qs_del.delete()
        else:
            lookup = {
                "user": request.user,
                "entry_date": target_date,
            }
            defaults = {
                "workspace": ws,
                "project": project,
                "hours": hours,
                "is_billable": True,
                "approval_status": dm.TimesheetEntry.ApprovalStatus.DRAFT,
            }
            if task:
                lookup["task"] = task
            else:
                lookup["task"] = None
                lookup["project"] = project
                lookup["description__iexact"] = description
                defaults["description"] = description

            # update_or_create ne supporte pas les __ dans lookup, on fait à la main
            existing = dm.TimesheetEntry.objects.filter(**lookup).first()
            if existing:
                for k, v in defaults.items():
                    setattr(existing, k, v)
                existing.save()
            else:
                # nettoyer le lookup pour create
                clean_lookup = {
                    k.replace("__iexact", ""): v
                    for k, v in lookup.items()
                }
                dm.TimesheetEntry.objects.create(**{**clean_lookup, **defaults})

        # Recalcul totaux
        week_qs = dm.TimesheetEntry.objects.filter(
            user=request.user, entry_date__gte=monday, entry_date__lte=sunday,
        )
        day_total = float(
            week_qs.filter(entry_date=target_date).aggregate(s=Sum("hours"))["s"] or 0
        )
        week_total = float(week_qs.aggregate(s=Sum("hours"))["s"] or 0)

        return JsonResponse({
            "success": True,
            "task_id": task.id if task else None,
            "project_id": project.id if project else None,
            "description": description or None,
            "date": target_date.isoformat(),
            "hours": float(hours),
            "day_total": day_total,
            "week_total": week_total,
        })


class DashboardSnapshotListView(DevflowListView):
    model = dm.DashboardSnapshot
    template_name = "project/dashboard_snapshot/list.html"
    section = "analytics"
    page_title = "Snapshots dashboard"


class DashboardSnapshotDetailView(DevflowDetailView):
    model = dm.DashboardSnapshot
    template_name = "project/dashboard_snapshot/detail.html"
    section = "analytics"
    page_title = "Détail snapshot"


class DashboardSnapshotCreateView(DevflowCreateView):
    model = dm.DashboardSnapshot
    form_class = DashboardSnapshotForm
    section = "analytics"
    page_title = "Créer snapshot"
    success_list_url_name = "dashboard_snapshot_list"


class DashboardSnapshotUpdateView(DevflowUpdateView):
    model = dm.DashboardSnapshot
    form_class = DashboardSnapshotForm
    section = "analytics"
    page_title = "Modifier snapshot"
    success_list_url_name = "dashboard_snapshot_list"


class DashboardSnapshotDeleteView(DevflowDeleteView):
    model = dm.DashboardSnapshot
    section = "analytics"
    page_title = "Supprimer snapshot"
    success_list_url_name = "dashboard_snapshot_list"


class UserPreferenceListView(DevflowListView):
    model = dm.UserPreference
    template_name = "project/user_preference/list.html"
    section = "settings"
    page_title = "Préférences utilisateur"
    search_fields = ("default_view",)


class UserPreferenceDetailView(DevflowDetailView):
    model = dm.UserPreference
    template_name = "project/user_preference/detail.html"
    section = "settings"
    page_title = "Détail préférence utilisateur"


class UserPreferenceCreateView(DevflowCreateView):
    model = dm.UserPreference
    form_class = UserPreferenceForm
    section = "settings"
    page_title = "Créer préférence utilisateur"
    success_list_url_name = "user_preference_list"


class UserPreferenceUpdateView(DevflowUpdateView):
    model = dm.UserPreference
    form_class = UserPreferenceForm
    section = "settings"
    page_title = "Modifier préférence utilisateur"
    success_list_url_name = "user_preference_list"


class UserPreferenceDeleteView(DevflowDeleteView):
    model = dm.UserPreference
    section = "settings"
    page_title = "Supprimer préférence utilisateur"
    success_list_url_name = "user_preference_list"


# =============================================================================
# SUPPORT MODELS
# =============================================================================
class LabelListView(DevflowListView):
    model = dm.Label
    template_name = "project/label/list.html"
    section = "settings"
    page_title = "Labels"
    search_fields = ("name", "description")


class LabelDetailView(DevflowDetailView):
    model = dm.Label
    template_name = "project/label/detail.html"
    section = "settings"
    page_title = "Détail label"


class LabelCreateView(DevflowCreateView):
    model = dm.Label
    form_class = LabelForm
    section = "settings"
    page_title = "Créer label"
    success_list_url_name = "label_list"


class LabelUpdateView(DevflowUpdateView):
    model = dm.Label
    form_class = LabelForm
    section = "settings"
    page_title = "Modifier label"
    success_list_url_name = "label_list"


class LabelDeleteView(DevflowDeleteView):
    model = dm.Label
    section = "settings"
    page_title = "Supprimer label"
    success_list_url_name = "label_list"


class TaskLabelListView(DevflowListView):
    model = dm.TaskLabel
    template_name = "project/task_label/list.html"
    section = "task"
    page_title = "Labels de tâches"
    search_fields = ("label__name", "task__title")


class TaskLabelDetailView(DevflowDetailView):
    model = dm.TaskLabel
    template_name = "project/task_label/detail.html"
    section = "task"
    page_title = "Détail label de tâche"


class TaskLabelCreateView(DevflowCreateView):
    model = dm.TaskLabel
    form_class = TaskLabelForm
    section = "task"
    page_title = "Créer label de tâche"
    success_list_url_name = "task_label_list"


class TaskLabelUpdateView(DevflowUpdateView):
    model = dm.TaskLabel
    form_class = TaskLabelForm
    section = "task"
    page_title = "Modifier label de tâche"
    success_list_url_name = "task_label_list"


class TaskLabelDeleteView(DevflowDeleteView):
    model = dm.TaskLabel
    section = "task"
    page_title = "Supprimer label de tâche"
    success_list_url_name = "task_label_list"


class ProjectLabelListView(DevflowListView):
    model = dm.ProjectLabel
    template_name = "project/project_label/list.html"
    section = "project"
    page_title = "Labels de projets"
    search_fields = ("label__name", "project__name")


class ProjectLabelDetailView(DevflowDetailView):
    model = dm.ProjectLabel
    template_name = "project/project_label/detail.html"
    section = "project"
    page_title = "Détail label de projet"


class ProjectLabelCreateView(DevflowCreateView):
    model = dm.ProjectLabel
    form_class = ProjectLabelForm
    section = "project"
    page_title = "Créer label de projet"
    success_list_url_name = "project_label_list"


class ProjectLabelUpdateView(DevflowUpdateView):
    model = dm.ProjectLabel
    form_class = ProjectLabelForm
    section = "project"
    page_title = "Modifier label de projet"
    success_list_url_name = "project_label_list"


class ProjectLabelDeleteView(DevflowDeleteView):
    model = dm.ProjectLabel
    section = "project"
    page_title = "Supprimer label de projet"
    success_list_url_name = "project_label_list"


class TaskDependencyListView(DevflowListView):
    model = dm.TaskDependency
    template_name = "project/task_dependency/list.html"
    section = "task"
    page_title = "Dépendances de tâches"
    search_fields = ("dependency_type", "from_task__title", "to_task__title")


class TaskDependencyDetailView(DevflowDetailView):
    model = dm.TaskDependency
    template_name = "project/task_dependency/detail.html"
    section = "task"
    page_title = "Détail dépendance tâche"


class TaskDependencyCreateView(DevflowCreateView):
    model = dm.TaskDependency
    form_class = TaskDependencyForm
    section = "task"
    page_title = "Créer dépendance tâche"
    success_list_url_name = "task_dependency_list"


class TaskDependencyUpdateView(DevflowUpdateView):
    model = dm.TaskDependency
    form_class = TaskDependencyForm
    section = "task"
    page_title = "Modifier dépendance tâche"
    success_list_url_name = "task_dependency_list"


class TaskDependencyDeleteView(DevflowDeleteView):
    model = dm.TaskDependency
    section = "task"
    page_title = "Supprimer dépendance tâche"
    success_list_url_name = "task_dependency_list"


class TaskChecklistListView(DevflowListView):
    model = dm.TaskChecklist
    template_name = "project/task_checklist/list.html"
    section = "task"
    page_title = "Checklists"
    search_fields = ("title", "task__title")


class TaskChecklistDetailView(DevflowDetailView):
    model = dm.TaskChecklist
    template_name = "project/task_checklist/detail.html"
    section = "task"
    page_title = "Détail checklist"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["items"] = self.object.items.all()
        return ctx


class TaskChecklistCreateView(DevflowCreateView):
    model = dm.TaskChecklist
    form_class = TaskChecklistForm
    section = "task"
    page_title = "Créer checklist"
    success_list_url_name = "task_checklist_list"


class TaskChecklistUpdateView(DevflowUpdateView):
    model = dm.TaskChecklist
    form_class = TaskChecklistForm
    section = "task"
    page_title = "Modifier checklist"
    success_list_url_name = "task_checklist_list"


class TaskChecklistDeleteView(DevflowDeleteView):
    model = dm.TaskChecklist
    section = "task"
    page_title = "Supprimer checklist"
    success_list_url_name = "task_checklist_list"


class ChecklistItemListView(DevflowListView):
    model = dm.ChecklistItem
    template_name = "project/checklist_item/list.html"
    section = "task"
    page_title = "Éléments de checklist"
    search_fields = ("text",)


class ChecklistItemDetailView(DevflowDetailView):
    model = dm.ChecklistItem
    template_name = "project/checklist_item/detail.html"
    section = "task"
    page_title = "Détail élément checklist"


class ChecklistItemCreateView(DevflowCreateView):
    model = dm.ChecklistItem
    form_class = ChecklistItemForm
    section = "task"
    page_title = "Créer élément checklist"
    success_list_url_name = "checklist_item_list"


class ChecklistItemUpdateView(DevflowUpdateView):
    model = dm.ChecklistItem
    form_class = ChecklistItemForm
    section = "task"
    page_title = "Modifier élément checklist"
    success_list_url_name = "checklist_item_list"


class ChecklistItemDeleteView(DevflowDeleteView):
    model = dm.ChecklistItem
    section = "task"
    page_title = "Supprimer élément checklist"
    success_list_url_name = "checklist_item_list"


class MilestoneListView(DevflowListView):
    model = dm.Milestone
    template_name = "project/milestone/list.html"
    section = "project"
    page_title = "Jalons"
    search_fields = ("name", "description", "status", "project__name", "owner__username")

    def get_queryset(self):
        today = timezone.localdate()
        return (
            dm.Milestone.objects.select_related("workspace", "project", "owner")
            .annotate(
                tasks_count=Count("milestone_tasks", distinct=True),
                completed_tasks_count=Count(
                    "milestone_tasks",
                    filter=Q(milestone_tasks__task__status=dm.Task.Status.DONE),
                    distinct=True,
                ),
                overdue_flag=Count(
                    "id",
                    filter=Q(due_date__lt=today) & ~Q(status=dm.Milestone.Status.DONE),
                    distinct=True,
                ),
            )
            .order_by("due_date", "id")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = ctx["object_list"]
        today = timezone.localdate()

        ctx.update({
            "today": today,
            "stats": qs.aggregate(
                total=Count("id"),
                planned=Count("id", filter=Q(status=dm.Milestone.Status.PLANNED)),
                in_progress=Count("id", filter=Q(status=dm.Milestone.Status.IN_PROGRESS)),
                at_risk=Count("id", filter=Q(status=dm.Milestone.Status.AT_RISK)),
                done=Count("id", filter=Q(status=dm.Milestone.Status.DONE)),
                missed=Count("id", filter=Q(status=dm.Milestone.Status.MISSED)),
            )
        })
        return ctx


class MilestoneDetailView(DevflowDetailView):
    model = dm.Milestone
    template_name = "project/milestone/detail.html"
    section = "project"
    page_title = "Détail jalon"

    def get_queryset(self):
        return (
            dm.Milestone.objects.select_related("workspace", "project", "owner")
            .prefetch_related(
                "milestone_tasks__task__assignee",
                "milestone_tasks__task__reporter",
                "milestone_tasks__task__sprint",
            )
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        milestone = self.object
        today = timezone.localdate()

        milestone_tasks = (
            milestone.milestone_tasks
            .select_related("task", "task__assignee", "task__reporter", "task__sprint")
            .all()
        )

        tasks_qs = dm.Task.objects.filter(
            milestones__milestone=milestone,
            is_archived=False,
        ).select_related("assignee", "reporter", "sprint")

        task_stats = tasks_qs.aggregate(
            total=Count("id"),
            todo=Count("id", filter=Q(status=dm.Task.Status.TODO)),
            in_progress=Count("id", filter=Q(status=dm.Task.Status.IN_PROGRESS)),
            review=Count("id", filter=Q(status=dm.Task.Status.REVIEW)),
            blocked=Count("id", filter=Q(status=dm.Task.Status.BLOCKED)),
            done=Count("id", filter=Q(status=dm.Task.Status.DONE)),
        )

        ctx.update({
            "today": today,
            "milestone_tasks": milestone_tasks,
            "linked_tasks": tasks_qs,
            "task_stats": task_stats,
            "is_overdue": milestone.due_date < today and milestone.status != dm.Milestone.Status.DONE,
        })
        return ctx


class MilestoneCreateView(DevflowCreateView):
    model = dm.Milestone
    form_class = MilestoneForm
    template_name = "project/milestone/form.html"
    section = "project"
    page_title = "Créer jalon"
    success_message = "Jalon créé avec succès."
    success_list_url_name = "project:milestone_list"

    def get_initial(self):
        initial = super().get_initial()
        project_id = self.request.GET.get("project")
        workspace = self.get_workspace()

        if project_id:
            initial["project"] = project_id

        if workspace:
            initial["workspace"] = workspace.pk

        return initial

    def get_workspace(self):
        workspace = getattr(self.request, "workspace", None)

        if not workspace:
            workspace = getattr(getattr(self.request.user, "profile", None), "workspace", None)

        return workspace

    def form_valid(self, form):
        workspace = form.cleaned_data.get("workspace") or self.get_workspace()
        project = form.cleaned_data.get("project")

        if not workspace:
            form.add_error("workspace", "Veuillez sélectionner un workspace.")
            form.add_error(None, "Aucun workspace actif n'est associé à cette création de jalon.")
            return self.form_invalid(form)

        if not project:
            form.add_error("project", "Veuillez sélectionner un projet.")
            return self.form_invalid(form)

        if project.workspace_id != workspace.id:
            form.add_error("project", "Le projet sélectionné n'appartient pas au workspace choisi.")
            return self.form_invalid(form)

        try:
            with transaction.atomic():
                obj = form.save(commit=False)
                obj.workspace = workspace
                obj.project = project

                if not obj.owner_id and self.request.user.is_authenticated:
                    obj.owner = self.request.user

                if obj.progress_percent is None:
                    obj.progress_percent = 0

                obj.save()
                form.save_m2m()
                self.object = obj

        except ValidationError as e:
            form.add_error(None, "; ".join(e.messages))
            return self.form_invalid(form)

        except Exception as e:
            form.add_error(None, f"Erreur lors de l'enregistrement : {e}")
            return self.form_invalid(form)

        messages.success(self.request, self.success_message)
        return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy("milestone_list")

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class MilestoneUpdateView(DevflowUpdateView):
    model = dm.Milestone
    form_class = MilestoneForm
    template_name = "project/milestone/form.html"
    section = "project"
    page_title = "Modifier jalon"
    success_list_url_name = "project:milestone_list"

    def form_valid(self, form):
        messages.success(self.request, "Jalon modifié avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class MilestoneDeleteView(DevflowDeleteView):
    model = dm.Milestone
    section = "project"
    page_title = "Supprimer jalon"
    success_list_url_name = "project:milestone_list"


class MilestoneArchiveView(ArchiveObjectView):
    model = dm.Milestone
    success_list_url_name = "project:milestone_list"


class MilestoneTaskListView(DevflowListView):
    model = dm.MilestoneTask
    template_name = "project/milestone_task/list.html"
    section = "project"
    page_title = "Tâches par jalon"
    search_fields = ("task__title", "milestone__name", "milestone__project__name")

    def get_queryset(self):
        return (
            dm.MilestoneTask.objects
            .select_related(
                "milestone",
                "milestone__project",
                "task",
                "task__assignee",
                "task__reporter",
                "task__sprint",
            )
            .order_by("milestone__due_date", "milestone__name", "task__title")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = ctx["object_list"]

        ctx.update({
            "stats": qs.aggregate(
                total=Count("id"),
                unique_milestones=Count("milestone", distinct=True),
                unique_tasks=Count("task", distinct=True),
                done_tasks=Count("id", filter=Q(task__status=dm.Task.Status.DONE)),
                blocked_tasks=Count("id", filter=Q(task__status=dm.Task.Status.BLOCKED)),
            )
        })
        return ctx


class MilestoneTaskDetailView(DevflowDetailView):
    model = dm.MilestoneTask
    template_name = "project/milestone_task/detail.html"
    section = "project"
    page_title = "Détail tâche jalon"

    def get_queryset(self):
        return (
            dm.MilestoneTask.objects
            .select_related(
                "milestone",
                "milestone__project",
                "milestone__owner",
                "task",
                "task__assignee",
                "task__reporter",
                "task__sprint",
                "task__backlog_item",
            )
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        obj = self.object
        ctx.update({
            "is_task_overdue": bool(
                obj.task.due_date and obj.task.due_date < timezone.localdate() and obj.task.status != dm.Task.Status.DONE),
            "is_milestone_overdue": obj.milestone.due_date < timezone.localdate() and obj.milestone.status != dm.Milestone.Status.DONE,
        })
        return ctx


class MilestoneTaskCreateView(DevflowCreateView):
    model = dm.MilestoneTask
    form_class = MilestoneTaskForm
    template_name = "project/milestone_task/form.html"
    section = "project"
    page_title = "Créer tâche jalon"
    success_list_url_name = "milestone_task_list"

    def get_initial(self):
        initial = super().get_initial()
        milestone_id = self.request.GET.get("milestone")
        task_id = self.request.GET.get("task")
        if milestone_id:
            initial["milestone"] = milestone_id
        if task_id:
            initial["task"] = task_id
        return initial

    def form_valid(self, form):
        messages.success(self.request, "Tâche liée au jalon avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class MilestoneTaskUpdateView(DevflowUpdateView):
    model = dm.MilestoneTask
    form_class = MilestoneTaskForm
    template_name = "project/milestone_task/form.html"
    section = "project"
    page_title = "Modifier tâche jalon"
    success_list_url_name = "milestone_task_list"

    def form_valid(self, form):
        messages.success(self.request, "Liaison tâche / jalon modifiée avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class MilestoneTaskDeleteView(DevflowDeleteView):
    model = dm.MilestoneTask
    section = "project"
    page_title = "Supprimer tâche jalon"
    success_list_url_name = "project:milestone_task_list"


class ReleaseListView(DevflowListView):
    model = dm.Release
    template_name = "project/release/list.html"
    section = "project"
    page_title = "Releases"
    search_fields = ("name", "tag", "description", "changelog", "project__name")

    def get_queryset(self):
        today = timezone.localdate()
        return (
            dm.Release.objects
            .select_related("workspace", "project")
            .prefetch_related("tasks", "sprints")
            .annotate(
                tasks_count=Count("tasks", distinct=True),
                sprints_count=Count("sprints", distinct=True),
                done_tasks_count=Count(
                    "tasks",
                    filter=Q(tasks__status=dm.Task.Status.DONE),
                    distinct=True,
                ),
            )
            .order_by("-release_date", "-id")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = ctx["object_list"]
        today = timezone.localdate()

        ctx.update({
            "today": today,
            "stats": qs.aggregate(
                total=Count("id"),
                planned=Count("id", filter=Q(status=dm.Release.Status.PLANNED)),
                in_progress=Count("id", filter=Q(status=dm.Release.Status.IN_PROGRESS)),
                released=Count("id", filter=Q(status=dm.Release.Status.RELEASED)),
                cancelled=Count("id", filter=Q(status=dm.Release.Status.CANCELLED)),
            ),
        })
        return ctx


class ReleaseDetailView(DevflowDetailView):
    model = dm.Release
    template_name = "project/release/detail.html"
    section = "project"
    page_title = "Détail release"

    def get_queryset(self):
        return (
            dm.Release.objects
            .select_related("workspace", "project")
            .prefetch_related(
                "tasks__assignee",
                "tasks__reporter",
                "tasks__sprint",
                "sprints__team",
            )
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        release = self.object
        today = timezone.localdate()

        tasks_qs = release.tasks.select_related(
            "assignee", "reporter", "sprint"
        ).filter(is_archived=False)

        sprints_qs = release.sprints.select_related("team").filter(is_archived=False)

        task_stats = tasks_qs.aggregate(
            total=Count("id"),
            todo=Count("id", filter=Q(status=dm.Task.Status.TODO)),
            in_progress=Count("id", filter=Q(status=dm.Task.Status.IN_PROGRESS)),
            review=Count("id", filter=Q(status=dm.Task.Status.REVIEW)),
            blocked=Count("id", filter=Q(status=dm.Task.Status.BLOCKED)),
            done=Count("id", filter=Q(status=dm.Task.Status.DONE)),
        )

        ctx.update({
            "today": today,
            "release_tasks": tasks_qs,
            "release_sprints": sprints_qs,
            "task_stats": task_stats,
            "is_late_release": (
                    release.release_date
                    and release.release_date < today
                    and release.status != dm.Release.Status.RELEASED
                    and release.status != dm.Release.Status.CANCELLED
            ),
        })
        return ctx


class ReleaseCreateView(DevflowCreateView):
    model = dm.Release
    form_class = ReleaseForm
    template_name = "project/release/form.html"
    section = "project"
    page_title = "Créer release"
    success_list_url_name = "release_list"

    def form_valid(self, form):
        messages.success(self.request, "Release créée avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class ReleaseUpdateView(DevflowUpdateView):
    model = dm.Release
    form_class = ReleaseForm
    template_name = "project/release/form.html"
    section = "project"
    page_title = "Modifier release"
    success_list_url_name = "release_list"

    def form_valid(self, form):
        messages.success(self.request, "Release modifiée avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class ReleaseDeleteView(DevflowDeleteView):
    model = dm.Release
    section = "project"
    page_title = "Supprimer release"
    success_list_url_name = "release_list"


class ReleaseArchiveView(ArchiveObjectView):
    model = dm.Release
    success_list_url_name = "release_list"


class RoadmapListView(DevflowListView):
    model = dm.Roadmap
    template_name = "project/roadmap/list.html"
    section = "project"
    page_title = "Roadmaps"
    search_fields = ("name", "description", "owner__username", "owner__first_name", "owner__last_name")

    def get_queryset(self):
        return (
            dm.Roadmap.objects.select_related("workspace", "owner")
            .annotate(
                items_count=Count("items", distinct=True),
                planned_items_count=Count(
                    "items",
                    filter=Q(items__status=dm.RoadmapItem.ItemStatus.PLANNED),
                    distinct=True,
                ),
                in_progress_items_count=Count(
                    "items",
                    filter=Q(items__status=dm.RoadmapItem.ItemStatus.IN_PROGRESS),
                    distinct=True,
                ),
                done_items_count=Count(
                    "items",
                    filter=Q(items__status=dm.RoadmapItem.ItemStatus.DONE),
                    distinct=True,
                ),
                at_risk_items_count=Count(
                    "items",
                    filter=Q(items__status=dm.RoadmapItem.ItemStatus.AT_RISK),
                    distinct=True,
                ),
                first_item_start=Min("items__start_date"),
                last_item_end=Max("items__end_date"),
            )
            .order_by("-start_date", "-id")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = ctx["object_list"]
        today = timezone.localdate()

        ctx.update({
            "stats": qs.aggregate(
                total=Count("id"),
                public_count=Count("id", filter=Q(is_public=True)),
                private_count=Count("id", filter=Q(is_public=False)),
                total_items=Coalesce(Count("items"), 0),
            ),
            "today": today,
        })
        return ctx


# class RoadmapDetailView(DevflowDetailView):
#     model = dm.Roadmap
#     template_name = "project/roadmap/detail.html"
#     section = "project"
#     page_title = "Détail roadmap"
#
#     def get_queryset(self):
#         return (
#             dm.Roadmap.objects.select_related("workspace", "owner")
#             .prefetch_related("items__project", "items__milestone")
#         )
#
#     def get_context_data(self, **kwargs):
#         ctx = super().get_context_data(**kwargs)
#         roadmap = self.object
#         today = timezone.localdate()
#
#         items_qs = (
#             roadmap.items.select_related("project", "milestone")
#             .order_by("row", "start_date", "id")
#         )
#
#         ctx.update({
#             "items": items_qs,
#             "items_by_row": {},
#             "stats": items_qs.aggregate(
#                 total=Count("id"),
#                 planned=Count("id", filter=Q(status=dm.RoadmapItem.ItemStatus.PLANNED)),
#                 in_progress=Count("id", filter=Q(status=dm.RoadmapItem.ItemStatus.IN_PROGRESS)),
#                 done=Count("id", filter=Q(status=dm.RoadmapItem.ItemStatus.DONE)),
#                 at_risk=Count("id", filter=Q(status=dm.RoadmapItem.ItemStatus.AT_RISK)),
#                 projects_count=Count("project", distinct=True),
#                 milestones_count=Count("milestone", distinct=True),
#                 rows_count=Count("row", distinct=True),
#             ),
#             "today": today,
#             "timeline_start": items_qs.aggregate(dt=Min("start_date"))["dt"] or roadmap.start_date,
#             "timeline_end": items_qs.aggregate(dt=Max("end_date"))["dt"] or roadmap.end_date,
#         })
#
#         grouped = {}
#         for item in items_qs:
#             grouped.setdefault(item.row, []).append(item)
#         ctx["items_by_row"] = grouped
#
#         return ctx
@login_required
@require_POST
def roadmap_item_shift_dates(request):
    try:
        payload = json.loads(request.body.decode("utf-8"))
        item_id = payload.get("item_id")
        delta_days = int(payload.get("delta_days", 0))

        roadmap_item = get_object_or_404(dm.RoadmapItem, pk=item_id)

        if delta_days == 0:
            return JsonResponse({"success": True, "message": "Aucun changement."})

        new_start = roadmap_item.start_date + timedelta(days=delta_days)
        new_end = roadmap_item.end_date + timedelta(days=delta_days)

        roadmap = roadmap_item.roadmap
        if new_start < roadmap.start_date or new_end > roadmap.end_date:
            return JsonResponse(
                {
                    "success": False,
                    "message": "Le déplacement sort de la période autorisée de la roadmap."
                },
                status=400
            )

        roadmap_item.start_date = new_start
        roadmap_item.end_date = new_end
        roadmap_item.full_clean()
        roadmap_item.save(update_fields=["start_date", "end_date", "updated_at"])

        return JsonResponse({
            "success": True,
            "item_id": roadmap_item.pk,
            "start_date": roadmap_item.start_date.strftime("%Y-%m-%d"),
            "end_date": roadmap_item.end_date.strftime("%Y-%m-%d"),
            "start_date_display": roadmap_item.start_date.strftime("%d/%m/%Y"),
            "end_date_display": roadmap_item.end_date.strftime("%d/%m/%Y"),
        })

    except Exception as exc:
        return JsonResponse(
            {"success": False, "message": str(exc)},
            status=400
        )


class RoadmapDetailView(DevflowDetailView):
    model = dm.Roadmap
    template_name = "project/roadmap/detail.html"
    section = "project"
    page_title = "Détail roadmap"

    def get_queryset(self):
        return (
            dm.Roadmap.objects.select_related("workspace", "owner")
            .prefetch_related("items__project", "items__milestone")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        roadmap = self.object
        today = timezone.localdate()

        items_qs = (
            roadmap.items.select_related("project", "milestone")
            .order_by("row", "start_date", "id")
        )

        timeline_start = items_qs.aggregate(dt=Min("start_date"))["dt"] or roadmap.start_date
        timeline_end = items_qs.aggregate(dt=Max("end_date"))["dt"] or roadmap.end_date
        timeline_days = max((timeline_end - timeline_start).days + 1, 1)

        grouped = OrderedDict()
        for item in items_qs:
            start_offset_days = max((item.start_date - timeline_start).days, 0)
            duration_days = max((item.end_date - item.start_date).days + 1, 1)

            item.gantt_left_pct = (start_offset_days / timeline_days) * 100
            item.gantt_width_pct = max((duration_days / timeline_days) * 100, 2.2)
            item.gantt_duration_days = duration_days
            item.is_overdue = item.end_date < today and item.status != dm.RoadmapItem.ItemStatus.DONE

            grouped.setdefault(item.row, []).append(item)

        month_markers = []
        cursor = timeline_start.replace(day=1)
        while cursor <= timeline_end:
            next_month = (cursor.replace(day=28) + timedelta(days=4)).replace(day=1)
            segment_end = min(timeline_end, next_month - timedelta(days=1))
            days_in_segment = (segment_end - cursor).days + 1
            month_markers.append({
                "label": cursor.strftime("%b %Y"),
                "days": days_in_segment,
                "width_pct": (days_in_segment / timeline_days) * 100,
            })
            cursor = next_month

        quarter_markers = []
        quarter_cursor = timeline_start.replace(day=1)
        while quarter_cursor <= timeline_end:
            quarter = ((quarter_cursor.month - 1) // 3) + 1
            q_start_month = ((quarter - 1) * 3) + 1
            q_start = quarter_cursor.replace(month=q_start_month, day=1)

            if q_start < timeline_start.replace(day=1):
                q_start = quarter_cursor

            q_end_month = q_start_month + 2
            year = q_start.year
            if q_end_month > 12:
                q_end_month -= 12
                year += 1

            tmp = q_start.replace(year=year, month=q_end_month, day=28) + timedelta(days=4)
            q_end = tmp.replace(day=1) - timedelta(days=1)
            segment_start = max(q_start, timeline_start)
            segment_end = min(q_end, timeline_end)

            if segment_start <= segment_end:
                days_in_segment = (segment_end - segment_start).days + 1
                quarter_markers.append({
                    "label": f"T{quarter} {segment_start.year}",
                    "days": days_in_segment,
                    "width_pct": (days_in_segment / timeline_days) * 100,
                })

            quarter_cursor = q_end + timedelta(days=1)

        if timeline_start <= today <= timeline_end:
            today_offset_days = (today - timeline_start).days
            today_left_pct = (today_offset_days / timeline_days) * 100
        else:
            today_left_pct = None

        ctx.update({
            "items": items_qs,
            "items_by_row": grouped,
            "stats": items_qs.aggregate(
                total=Count("id"),
                planned=Count("id", filter=Q(status=dm.RoadmapItem.ItemStatus.PLANNED)),
                in_progress=Count("id", filter=Q(status=dm.RoadmapItem.ItemStatus.IN_PROGRESS)),
                done=Count("id", filter=Q(status=dm.RoadmapItem.ItemStatus.DONE)),
                at_risk=Count("id", filter=Q(status=dm.RoadmapItem.ItemStatus.AT_RISK)),
                projects_count=Count("project", distinct=True),
                milestones_count=Count("milestone", distinct=True),
                rows_count=Count("row", distinct=True),
            ),
            "today": today,
            "timeline_start": timeline_start,
            "timeline_end": timeline_end,
            "timeline_days": timeline_days,
            "month_markers": month_markers,
            "quarter_markers": quarter_markers,
            "today_left_pct": today_left_pct,
        })
        return ctx


class RoadmapCreateView(DevflowCreateView):
    model = dm.Roadmap
    form_class = RoadmapForm
    template_name = "project/roadmap/form.html"
    section = "project"
    page_title = "Créer roadmap"
    success_list_url_name = "roadmap_list"

    def get_initial(self):
        initial = super().get_initial()
        today = timezone.localdate()
        initial.setdefault("start_date", today)
        initial.setdefault("end_date", today)
        return initial

    def form_valid(self, form):
        messages.success(self.request, "Roadmap créée avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class RoadmapUpdateView(DevflowUpdateView):
    model = dm.Roadmap
    form_class = RoadmapForm
    section = "project"
    template_name = "project/roadmap/form.html"
    page_title = "Modifier roadmap"
    success_list_url_name = "project:roadmap_list"

    def form_valid(self, form):
        messages.success(self.request, "Roadmap modifiée avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class RoadmapDeleteView(DevflowDeleteView):
    model = dm.Roadmap
    section = "project"
    page_title = "Supprimer roadmap"
    success_list_url_name = "roadmap_list"


class RoadmapArchiveView(ArchiveObjectView):
    model = dm.Roadmap
    success_list_url_name = "roadmap_list"


class RoadmapItemListView(DevflowListView):
    model = dm.RoadmapItem
    template_name = "project/roadmap_item/list.html"
    section = "project"
    page_title = "Éléments roadmap"
    search_fields = (
        "title",
        "status",
        "roadmap__name",
        "project__name",
        "milestone__name",
    )

    def get_queryset(self):
        return (
            dm.RoadmapItem.objects.select_related(
                "roadmap",
                "project",
                "milestone",
                "roadmap__owner",
                "roadmap__workspace",
            )
            .order_by("roadmap", "row", "start_date", "id")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = ctx["object_list"]
        today = timezone.localdate()

        ctx.update({
            "stats": qs.aggregate(
                total=Count("id"),
                planned=Count("id", filter=Q(status=dm.RoadmapItem.ItemStatus.PLANNED)),
                in_progress=Count("id", filter=Q(status=dm.RoadmapItem.ItemStatus.IN_PROGRESS)),
                done=Count("id", filter=Q(status=dm.RoadmapItem.ItemStatus.DONE)),
                at_risk=Count("id", filter=Q(status=dm.RoadmapItem.ItemStatus.AT_RISK)),
                roadmaps_count=Count("roadmap", distinct=True),
            ),
            "today": today,
        })
        return ctx


class RoadmapItemDetailView(DevflowDetailView):
    model = dm.RoadmapItem
    template_name = "project/roadmap_item/detail.html"
    section = "project"
    page_title = "Détail élément roadmap"

    def get_queryset(self):
        return (
            dm.RoadmapItem.objects.select_related(
                "roadmap",
                "project",
                "milestone",
                "roadmap__owner",
                "roadmap__workspace",
            )
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        item = self.object
        today = timezone.localdate()

        ctx.update({
            "today": today,
            "is_overdue": item.end_date < today and item.status != dm.RoadmapItem.ItemStatus.DONE,
            "duration_days": (item.end_date - item.start_date).days + 1,
            "related_items": (
                item.roadmap.items.exclude(pk=item.pk)
                .select_related("project", "milestone")
                .order_by("row", "start_date")[:8]
            ),
        })
        return ctx


class RoadmapItemCreateView(DevflowCreateView):
    model = dm.RoadmapItem
    form_class = RoadmapItemForm
    section = "project"
    template_name = "project/roadmap_item/form.html"
    page_title = "Créer élément roadmap"
    success_list_url_name = "roadmap_item_list"

    def form_valid(self, form):
        messages.success(self.request, "Élément roadmap créé avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class RoadmapItemUpdateView(DevflowUpdateView):
    model = dm.RoadmapItem
    form_class = RoadmapItemForm
    section = "project"
    template_name = "project/roadmap_item/form.html"
    page_title = "Modifier élément roadmap"
    success_list_url_name = "project:roadmap_item_list"

    def form_valid(self, form):
        messages.success(self.request, "Élément roadmap modifié avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class RoadmapItemDeleteView(DevflowDeleteView):
    model = dm.RoadmapItem
    section = "project"
    page_title = "Supprimer élément roadmap"
    success_list_url_name = "roadmap_item_list"


class BoardColumnListView(DevflowListView):
    model = dm.BoardColumn
    template_name = "project/board_column/list.html"
    section = "project"
    page_title = "Colonnes Kanban"
    search_fields = ("name", "mapped_status")


class BoardColumnDetailView(DevflowDetailView):
    model = dm.BoardColumn
    template_name = "project/board_column/detail.html"
    section = "project"
    page_title = "Détail colonne Kanban"


class BoardColumnCreateView(DevflowCreateView):
    model = dm.BoardColumn
    form_class = BoardColumnForm
    section = "project"
    page_title = "Créer colonne Kanban"
    success_list_url_name = "board_column_list"


class BoardColumnUpdateView(DevflowUpdateView):
    model = dm.BoardColumn
    form_class = BoardColumnForm
    section = "project"
    page_title = "Modifier colonne Kanban"
    success_list_url_name = "board_column_list"


class BoardColumnDeleteView(DevflowDeleteView):
    model = dm.BoardColumn
    section = "project"
    page_title = "Supprimer colonne Kanban"
    success_list_url_name = "board_column_list"


class WorkspaceInvitationListView(DevflowListView):
    model = dm.WorkspaceInvitation
    template_name = "project/workspace_invitation/list.html"
    section = "workspace"
    page_title = "Invitations workspace"
    search_placeholder = "Rechercher par email…"
    search_fields = ("email", "invited_by__username", "invited_by__first_name", "invited_by__last_name")
    filter_fields = [
        {"name": "status", "label": "Statut", "type": "choices",
         "choices": dm.WorkspaceInvitation.Status.choices},
        {"name": "role", "label": "Rôle", "type": "choices",
         "choices": dm.TeamMembership.Role.choices},
        {"name": "team", "label": "Équipe", "type": "model",
         "queryset": "_teams_qs", "lookup": "team_id"},
    ]

    def _teams_qs(self):
        ws = self.get_current_workspace()
        return dm.Team.objects.filter(workspace=ws, is_archived=False).order_by("name") if ws else dm.Team.objects.none()

    def get_queryset(self):
        return super().get_queryset().select_related("invited_by", "team", "workspace")


class WorkspaceInvitationDetailView(DevflowDetailView):
    model = dm.WorkspaceInvitation
    template_name = "project/workspace_invitation/detail.html"
    section = "workspace"
    page_title = "Détail invitation workspace"


class WorkspaceInvitationCreateView(DevflowCreateView):
    model = dm.WorkspaceInvitation
    form_class = WorkspaceInvitationForm
    template_name = "project/workspace_invitation/form.html"
    section = "workspace"
    page_title = "Inviter un membre"
    success_list_url_name = "workspace_invitation_list"

    def get_initial(self):
        initial = super().get_initial()
        team_id = self.request.GET.get("team")
        role = self.request.GET.get("role")
        email = self.request.GET.get("email")
        if team_id:
            initial["team"] = team_id
        if role:
            initial["role"] = role
        if email:
            initial["email"] = email
        return initial

    def form_valid(self, form):
        invitation = form.save(
            invited_by=self.request.user,
            workspace=self.get_current_workspace(),
        )
        try:
            from project.services.invitations import send_invitation_email
            send_invitation_email(invitation, request=self.request)
        except Exception as exc:
            messages.warning(
                self.request,
                f"Invitation enregistrée mais l'e-mail n'a pas pu être envoyé : {exc}"
            )
        else:
            messages.success(
                self.request,
                f"Invitation envoyée à {invitation.email}."
            )
        return redirect(self.get_success_url())


class WorkspaceInvitationResendView(WorkspaceSecurityMixin, DevflowBaseMixin, View):
    """Renvoie l'email d'invitation pour une invitation PENDING."""

    def post(self, request, pk):
        from datetime import timedelta as _td
        invitation = self.filter_by_workspace(
            dm.WorkspaceInvitation.objects.all()
        ).filter(pk=pk).first()
        if not invitation:
            messages.error(request, "Invitation introuvable.")
            return redirect("workspace_invitation_list")
        if invitation.status != dm.WorkspaceInvitation.Status.PENDING:
            messages.warning(request, "Seules les invitations en attente peuvent être renvoyées.")
            return redirect("workspace_invitation_list")

        # Prolonge l'expiration si proche / dépassée
        if not invitation.expires_at or invitation.expires_at <= timezone.now() + _td(days=1):
            invitation.expires_at = timezone.now() + _td(days=14)
            invitation.save(update_fields=["expires_at", "updated_at"])

        try:
            from project.services.invitations import send_invitation_email
            send_invitation_email(invitation, request=request)
            messages.success(request, f"Invitation renvoyée à {invitation.email}.")
        except Exception as exc:
            messages.error(request, f"Échec d'envoi : {exc}")
        return redirect("workspace_invitation_list")


class WorkspaceInvitationRevokeView(WorkspaceSecurityMixin, DevflowBaseMixin, View):
    """Révoque une invitation (PENDING → REVOKED)."""

    def post(self, request, pk):
        invitation = self.filter_by_workspace(
            dm.WorkspaceInvitation.objects.all()
        ).filter(pk=pk).first()
        if not invitation:
            messages.error(request, "Invitation introuvable.")
            return redirect("workspace_invitation_list")
        if invitation.status != dm.WorkspaceInvitation.Status.PENDING:
            messages.warning(request, "Seules les invitations en attente peuvent être révoquées.")
            return redirect("workspace_invitation_list")
        invitation.status = dm.WorkspaceInvitation.Status.REVOKED
        invitation.save(update_fields=["status", "updated_at"])
        messages.success(request, f"Invitation pour {invitation.email} révoquée.")
        return redirect("workspace_invitation_list")


class WorkspaceInvitationUpdateView(DevflowUpdateView):
    model = dm.WorkspaceInvitation
    form_class = WorkspaceInvitationForm
    template_name = "project/workspace_invitation/form.html"
    section = "workspace"
    page_title = "Modifier invitation workspace"
    success_list_url_name = "workspace_invitation_list"


class WorkspaceInvitationDeleteView(DevflowDeleteView):
    model = dm.WorkspaceInvitation
    section = "workspace"
    page_title = "Supprimer invitation workspace"
    success_list_url_name = "workspace_invitation_list"


class WorkspaceInvitationAcceptView(DevflowBaseMixin, View):
    """Accept admin-side (réservée aux managers du workspace)."""

    def post(self, request, pk):
        invitation = self.filter_by_workspace(
            dm.WorkspaceInvitation.objects.all()
        ).get(pk=pk)
        invitation.status = dm.WorkspaceInvitation.Status.ACCEPTED
        invitation.accepted_at = timezone.now()
        invitation.save(update_fields=["status", "accepted_at", "updated_at"])
        messages.success(request, "Invitation acceptée.")
        return redirect("workspace_invitation_list")


class WorkspaceInvitationPublicAcceptView(View):
    """
    Vue publique : un invité clique sur le lien de l'email reçu.
    GET → page d'accueil (présentation de l'invitation + form de signup si l'user n'existe pas).
    POST → finalise : crée le User si nécessaire, le UserProfile et le TeamMembership.
    """

    template_name = "project/workspace_invitation/accept.html"

    def get_invitation_or_404(self, token):
        invitation = get_object_or_404(
            dm.WorkspaceInvitation.objects.select_related("workspace", "team"),
            token=token,
        )
        if invitation.status != dm.WorkspaceInvitation.Status.PENDING:
            return invitation, "not_pending"
        if invitation.is_expired():
            invitation.status = dm.WorkspaceInvitation.Status.EXPIRED
            invitation.save(update_fields=["status", "updated_at"])
            return invitation, "expired"
        return invitation, "ok"

    def get(self, request, token):
        invitation, state = self.get_invitation_or_404(token)
        existing_user = User.objects.filter(email__iexact=invitation.email).first()
        return render(request, self.template_name, {
            "invitation": invitation,
            "state": state,
            "existing_user": existing_user,
        })

    def post(self, request, token):
        invitation, state = self.get_invitation_or_404(token)
        if state != "ok":
            return render(request, self.template_name, {
                "invitation": invitation,
                "state": state,
            })

        from django.db import transaction
        from django.contrib.auth import login

        existing_user = User.objects.filter(email__iexact=invitation.email).first()
        with transaction.atomic():
            if existing_user:
                user = existing_user
            else:
                first_name = (request.POST.get("first_name") or "").strip()
                last_name = (request.POST.get("last_name") or "").strip()
                password = request.POST.get("password") or ""
                if not (first_name and last_name and len(password) >= 8):
                    messages.error(
                        request,
                        "Renseigne prénom, nom et un mot de passe d'au moins 8 caractères."
                    )
                    return render(request, self.template_name, {
                        "invitation": invitation, "state": "ok",
                        "form_errors": True,
                    })
                username = invitation.email.split("@")[0][:30]
                # Évite les collisions de username
                base_username = username
                idx = 1
                while User.objects.filter(username=username).exists():
                    username = f"{base_username}{idx}"
                    idx += 1
                user = User(
                    username=username,
                    email=invitation.email,
                    first_name=first_name,
                    last_name=last_name,
                )
                user.set_password(password)
                # Posé pour le signal create_user_profile
                user._invited_workspace = invitation.workspace
                user.save()

            # UserProfile
            dm.UserProfile.objects.get_or_create(
                user=user, workspace=invitation.workspace,
            )
            # TeamMembership
            dm.TeamMembership.objects.update_or_create(
                workspace=invitation.workspace,
                user=user,
                team=invitation.team,
                defaults={
                    "role": invitation.role,
                    "status": dm.TeamMembership.Status.ACTIVE,
                },
            )
            # Notification au demandeur
            if invitation.invited_by_id:
                try:
                    from project.services.notifications import (
                        create_in_app_notification,
                    )
                    create_in_app_notification(
                        recipient=invitation.invited_by,
                        workspace=invitation.workspace,
                        notification_type=dm.Notification.NotificationType.INFO
                        if hasattr(dm.Notification.NotificationType, "INFO")
                        else dm.Notification.NotificationType.TASK,
                        title="Invitation acceptée",
                        body=f"{user} a accepté votre invitation au workspace.",
                        url="/workspace-invitations/",
                    )
                except Exception:
                    pass
            invitation.status = dm.WorkspaceInvitation.Status.ACCEPTED
            invitation.accepted_at = timezone.now()
            invitation.save(update_fields=["status", "accepted_at", "updated_at"])

        if not existing_user:
            login(request, user, backend="django.contrib.auth.backends.ModelBackend")

        messages.success(
            request,
            f"Bienvenue dans {invitation.workspace.name} !"
        )
        return redirect("dashboard")


class IntegrationListView(DevflowListView):
    model = dm.Integration
    template_name = "project/integration/list.html"
    section = "settings"
    page_title = "Intégrations"
    search_fields = ("provider", "name", "status", "error_message")


class IntegrationDetailView(DevflowDetailView):
    model = dm.Integration
    template_name = "project/integration/detail.html"
    section = "settings"
    page_title = "Détail intégration"


class IntegrationCreateView(DevflowCreateView):
    model = dm.Integration
    form_class = IntegrationForm
    section = "settings"
    page_title = "Créer intégration"
    success_list_url_name = "integration_list"


class IntegrationUpdateView(DevflowUpdateView):
    model = dm.Integration
    form_class = IntegrationForm
    section = "settings"
    page_title = "Modifier intégration"
    success_list_url_name = "integration_list"


class IntegrationDeleteView(DevflowDeleteView):
    model = dm.Integration
    section = "settings"
    page_title = "Supprimer intégration"
    success_list_url_name = "integration_list"


class WebhookListView(DevflowListView):
    model = dm.Webhook
    template_name = "project/webhook/list.html"
    section = "settings"
    page_title = "Webhooks"
    search_fields = ("url", "secret")


class WebhookDetailView(DevflowDetailView):
    model = dm.Webhook
    template_name = "project/webhook/detail.html"
    section = "settings"
    page_title = "Détail webhook"


class WebhookCreateView(DevflowCreateView):
    model = dm.Webhook
    form_class = WebhookForm
    section = "settings"
    page_title = "Créer webhook"
    success_list_url_name = "webhook_list"


class WebhookUpdateView(DevflowUpdateView):
    model = dm.Webhook
    form_class = WebhookForm
    section = "settings"
    page_title = "Modifier webhook"
    success_list_url_name = "webhook_list"


class WebhookDeleteView(DevflowDeleteView):
    model = dm.Webhook
    section = "settings"
    page_title = "Supprimer webhook"
    success_list_url_name = "webhook_list"


class ReactionListView(DevflowListView):
    model = dm.Reaction
    template_name = "project/reaction/list.html"
    section = "messages"
    page_title = "Réactions"
    search_fields = ("emoji", "user__username")


class ReactionDetailView(DevflowDetailView):
    model = dm.Reaction
    template_name = "project/reaction/detail.html"
    section = "messages"
    page_title = "Détail réaction"


class ReactionCreateView(DevflowCreateView):
    model = dm.Reaction
    form_class = ReactionForm
    section = "messages"
    page_title = "Créer réaction"
    success_list_url_name = "reaction_list"


class ReactionUpdateView(DevflowUpdateView):
    model = dm.Reaction
    form_class = ReactionForm
    section = "messages"
    page_title = "Modifier réaction"
    success_list_url_name = "reaction_list"


class ReactionDeleteView(DevflowDeleteView):
    model = dm.Reaction
    section = "messages"
    page_title = "Supprimer réaction"
    success_list_url_name = "reaction_list"


class MessageAttachmentListView(DevflowListView):
    model = dm.MessageAttachment
    template_name = "project/message_attachment/list.html"
    section = "messages"
    page_title = "Pièces jointes messages"
    search_fields = ("name", "mime_type")


class MessageAttachmentDetailView(DevflowDetailView):
    model = dm.MessageAttachment
    template_name = "project/message_attachment/detail.html"
    section = "messages"
    page_title = "Détail pièce jointe message"


class MessageAttachmentCreateView(DevflowCreateView):
    model = dm.MessageAttachment
    form_class = MessageAttachmentForm
    section = "messages"
    page_title = "Créer pièce jointe message"
    success_list_url_name = "message_attachment_list"


class MessageAttachmentUpdateView(DevflowUpdateView):
    model = dm.MessageAttachment
    form_class = MessageAttachmentForm
    section = "messages"
    page_title = "Modifier pièce jointe message"
    success_list_url_name = "message_attachment_list"


class MessageAttachmentDeleteView(DevflowDeleteView):
    model = dm.MessageAttachment
    section = "messages"
    page_title = "Supprimer pièce jointe message"
    success_list_url_name = "message_attachment_list"


class SprintReviewListView(DevflowListView):
    model = dm.SprintReview
    template_name = "project/sprint_review/list.html"
    section = "sprint"
    page_title = "Sprint reviews"
    search_fields = ("demo_notes", "stakeholder_feedback")


class SprintReviewDetailView(DevflowDetailView):
    model = dm.SprintReview
    template_name = "project/sprint_review/detail.html"
    section = "sprint"
    page_title = "Détail sprint review"


class SprintReviewCreateView(DevflowCreateView):
    model = dm.SprintReview
    form_class = SprintReviewForm
    section = "sprint"
    page_title = "Créer sprint review"
    success_list_url_name = "sprint_review_list"


class SprintReviewUpdateView(DevflowUpdateView):
    model = dm.SprintReview
    form_class = SprintReviewForm
    section = "sprint"
    page_title = "Modifier sprint review"
    success_list_url_name = "sprint_review_list"


class SprintReviewDeleteView(DevflowDeleteView):
    model = dm.SprintReview
    section = "sprint"
    page_title = "Supprimer sprint review"
    success_list_url_name = "sprint_review_list"


class SprintRetrospectiveListView(DevflowListView):
    model = dm.SprintRetrospective
    template_name = "project/sprint_retrospective/list.html"
    section = "sprint"
    page_title = "Rétrospectives sprint"
    search_fields = ("went_well", "to_improve", "action_items")


class SprintRetrospectiveDetailView(DevflowDetailView):
    model = dm.SprintRetrospective
    template_name = "project/sprint_retrospective/detail.html"
    section = "sprint"
    page_title = "Détail rétrospective sprint"


class SprintRetrospectiveCreateView(DevflowCreateView):
    model = dm.SprintRetrospective
    form_class = SprintRetrospectiveForm
    section = "sprint"
    page_title = "Créer rétrospective sprint"
    success_list_url_name = "sprint_retrospective_list"


class SprintRetrospectiveUpdateView(DevflowUpdateView):
    model = dm.SprintRetrospective
    form_class = SprintRetrospectiveForm
    section = "sprint"
    page_title = "Modifier rétrospective sprint"
    success_list_url_name = "sprint_retrospective_list"


class SprintRetrospectiveDeleteView(DevflowDeleteView):
    model = dm.SprintRetrospective
    section = "sprint"
    page_title = "Supprimer rétrospective sprint"
    success_list_url_name = "sprint_retrospective_list"


class APIKeyListView(DevflowListView):
    model = dm.APIKey
    template_name = "project/api_key/list.html"
    section = "settings"
    page_title = "Clés API"
    search_fields = ("name", "key_prefix", "scope")


class APIKeyDetailView(DevflowDetailView):
    model = dm.APIKey
    template_name = "project/api_key/detail.html"
    section = "settings"
    page_title = "Détail clé API"


class APIKeyCreateView(DevflowCreateView):
    model = dm.APIKey
    form_class = APIKeyForm
    section = "settings"
    page_title = "Créer clé API"
    success_list_url_name = "api_key_list"


class APIKeyUpdateView(DevflowUpdateView):
    model = dm.APIKey
    form_class = APIKeyForm
    section = "settings"
    page_title = "Modifier clé API"
    success_list_url_name = "api_key_list"


class APIKeyDeleteView(DevflowDeleteView):
    model = dm.APIKey
    section = "settings"
    page_title = "Supprimer clé API"
    success_list_url_name = "api_key_list"


class WorkspaceSettingsListView(DevflowListView):
    model = dm.WorkspaceSettings
    template_name = "project/workspace_settings/list.html"
    section = "settings"
    page_title = "Paramètres workspace"


class WorkspaceSettingsDetailView(DevflowDetailView):
    model = dm.WorkspaceSettings
    template_name = "project/workspace_settings/detail.html"
    section = "settings"
    page_title = "Détail paramètres workspace"


class WorkspaceSettingsCreateView(DevflowCreateView):
    model = dm.WorkspaceSettings
    form_class = WorkspaceSettingsForm
    section = "settings"
    page_title = "Créer paramètres workspace"
    success_list_url_name = "workspace_settings_list"


class WorkspaceSettingsUpdateView(DevflowUpdateView):
    model = dm.WorkspaceSettings
    form_class = WorkspaceSettingsForm
    section = "settings"
    page_title = "Modifier paramètres workspace"
    success_list_url_name = "workspace_settings_list"


class WorkspaceSettingsDeleteView(DevflowDeleteView):
    model = dm.WorkspaceSettings
    section = "settings"
    page_title = "Supprimer paramètres workspace"
    success_list_url_name = "workspace_settings_list"


class ObjectiveListView(DevflowListView):
    model = dm.Objective
    template_name = "project/objective/list.html"
    section = "analytics"
    page_title = "Objectifs"
    search_fields = ("title", "description", "level", "status", "quarter_label")


class ObjectiveDetailView(DevflowDetailView):
    model = dm.Objective
    template_name = "project/objective/detail.html"
    section = "analytics"
    page_title = "Détail objectif"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["key_results"] = self.object.key_results.all()
        return ctx


class ObjectiveCreateView(DevflowCreateView):
    model = dm.Objective
    template_name = "project/objective/form.html"
    form_class = ObjectiveForm
    section = "analytics"
    page_title = "Créer objectif"
    success_list_url_name = "objective_list"


class ObjectiveUpdateView(DevflowUpdateView):
    model = dm.Objective
    form_class = ObjectiveForm
    section = "analytics"
    page_title = "Modifier objectif"
    success_list_url_name = "objective_list"


class ObjectiveDeleteView(DevflowDeleteView):
    model = dm.Objective
    section = "analytics"
    page_title = "Supprimer objectif"
    success_list_url_name = "objective_list"


class ObjectiveArchiveView(ArchiveObjectView):
    model = dm.Objective
    success_list_url_name = "objective_list"


class KeyResultListView(DevflowListView):
    model = dm.KeyResult
    template_name = "project/key_result/list.html"
    section = "analytics"
    page_title = "Key results"
    search_fields = (
        "title",
        "unit",
        "objective__title",
        "objective__quarter_label",
        "owner__username",
    )

    def get_queryset(self):
        return (
            dm.KeyResult.objects
            .select_related("objective", "objective__team", "objective__owner", "owner")
            .order_by("objective__start_date", "objective__title", "title")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = ctx["object_list"]
        today = timezone.localdate()

        ctx.update({
            "today": today,
            "stats": qs.aggregate(
                total=Count("id"),
                percentage=Count("id", filter=Q(result_type=dm.KeyResult.ResultType.PERCENTAGE)),
                number=Count("id", filter=Q(result_type=dm.KeyResult.ResultType.NUMBER)),
                boolean=Count("id", filter=Q(result_type=dm.KeyResult.ResultType.BOOLEAN)),
                currency=Count("id", filter=Q(result_type=dm.KeyResult.ResultType.CURRENCY)),
                avg_target=Avg("target_value"),
                avg_current=Avg("current_value"),
            ),
        })
        return ctx


class KeyResultDetailView(DevflowDetailView):
    model = dm.KeyResult
    template_name = "project/key_result/detail.html"
    section = "analytics"
    page_title = "Détail key result"

    def get_queryset(self):
        return (
            dm.KeyResult.objects
            .select_related("objective", "objective__team", "objective__owner", "owner")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        kr = self.object
        objective = kr.objective
        today = timezone.localdate()

        sibling_krs = objective.key_results.exclude(pk=kr.pk).select_related("owner").order_by("title")

        ctx.update({
            "today": today,
            "objective": objective,
            "sibling_key_results": sibling_krs,
            "is_objective_overdue": objective.end_date < today and objective.status != dm.Objective.Status.DONE,
        })
        return ctx


class KeyResultCreateView(DevflowCreateView):
    model = dm.KeyResult
    form_class = KeyResultForm
    template_name = "project/key_result/form.html"
    section = "analytics"
    page_title = "Créer key result"
    success_list_url_name = "key_result_list"

    def form_valid(self, form):
        messages.success(self.request, "Key result créé avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class KeyResultUpdateView(DevflowUpdateView):
    model = dm.KeyResult
    form_class = KeyResultForm
    template_name = "project/key_result/form.html"
    section = "analytics"
    page_title = "Modifier key result"
    success_list_url_name = "key_result_list"

    def form_valid(self, form):
        messages.success(self.request, "Key result modifié avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Veuillez corriger les erreurs du formulaire.")
        return self.render_to_response(self.get_context_data(form=form))


class KeyResultDeleteView(DevflowDeleteView):
    model = dm.KeyResult
    section = "analytics"
    page_title = "Supprimer key result"
    success_list_url_name = "key_result_list"


# =============================================================================
# FACTURATION
# =============================================================================
class InvoiceClientListView(DevflowListView):
    model = dm.InvoiceClient
    template_name = "project/invoice_client/list.html"
    section = "billing"
    page_title = "Clients de facturation"
    search_placeholder = "Rechercher par nom, raison sociale, ID fiscal, ville…"
    search_fields = ("name", "legal_name", "tax_id", "email", "city", "country")
    filter_fields = [
        {"name": "country", "label": "Pays", "type": "exact"},
    ]


class InvoiceClientDetailView(DevflowDetailView):
    model = dm.InvoiceClient
    template_name = "project/invoice_client/detail.html"
    section = "billing"
    page_title = "Détail client"


class InvoiceClientCreateView(DevflowCreateView):
    model = dm.InvoiceClient
    form_class = InvoiceClientForm
    template_name = "project/invoice_client/form.html"
    section = "billing"
    page_title = "Nouveau client"
    success_list_url_name = "invoice_client_list"


class InvoiceClientUpdateView(DevflowUpdateView):
    model = dm.InvoiceClient
    form_class = InvoiceClientForm
    template_name = "project/invoice_client/form.html"
    section = "billing"
    page_title = "Modifier client"
    success_list_url_name = "invoice_client_list"


class InvoiceClientDeleteView(DevflowDeleteView):
    model = dm.InvoiceClient
    section = "billing"
    page_title = "Supprimer client"
    success_list_url_name = "invoice_client_list"


class InvoiceListView(DevflowListView):
    model = dm.Invoice
    template_name = "project/invoice/list.html"
    section = "billing"
    page_title = "Factures"
    paginate_by = 25
    search_placeholder = "Rechercher par numéro, titre, projet ou client…"
    search_fields = ("number", "title", "project__name", "client__name", "status")
    filter_fields = [
        {"name": "status", "label": "Statut", "type": "choices",
         "choices": dm.Invoice.Status.choices},
        {"name": "billing_mode", "label": "Mode", "type": "choices",
         "choices": dm.Invoice.BillingMode.choices},
        {"name": "project", "label": "Projet", "type": "model",
         "queryset": "_projects_qs", "lookup": "project_id"},
        {"name": "client", "label": "Client", "type": "model",
         "queryset": "_clients_qs", "lookup": "client_id"},
    ]

    def _projects_qs(self):
        ws = self.get_current_workspace()
        return dm.Project.objects.filter(workspace=ws, is_archived=False).order_by("name") if ws else dm.Project.objects.none()

    def _clients_qs(self):
        ws = self.get_current_workspace()
        return dm.InvoiceClient.objects.filter(workspace=ws, is_archived=False).order_by("name") if ws else dm.InvoiceClient.objects.none()

    def get_queryset(self):
        return (
            super().get_queryset()
            .select_related("project", "client", "issued_by")
            .order_by("-issue_date", "-id")
        )

    def get_context_data(self, **kwargs):
        from django.db.models import Sum
        ctx = super().get_context_data(**kwargs)
        qs = self.get_queryset()
        agg = qs.aggregate(
            total_ttc=Sum("total_ttc"),
            total_paid=Sum("paid_amount"),
        )
        ctx["agg_total_ttc"] = agg["total_ttc"] or 0
        ctx["agg_total_paid"] = agg["total_paid"] or 0
        ctx["agg_total_due"] = (agg["total_ttc"] or 0) - (agg["total_paid"] or 0)
        ctx["count_overdue"] = qs.filter(status=dm.Invoice.Status.OVERDUE).count()
        ctx["count_draft"] = qs.filter(status=dm.Invoice.Status.DRAFT).count()
        return ctx


class InvoiceDetailView(DevflowDetailView):
    model = dm.Invoice
    template_name = "project/invoice/detail.html"
    section = "billing"
    page_title = "Facture"
    context_object_name = "invoice"

    def get_queryset(self):
        return (
            super().get_queryset()
            .select_related("project", "client", "issued_by", "workspace")
            .prefetch_related("lines", "payments")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        invoice = self.object
        ctx["lines"] = invoice.lines.all().order_by("position", "id")
        ctx["payments"] = invoice.payments.all().order_by("-received_at")
        ctx["payment_form"] = InvoicePaymentForm(
            current_workspace=invoice.workspace,
            initial={"invoice": invoice.pk},
        )
        ctx["remaining"] = invoice.remaining_due
        return ctx


class InvoiceCreateView(DevflowCreateView):
    model = dm.Invoice
    form_class = InvoiceForm
    template_name = "project/invoice/form.html"
    section = "billing"
    page_title = "Nouvelle facture"
    success_list_url_name = "invoice_list"

    def get_initial(self):
        initial = super().get_initial()
        project_id = self.request.GET.get("project")
        if project_id:
            initial["project"] = project_id
        return initial

    def form_valid(self, form):
        invoice = form.save(commit=False)
        invoice.issued_by = self.request.user
        if not invoice.workspace_id and invoice.project_id:
            invoice.workspace = invoice.project.workspace
        invoice.save()
        invoice.recompute_totals()
        messages.success(self.request, f"Facture {invoice.number or 'brouillon'} créée.")
        return redirect("invoice_detail", pk=invoice.pk)


class InvoiceUpdateView(DevflowUpdateView):
    model = dm.Invoice
    form_class = InvoiceForm
    template_name = "project/invoice/form.html"
    section = "billing"
    page_title = "Modifier facture"
    success_list_url_name = "invoice_list"

    def form_valid(self, form):
        response = super().form_valid(form)
        self.object.recompute_totals()
        return response


class InvoiceDeleteView(DevflowDeleteView):
    model = dm.Invoice
    section = "billing"
    page_title = "Supprimer facture"
    success_list_url_name = "invoice_list"


class InvoiceIssueView(DevflowBaseMixin, View):
    """Passe une facture en statut ISSUED (lui attribue un numéro)."""

    def post(self, request, pk):
        invoice = self.filter_by_workspace(dm.Invoice.objects.all()).get(pk=pk)
        if invoice.status == dm.Invoice.Status.DRAFT:
            invoice.status = dm.Invoice.Status.ISSUED
            invoice.save()  # Trigger numérotation
            invoice.recompute_totals()
            messages.success(request, f"Facture émise sous le numéro {invoice.number}.")
        else:
            messages.warning(request, "Seules les factures en brouillon peuvent être émises.")
        return redirect("invoice_detail", pk=invoice.pk)


class InvoiceMarkSentView(DevflowBaseMixin, View):
    def post(self, request, pk):
        invoice = self.filter_by_workspace(dm.Invoice.objects.all()).get(pk=pk)
        invoice.status = dm.Invoice.Status.SENT
        invoice.sent_at = timezone.now()
        invoice.save(update_fields=["status", "sent_at", "updated_at"])
        messages.success(request, "Facture marquée comme envoyée.")
        return redirect("invoice_detail", pk=invoice.pk)


class InvoiceCancelView(DevflowBaseMixin, View):
    def post(self, request, pk):
        invoice = self.filter_by_workspace(dm.Invoice.objects.all()).get(pk=pk)
        invoice.status = dm.Invoice.Status.CANCELLED
        invoice.save(update_fields=["status", "updated_at"])
        messages.success(request, "Facture annulée.")
        return redirect("invoice_detail", pk=invoice.pk)


class InvoicePrintView(DevflowDetailView):
    model = dm.Invoice
    template_name = "project/invoice/print.html"
    section = "billing"
    page_title = "Aperçu facture"
    context_object_name = "invoice"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["lines"] = self.object.lines.all().order_by("position", "id")
        ctx["payments"] = self.object.payments.filter(
            status=dm.InvoicePayment.Status.CONFIRMED
        )
        return ctx


class InvoicePDFView(DevflowBaseMixin, View):
    """
    Génère et télécharge un PDF A4 avec papier en-tête du workspace.
    URL: /billing/invoices/<pk>/pdf/  → name="invoice_pdf"
    Param ?inline=1 pour afficher dans le navigateur au lieu de télécharger.
    """

    def get(self, request, pk):
        from django.http import HttpResponse, Http404
        from project.services.invoice_pdf import render_invoice_pdf

        invoice = self.filter_by_workspace(
            dm.Invoice.objects.all()
        ).select_related("project", "client", "workspace", "issued_by").filter(pk=pk).first()
        if not invoice:
            raise Http404("Facture introuvable.")

        try:
            pdf_bytes = render_invoice_pdf(invoice, request=request)
        except Exception as exc:
            messages.error(request, f"Impossible de générer le PDF : {exc}")
            return redirect("invoice_detail", pk=invoice.pk)

        filename = (invoice.number or f"facture-{invoice.pk}").replace("/", "-")
        disposition = "inline" if request.GET.get("inline") == "1" else "attachment"
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'{disposition}; filename="{filename}.pdf"'
        return response


class InvoiceLineCreateView(DevflowCreateView):
    model = dm.InvoiceLine
    form_class = InvoiceLineForm
    template_name = "project/invoice_line/form.html"
    section = "billing"
    page_title = "Nouvelle ligne"
    success_list_url_name = "invoice_list"

    def get_initial(self):
        initial = super().get_initial()
        invoice_id = self.request.GET.get("invoice")
        if invoice_id:
            initial["invoice"] = invoice_id
        return initial

    def get_success_url(self):
        return reverse_lazy("invoice_detail", args=[self.object.invoice_id])

    def form_valid(self, form):
        response = super().form_valid(form)
        self.object.invoice.recompute_totals()
        return response


class InvoiceLineUpdateView(DevflowUpdateView):
    model = dm.InvoiceLine
    form_class = InvoiceLineForm
    template_name = "project/invoice_line/form.html"
    section = "billing"
    page_title = "Modifier ligne"
    success_list_url_name = "invoice_list"

    def get_success_url(self):
        return reverse_lazy("invoice_detail", args=[self.object.invoice_id])

    def form_valid(self, form):
        response = super().form_valid(form)
        self.object.invoice.recompute_totals()
        return response


class InvoiceLineDeleteView(DevflowDeleteView):
    model = dm.InvoiceLine
    section = "billing"
    page_title = "Supprimer ligne"
    success_list_url_name = "invoice_list"

    def get_success_url(self):
        invoice_id = self.object.invoice_id if hasattr(self, "object") else None
        if invoice_id:
            return reverse_lazy("invoice_detail", args=[invoice_id])
        return reverse_lazy("invoice_list")

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        invoice = self.object.invoice
        response = super().delete(request, *args, **kwargs)
        invoice.recompute_totals()
        return response


class InvoicePaymentCreateView(DevflowBaseMixin, View):
    """Création de paiement depuis la fiche facture."""

    def post(self, request, pk):
        invoice = self.filter_by_workspace(dm.Invoice.objects.all()).get(pk=pk)
        form = InvoicePaymentForm(
            request.POST,
            current_workspace=invoice.workspace,
        )
        if not form.is_valid():
            for err in form.errors.values():
                messages.error(request, " ".join(err))
            return redirect("invoice_detail", pk=invoice.pk)

        payment = form.save(commit=False)
        payment.invoice = invoice
        payment.recorded_by = request.user
        payment.save()
        invoice.recompute_totals()
        messages.success(request, "Paiement enregistré.")
        return redirect("invoice_detail", pk=invoice.pk)


class InvoiceGenerateFromProjectView(DevflowBaseMixin, View):
    """
    GET : affiche le form de génération.
    POST : appelle le service InvoiceGenerator selon le mode et redirige
    vers la facture créée.
    """

    template_name = "project/invoice/generate.html"

    def get(self, request, project_pk):
        project = self.filter_by_workspace(dm.Project.objects.all()).get(pk=project_pk)
        form = InvoiceGenerateForm()
        return render(request, self.template_name, {
            "project": project, "form": form,
        })

    def post(self, request, project_pk):
        from project.services.invoicing import generate_invoice_for_project

        project = self.filter_by_workspace(dm.Project.objects.all()).get(pk=project_pk)
        form = InvoiceGenerateForm(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {
                "project": project, "form": form,
            })

        cleaned = form.cleaned_data
        try:
            invoice = generate_invoice_for_project(
                project,
                mode=cleaned["mode"],
                issued_by=request.user,
                tax_rate=cleaned["tax_rate"],
                currency=cleaned["currency"],
                title=cleaned.get("title") or None,
                period_start=cleaned.get("period_start"),
                period_end=cleaned.get("period_end"),
                notes=cleaned.get("notes") or "",
            )
        except Exception as exc:
            messages.error(request, f"Génération impossible : {exc}")
            return render(request, self.template_name, {
                "project": project, "form": form,
            })

        if invoice.lines.count() == 0:
            messages.warning(
                request,
                "Aucune ligne facturable n'a été trouvée pour ce mode. "
                "Ajustez les estimations / timesheets / jalons puis relancez."
            )
        else:
            messages.success(
                request,
                f"Brouillon de facture généré ({invoice.lines.count()} lignes)."
            )
        return redirect("invoice_detail", pk=invoice.pk)
