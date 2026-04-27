"""
Service de conversation DevFlow AI (panneau latéral).

Architecture :
1. Contexte injecté à chaque requête : sprint actif, vélocité, projets
   à risque, charge équipe — tirés des modèles DevFlow réels.
2. Prompt système qui borne strictement l'IA au domaine DevFlow.
3. Provider hybride (OpenAI / local / heuristique) déjà configuré dans
   project/services/ai/factory.py.
4. Persistance complète : AIChatSession + AIChatMessage (audit, multi-tour).
5. Intents pré-câblés pour les chips de suggestion (analyse sprint,
   projets à risque, rapport, charge équipe). Même quand l'IA est
   indisponible, ces intents répondent avec des données factuelles.
"""
from __future__ import annotations

import json
import logging
from dataclasses import dataclass, field
from decimal import Decimal
from typing import Any

import requests
from django.conf import settings
from django.contrib.auth import get_user_model
from django.db.models import Count, Sum
from django.utils import timezone

from project import models as dm
from project.services.ai.base import AIMessage
from project.services.ai.factory import get_ai_provider
from io import BytesIO
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from openpyxl import Workbook

logger = logging.getLogger(__name__)
User = get_user_model()


@dataclass
class ChatTurnResult:
    session: dm.AIChatSession
    user_message: dm.AIChatMessage
    assistant_message: dm.AIChatMessage
    used_provider: str
    context: dict = field(default_factory=dict)


class DevFlowContextBuilder:

    @staticmethod
    def _workspace_tasks(ws, project=None, limit=80) -> list[dict]:
        today = timezone.localdate()

        qs = (
            dm.Task.objects.filter(workspace=ws, is_archived=False)
            .select_related("project", "assignee", "sprint")
            .order_by("due_date", "-priority", "title")
        )

        if project:
            qs = qs.filter(project=project)

        out = []

        for task in qs[:limit]:
            status_key = getattr(task, "status", "")
            priority_key = getattr(task, "priority", "")

            is_overdue = (
                    bool(task.due_date)
                    and task.due_date < today
                    and status_key not in ["DONE", "CLOSED", "CANCELLED"]
            )

            out.append({
                "id": task.pk,
                "title": task.title,
                "project": task.project.name if task.project else "",
                "sprint": task.sprint.name if getattr(task, "sprint_id", None) else "",
                "status_key": status_key,
                "status": task.get_status_display(),
                "priority_key": priority_key,
                "priority": task.get_priority_display(),
                "assignee": task.assignee.get_full_name() or task.assignee.username if task.assignee else "",
                "due_date": task.due_date.isoformat() if task.due_date else None,
                "is_overdue": is_overdue,
            })

        return out
    @classmethod
    def build_for_user(cls, user, workspace=None, project=None) -> dict:
        ws = workspace or cls._infer_workspace(user)

        return {
            "user": cls._user_summary(user),
            "workspace": cls._workspace_summary(ws) if ws else None,
            "active_sprint": cls._active_sprint(ws, project) if ws else None,
            "at_risk_projects": cls._at_risk_projects(ws, limit=4) if ws else [],
            "team_workload": cls._team_workload(ws, limit=6) if ws else [],
            "user_open_tasks": cls._user_open_tasks(user, limit=5),
            "now": timezone.now().isoformat(),
            "workspace_tasks": cls._workspace_tasks(ws, project, limit=80) if ws else [],
        }

    @staticmethod
    def _user_summary(user) -> dict:
        return {
            "id": user.pk,
            "name": user.get_full_name() or user.username,
            "first_name": user.first_name or user.username,
            "email": user.email,
        }

    @staticmethod
    def _infer_workspace(user):
        membership = (
            dm.TeamMembership.objects.filter(user=user)
            .select_related("workspace")
            .order_by("-created_at")
            .first()
        )

        if membership and membership.workspace_id:
            return membership.workspace

        return dm.Workspace.objects.filter(owner=user).first()

    @staticmethod
    def _workspace_summary(ws) -> dict:
        active_projects = (
            ws.projects.filter(is_archived=False)
            .exclude(status__in=[dm.Project.Status.DONE, dm.Project.Status.CANCELLED])
        )

        return {
            "id": ws.pk,
            "name": ws.name,
            "active_projects_count": active_projects.count(),
            "members_count": ws.memberships.count(),
        }

    @staticmethod
    def _active_sprint(ws, project=None):
        qs = dm.Sprint.objects.filter(
            workspace=ws,
            status=dm.Sprint.Status.ACTIVE,
            is_archived=False,
        )

        if project:
            qs = qs.filter(project=project)

        sprint = qs.order_by("-start_date").first()

        if not sprint:
            return None

        today = timezone.localdate()
        days_remaining = (
            max((sprint.end_date - today).days, 0)
            if sprint.end_date
            else None
        )

        velocity_target = sprint.velocity_target or 0
        velocity_completed = sprint.velocity_completed or 0
        velocity_percent = (
            int((velocity_completed / velocity_target) * 100)
            if velocity_target > 0
            else 0
        )

        return {
            "id": sprint.pk,
            "name": sprint.name,
            "number": sprint.number,
            "project": sprint.project.name if sprint.project else "",
            "days_remaining": days_remaining,
            "velocity_target": velocity_target,
            "velocity_completed": velocity_completed,
            "velocity_percent": velocity_percent,
            "total_story_points": sprint.total_story_points,
            "completed_story_points": sprint.completed_story_points,
            "remaining_story_points": sprint.remaining_story_points,
        }

    @staticmethod
    def _at_risk_projects(ws, limit=4) -> list[dict]:
        today = timezone.localdate()

        qs = (
            ws.projects.filter(is_archived=False)
            .exclude(status__in=[dm.Project.Status.DONE, dm.Project.Status.CANCELLED])
            .order_by("-risk_score")
        )

        out = []

        for project in qs[:limit]:
            days_late = 0

            if (
                project.target_date
                and project.target_date < today
                and project.status != dm.Project.Status.DONE
            ):
                days_late = (today - project.target_date).days

            out.append({
                "id": project.pk,
                "name": project.name,
                "status": project.get_status_display(),
                "priority": project.get_priority_display(),
                "risk_score": project.risk_score,
                "ai_risk_label": project.ai_risk_label or "—",
                "progress_percent": project.progress_percent,
                "target_date": project.target_date.isoformat() if project.target_date else None,
                "days_late": days_late,
            })

        return out

    @staticmethod
    def _team_workload(ws, limit=6) -> list[dict]:
        rows = (
            dm.Task.objects.filter(
                workspace=ws,
                is_archived=False,
                assignee__isnull=False,
            )
            .exclude(status__in=[dm.Task.Status.DONE, dm.Task.Status.CANCELLED])
            .values(
                "assignee_id",
                "assignee__username",
                "assignee__first_name",
                "assignee__last_name",
            )
            .annotate(
                open_tasks=Count("id"),
                est_hours=Sum("estimate_hours"),
                spent_hours=Sum("spent_hours"),
            )
            .order_by("-open_tasks")[:limit]
        )

        out = []

        for row in rows:
            est = row["est_hours"] or Decimal("0")
            spent = row["spent_hours"] or Decimal("0")
            remaining = max(est - spent, Decimal("0"))

            label = (
                f"{row['assignee__first_name'] or ''} {row['assignee__last_name'] or ''}"
            ).strip() or row["assignee__username"]

            out.append({
                "user_id": row["assignee_id"],
                "user_label": label,
                "open_tasks": row["open_tasks"],
                "estimate_hours": float(est),
                "remaining_hours": float(remaining),
            })

        return out

    @staticmethod
    def _user_open_tasks(user, limit=5) -> list[dict]:
        qs = (
            dm.Task.objects.filter(assignee=user, is_archived=False)
            .exclude(status__in=[dm.Task.Status.DONE, dm.Task.Status.CANCELLED])
            .select_related("project")
            .order_by("due_date", "-priority")[:limit]
        )

        return [
            {
                "id": task.pk,
                "title": task.title,
                "project": task.project.name if task.project else "",
                "status": task.get_status_display(),
                "priority": task.get_priority_display(),
                "due_date": task.due_date.isoformat() if task.due_date else None,
            }
            for task in qs
        ]


def web_search(query: str) -> str:
    query = (query or "").strip()

    if not query:
        return ""

    if not getattr(settings, "AI_WEB_SEARCH_ENABLED", False):
        return ""

    api_key = getattr(settings, "TAVILY_API_KEY", "")

    if not api_key:
        logger.warning("Tavily web search skipped: missing TAVILY_API_KEY")
        return ""

    payload: dict[str, Any] = {
        "api_key": api_key,
        "query": query,
        "search_depth": "basic",
        "max_results": 5,
        "include_answer": False,
        "include_raw_content": False,
    }

    try:
        response = requests.post(
            "https://api.tavily.com/search",
            json=payload,
            timeout=20,
        )
        response.raise_for_status()
        data = response.json()
    except requests.RequestException as exc:
        logger.warning("Tavily web search HTTP error: %s", exc)
        return ""
    except ValueError as exc:
        logger.warning("Tavily web search JSON error: %s", exc)
        return ""

    results = data.get("results") or []

    if not isinstance(results, list):
        return ""

    lines = []

    for item in results[:5]:
        if not isinstance(item, dict):
            continue

        title = (item.get("title") or "").strip()
        content = (item.get("content") or "").strip()
        url = (item.get("url") or "").strip()

        if not title and not content:
            continue

        line = f"- {title}: {content}"

        if url:
            line += f" ({url})"

        lines.append(line)

    return "\n".join(lines)


class AIChatService:
    SYSTEM_PROMPT = (
        "Tu es DevFlow AI, l'assistant intelligent intégré à DevFlow. "
        "Tu aides l'utilisateur sur la gestion de projet, l'organisation, la productivité, "
        "l'analyse des risques, les sprints, les équipes, les budgets et les rapports. "
        "Tu peux aussi répondre aux questions générales quand elles sont utiles, "
        "même si elles ne sont pas directement liées aux données DevFlow. "

        "La plateforme DevFlow a été développée par Serge OGAH. "
        "Quand l'utilisateur demande qui est le développeur, le créateur ou l'auteur de la plateforme, "
        "tu dois répondre que le développeur est Serge OGAH. "
        "Tu peux le présenter comme un développeur full-stack, architecte logiciel, spécialisé dans "
        "les plateformes Django, les systèmes de gestion de projet, les solutions IA, les dashboards, "
        "les architectures Docker/Celery/Redis/PostgreSQL et les applications métiers modernes. "
        "Ne donne pas d'informations personnelles sensibles non fournies. "

        "Quand la question concerne DevFlow, utilise prioritairement le contexte fourni. "
        "Quand la question est générale, réponds avec tes connaissances générales. "
        "Quand des résultats web sont fournis, utilise-les pour répondre aux questions récentes. "

        "Si l'utilisateur demande de générer un fichier Excel ou PDF, ne réponds pas que tu es incapable. "
        "Explique que DevFlow peut générer ce fichier via le backend, puis retourne une réponse courte "
        "avec le lien ou l'action fournie par le système si elle existe. "

        "Tu réponds en français, de manière claire, utile et structurée. "
        "Tu ne dois pas inventer de données internes DevFlow absentes du contexte. "
        "Tu peux utiliser des emojis sobres et du HTML simple comme <strong>...</strong>. "
        "Évite le markdown lourd."
    )

    INTENT_KEYWORDS = {
        "sprint_analysis": ["analyse du sprint", "sprint analysis", "vélocité"],
        "risk_projects": ["projets à risque", "risk", "à risque", "risque"],
        "report": ["générer un rapport", "rapport"],
        "team_workload": ["charge équipe", "charge equipe", "workload"],

        "task_analysis": [
            "analyse des tâches",
            "analyse des taches",
            "état des tâches",
            "etat des taches",
            "suivi des tâches",
            "suivi des taches",
            "tâches du projet",
            "taches du projet",
        ],
        "pending_tasks": [
            "tâches en attente",
            "taches en attente",
            "tâches non réalisées",
            "taches non realisees",
            "tâches non terminées",
            "taches non terminees",
            "reste à faire",
            "reste a faire",
            "backlog ouvert",
        ],
        "critical_tasks": [
            "tâches critiques",
            "taches critiques",
            "tâches urgentes",
            "taches urgentes",
            "priorité critique",
            "priorite critique",
            "tâches bloquées",
            "taches bloquees",
            "blocked tasks",
        ],
        "overdue_tasks": [
            "tâches en retard",
            "taches en retard",
            "retard tâches",
            "retard taches",
            "échéance dépassée",
            "echeance depassee",
            "overdue tasks",
        ],
    }

    GENERAL_KEYWORDS = [
        "qu'est-ce que",
        "c'est quoi",
        "définis",
        "definis",
        "explique",
        "comment faire",
        "quel est",
        "quelle est",
        "quels sont",
        "quelles sont",
        "pourquoi",
        "différence entre",
        "difference entre",
    ]
    FILE_EXPORT_KEYWORDS = [
        "génère un fichier excel",
        "genere un fichier excel",
        "export excel",
        "fichier excel",
        "xlsx",
        "génère un pdf",
        "genere un pdf",
        "export pdf",
        "fichier pdf",
        "excel",

        "xlsx",

        "exporter",

        "exporte",

        "export",

        "fichier excel",

        "export excel",

        "peux tu exporter",

        "peux-tu exporter",

        "tableau excel",
    ]
    WEB_SEARCH_KEYWORDS = [
        "actualité",
        "actualite",
        "aujourd'hui",
        "aujourdhui",
        "maintenant",
        "récent",
        "recent",
        "récemment",
        "recemment",
        "dernier",
        "dernière",
        "derniere",
        "prix actuel",
        "taux actuel",
        "version actuelle",
        "en 2026",
        "news",
        "internet",
        "recherche web",
        "cherche sur internet",
    ]

    @classmethod
    def _wants_file_export(cls, message: str) -> bool:
        text = (message or "").lower()
        return any(keyword in text for keyword in cls.FILE_EXPORT_KEYWORDS)

    @classmethod
    def _generate_export_response(cls, message: str, ctx: dict) -> str:
        text = (message or "").lower()

        if "excel" in text or "xlsx" in text or "export" in text:
            return cls._generate_risk_projects_excel(ctx)

        return "Je peux générer un export Excel. Précisez les données à exporter."

    @classmethod
    def _generate_risk_projects_excel(cls, ctx: dict) -> str:
        risks = ctx.get("at_risk_projects") or []

        wb = Workbook()
        ws = wb.active
        ws.title = "Projets à risque"

        ws.append(["Nom du projet", "Risque", "Score", "Progression", "Retard"])

        for project in risks:
            ws.append([
                project.get("name", ""),
                project.get("ai_risk_label", ""),
                project.get("risk_score", 0),
                f"{project.get('progress_percent', 0)}%",
                f"{project.get('days_late', 0)} jours",
            ])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"exports/devflow_projets_risque_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        saved_path = default_storage.save(
            filename,
            ContentFile(buffer.getvalue())
        )

        file_url = default_storage.url(saved_path)

        return (
            "✅ <strong>Fichier Excel généré avec succès.</strong><br>"
            f"<a href='{file_url}' target='_blank' download>Télécharger le fichier Excel</a>"
        )
    @classmethod
    def get_or_create_session(cls, user, session_id=None, workspace=None, project=None):
        if session_id:
            session = (
                dm.AIChatSession.objects.filter(
                    pk=session_id,
                    user=user,
                    is_active=True,
                )
                .first()
            )

            if session:
                return session

        return dm.AIChatSession.objects.create(
            user=user,
            workspace=workspace or DevFlowContextBuilder._infer_workspace(user),
            project=project,
        )

    @classmethod
    def welcome_message(cls, user, workspace=None) -> dict:
        ctx = DevFlowContextBuilder.build_for_user(user, workspace=workspace)

        sprint = ctx.get("active_sprint")
        risks = ctx.get("at_risk_projects") or []
        first_name = ctx["user"]["first_name"]

        parts = [f"👋 Bonjour {first_name} ! Je suis DevFlow AI."]

        if sprint:
            days = sprint.get("days_remaining")
            days_label = f" avec {days} jours restants" if days is not None else ""

            velocity_label = (
                " — légèrement en dessous de la cible."
                if sprint["velocity_percent"] < 80
                else " — au-dessus de la cible."
            )

            parts.append(
                f"<br><br>Je vois que le <strong>{sprint['name']}</strong> "
                f"est en cours{days_label}. "
                f"La vélocité est à <strong>{sprint['velocity_percent']}%</strong>"
                f"{velocity_label}"
            )

        elif ctx.get("workspace"):
            ws = ctx["workspace"]
            parts.append(
                f"<br><br>Aucun sprint actif. "
                f"{ws['active_projects_count']} projet(s) actifs sur le workspace "
                f"<strong>{ws['name']}</strong>."
            )

        late_projects = [p for p in risks if p.get("days_late", 0) > 0]

        if late_projects:
            top = late_projects[0]
            parts.append(
                f"<br><br>⚠️ <strong>{top['name']}</strong> est en retard "
                f"({top['days_late']} jour(s)). Voulez-vous que j'analyse les risques ?"
            )
        elif risks:
            high = [p for p in risks if p.get("risk_score", 0) >= 60]
            if high:
                parts.append(
                    f"<br><br>⚠️ <strong>{high[0]['name']}</strong> présente "
                    f"un risque élevé (score {high[0]['risk_score']})."
                )

        return {
            "html": "".join(parts),
            "context": ctx,
            "suggestions": [
                "📊 Analyse du sprint",
                "⚠️ Projets à risque",
                "📋 Générer un rapport",
                "👥 Charge équipe",
            ],
        }

    @classmethod
    def process_user_message(
        cls,
        user,
        message: str,
        session_id=None,
        workspace=None,
        project=None,
    ) -> ChatTurnResult:
        message = (message or "").strip()

        if not message:
            raise ValueError("Message vide")

        session = cls.get_or_create_session(
            user=user,
            session_id=session_id,
            workspace=workspace,
            project=project,
        )

        intent = cls._detect_intent(message)

        user_msg = dm.AIChatMessage.objects.create(
            session=session,
            role=dm.AIChatMessage.Role.USER,
            content=message[:5000],
            intent=intent,
        )

        ctx = DevFlowContextBuilder.build_for_user(
            user=user,
            workspace=session.workspace,
            project=session.project,
        )

        deterministic = cls._answer_intent(intent, ctx)
        if cls._wants_file_export(message):
            export_text = cls._generate_export_response(message, ctx)
            assistant_msg = dm.AIChatMessage.objects.create(
                session=session,
                role=dm.AIChatMessage.Role.ASSISTANT,
                content=export_text[:8000],
                intent="file_export",
                used_provider="system",
                used_model="django-export",
                tokens_used=0,
                context_payload=ctx,
            )

            return ChatTurnResult(
                session=session,
                user_message=user_msg,
                assistant_message=assistant_msg,
                used_provider="system",
                context=ctx,
            )
        ai_text, used_provider, model_name, tokens = cls._call_ai(
            session=session,
            ctx=ctx,
            current_message=message,
            seed=deterministic,
        )

        final_text = (
            ai_text
            or deterministic
            or "Je n'ai pas réussi à interpréter votre demande. Pouvez-vous reformuler ?"
        )

        assistant_msg = dm.AIChatMessage.objects.create(
            session=session,
            role=dm.AIChatMessage.Role.ASSISTANT,
            content=final_text[:8000],
            intent=intent,
            used_provider=used_provider,
            used_model=model_name,
            tokens_used=tokens,
            context_payload=ctx,
        )

        if not session.title:
            session.title = message[:120] or "Conversation IA"
            session.save(update_fields=["title", "updated_at"])

        return ChatTurnResult(
            session=session,
            user_message=user_msg,
            assistant_message=assistant_msg,
            used_provider=used_provider,
            context=ctx,
        )

    @classmethod
    def _detect_intent(cls, message: str) -> str:
        text = (message or "").lower()

        for intent, keywords in cls.INTENT_KEYWORDS.items():
            if any(keyword in text for keyword in keywords):
                return intent

        return ""

    @classmethod
    def _is_general_question(cls, message: str) -> bool:
        text = (message or "").lower()
        return any(keyword in text for keyword in cls.GENERAL_KEYWORDS)

    @classmethod
    def _needs_web_search(cls, message: str) -> bool:
        text = (message or "").lower()
        return any(keyword in text for keyword in cls.WEB_SEARCH_KEYWORDS)

    @staticmethod
    def _answer_task_analysis(ctx: dict) -> str:
        tasks = ctx.get("workspace_tasks") or []

        if not tasks:
            return "ℹ️ Aucune tâche disponible pour l’analyse."

        total = len(tasks)
        done = len([t for t in tasks if t.get("status_key") in ["DONE", "CLOSED"]])
        blocked = len([t for t in tasks if t.get("status_key") == "BLOCKED"])
        overdue = len([t for t in tasks if t.get("is_overdue")])
        critical = len([t for t in tasks if t.get("priority_key") in ["CRITICAL", "HIGH"]])

        progress = round((done / total) * 100, 1) if total else 0

        return (
            f"📌 <strong>Analyse des tâches</strong><br>"
            f"• Total : <strong>{total}</strong><br>"
            f"• Terminées : <strong>{done}</strong> ({progress}%)<br>"
            f"• Bloquées : <strong>{blocked}</strong><br>"
            f"• En retard : <strong>{overdue}</strong><br>"
            f"• Critiques / haute priorité : <strong>{critical}</strong>"
        )

    @staticmethod
    def _answer_pending_tasks(ctx: dict) -> str:
        tasks = [
            t for t in (ctx.get("workspace_tasks") or [])
            if t.get("status_key") not in ["DONE", "CLOSED", "CANCELLED"]
        ]

        if not tasks:
            return "✅ Aucune tâche en attente détectée."

        lines = ["🕒 <strong>Tâches non terminées / en attente</strong>"]

        for task in tasks[:10]:
            lines.append(
                f"• <strong>{task['title']}</strong> — {task['status']} · "
                f"{task['priority']} · échéance : {task.get('due_date') or 'non définie'}"
            )

        if len(tasks) > 10:
            lines.append(f"… et {len(tasks) - 10} autre(s) tâche(s).")

        return "<br>".join(lines)

    @staticmethod
    def _answer_critical_tasks(ctx: dict) -> str:
        tasks = [
            t for t in (ctx.get("workspace_tasks") or [])
            if t.get("priority_key") in ["CRITICAL", "HIGH"] or t.get("status_key") == "BLOCKED"
        ]

        if not tasks:
            return "✅ Aucune tâche critique ou bloquée détectée."

        lines = ["🔥 <strong>Tâches critiques / bloquées</strong>"]

        for task in tasks[:10]:
            overdue_label = " · ⚠️ en retard" if task.get("is_overdue") else ""
            lines.append(
                f"• <strong>{task['title']}</strong> — {task['status']} · "
                f"{task['priority']} · {task.get('assignee') or 'non assignée'}{overdue_label}"
            )

        if len(tasks) > 10:
            lines.append(f"… et {len(tasks) - 10} autre(s) tâche(s) critique(s).")

        return "<br>".join(lines)

    @staticmethod
    def _answer_overdue_tasks(ctx: dict) -> str:
        tasks = [t for t in (ctx.get("workspace_tasks") or []) if t.get("is_overdue")]

        if not tasks:
            return "✅ Aucune tâche en retard détectée."

        lines = ["⚠️ <strong>Tâches en retard</strong>"]

        for task in tasks[:10]:
            lines.append(
                f"• <strong>{task['title']}</strong> — échéance : "
                f"{task.get('due_date') or 'non définie'} · "
                f"{task.get('assignee') or 'non assignée'}"
            )

        if len(tasks) > 10:
            lines.append(f"… et {len(tasks) - 10} autre(s) tâche(s) en retard.")

        return "<br>".join(lines)

    @classmethod
    def _answer_intent(cls, intent: str, ctx: dict) -> str:
        if intent == "sprint_analysis":
            return cls._answer_sprint_analysis(ctx)

        if intent == "risk_projects":
            return cls._answer_risk_projects(ctx)

        if intent == "team_workload":
            return cls._answer_team_workload(ctx)

        if intent == "report":
            return cls._answer_report(ctx)

        if intent == "task_analysis":
            return cls._answer_task_analysis(ctx)

        if intent == "pending_tasks":
            return cls._answer_pending_tasks(ctx)

        if intent == "critical_tasks":
            return cls._answer_critical_tasks(ctx)

        if intent == "overdue_tasks":
            return cls._answer_overdue_tasks(ctx)

        return ""

    @staticmethod
    def _answer_sprint_analysis(ctx: dict) -> str:
        sprint = ctx.get("active_sprint")

        if not sprint:
            return "ℹ️ Aucun sprint actif détecté sur votre workspace."

        lines = [
            f"📊 <strong>Analyse — {sprint['name']}</strong>",
            (
                f"• Vélocité : <strong>{sprint['velocity_completed']}/"
                f"{sprint['velocity_target']} pts</strong> "
                f"({sprint['velocity_percent']}%)"
            ),
            (
                f"• Story points : {sprint['completed_story_points']}/"
                f"{sprint['total_story_points']} terminés, "
                f"{sprint['remaining_story_points']} restants"
            ),
        ]

        if sprint["days_remaining"] is not None:
            lines.append(f"• Jours restants : <strong>{sprint['days_remaining']}</strong>")

        if sprint["velocity_percent"] < 70:
            lines.append(
                "⚠️ Vélocité en retrait — envisagez de réduire le scope "
                "ou de réallouer une ressource."
            )
        elif sprint["velocity_percent"] >= 95:
            lines.append(
                "✅ Sprint en avance — pensez à précharger des items du backlog."
            )

        return "<br>".join(lines)

    @staticmethod
    def _answer_risk_projects(ctx: dict) -> str:
        risks = ctx.get("at_risk_projects") or []

        if not risks:
            return "✅ Aucun projet à risque détecté."

        lines = ["⚠️ <strong>Projets à risque</strong>"]

        for project in risks:
            late_str = (
                f" · en retard de {project['days_late']}j"
                if project.get("days_late", 0) > 0
                else ""
            )

            lines.append(
                f"• <strong>{project['name']}</strong> — "
                f"{project['ai_risk_label']} "
                f"(score {project['risk_score']}), "
                f"{project['progress_percent']}% terminé{late_str}"
            )

        return "<br>".join(lines)

    @staticmethod
    def _answer_team_workload(ctx: dict) -> str:
        rows = ctx.get("team_workload") or []

        if not rows:
            return "ℹ️ Aucune charge à reporter actuellement."

        lines = ["👥 <strong>Charge équipe (tâches ouvertes)</strong>"]

        for row in rows:
            lines.append(
                f"• {row['user_label']} : "
                f"<strong>{row['open_tasks']} tâches</strong>, "
                f"{row['remaining_hours']:.0f}h restantes"
            )

        return "<br>".join(lines)

    @staticmethod
    def _answer_report(ctx: dict) -> str:
        ws = ctx.get("workspace") or {}
        sprint = ctx.get("active_sprint")
        risks = ctx.get("at_risk_projects") or []

        lines = [f"📋 <strong>Rapport hebdo — {ws.get('name', '—')}</strong>"]
        lines.append(f"• Projets actifs : {ws.get('active_projects_count', 0)}")

        if sprint:
            lines.append(
                f"• Sprint actif : {sprint['name']} "
                f"({sprint['velocity_percent']}% vélocité, "
                f"{sprint['days_remaining']}j restants)"
            )

        lines.append(f"• Projets à risque : {len(risks)}")

        if risks:
            top = risks[0]
            lines.append(
                f"↳ Le plus critique : {top['name']} "
                f"(risque {top['ai_risk_label']})"
            )

        return "<br>".join(lines)

    @classmethod
    def _call_ai(cls, session, ctx: dict, current_message: str, seed: str = ""):
        provider = get_ai_provider()

        if not provider or not provider.is_available():
            return seed, "heuristic", "", 0

        is_general = cls._is_general_question(current_message)
        needs_web = cls._needs_web_search(current_message)

        web_context = ""

        if needs_web:
            web_context = web_search(current_message)

        history = list(session.messages.order_by("-created_at")[:10])
        history.reverse()

        messages = [
            AIMessage(role="system", content=cls.SYSTEM_PROMPT),
        ]

        if is_general:
            messages.append(AIMessage(
                role="system",
                content=(
                    "La question semble générale. Réponds avec tes connaissances générales. "
                    "N'utilise le contexte DevFlow que s'il est pertinent ou demandé explicitement."
                ),
            ))
        else:
            messages.append(AIMessage(
                role="system",
                content=(
                    "Contexte DevFlow temps-réel :\n"
                    + json.dumps(ctx, ensure_ascii=False, default=str)[:4000]
                ),
            ))

        if web_context:
            messages.append(AIMessage(
                role="system",
                content=(
                    "Résultats web récents à utiliser si pertinents :\n"
                    + web_context[:4000]
                ),
            ))

        role_map = {
            dm.AIChatMessage.Role.USER: "user",
            dm.AIChatMessage.Role.ASSISTANT: "assistant",
            dm.AIChatMessage.Role.SYSTEM: "system",
        }

        for old_message in history[:-1]:
            messages.append(AIMessage(
                role=role_map.get(old_message.role, "user"),
                content=old_message.content[:2000],
            ))

        user_content = current_message

        if seed:
            user_content = (
                f"{current_message}\n\n"
                f"[Pré-analyse factuelle DevFlow disponible :\n{seed}\n]\n"
                "Utilise ces faits uniquement s'ils sont pertinents. "
                "Si la question est générale, réponds naturellement sans te limiter à DevFlow."
            )

        messages.append(AIMessage(role="user", content=user_content[:5000]))

        try:
            response = provider.generate(messages=messages, temperature=0.4)

            return (
                (response.text or "").strip(),
                provider.name,
                response.model or "",
                response.tokens_used or 0,
            )

        except Exception as exc:
            logger.warning("AI chat call failed: %s", exc)
            return seed, "heuristic", "", 0

# from __future__ import annotations
#
# import logging
# from dataclasses import dataclass, field
# from datetime import timedelta
# from decimal import Decimal
# from typing import Any
#
# from django.contrib.auth import get_user_model
# from django.db.models import Count, Q, Sum
# from django.utils import timezone
#
# from project import models as dm
# from project.services.ai.base import AIMessage
# from project.services.ai.factory import get_ai_provider
# from project.services.ai.openai_provider import OpenAIProvider
#
# logger = logging.getLogger(__name__)
# User = get_user_model()
#
#
# @dataclass
# class ChatTurnResult:
#     session: dm.AIChatSession
#     user_message: dm.AIChatMessage
#     assistant_message: dm.AIChatMessage
#     used_provider: str
#     context: dict = field(default_factory=dict)
#
#
# # =========================================================================
# # Builder de contexte : tire les vraies données DevFlow
# # =========================================================================
# class DevFlowContextBuilder:
#     @classmethod
#     def build_for_user(cls, user, workspace=None, project=None) -> dict:
#         ws = workspace or cls._infer_workspace(user)
#         ctx = {
#             "user": cls._user_summary(user),
#             "workspace": cls._workspace_summary(ws) if ws else None,
#             "active_sprint": cls._active_sprint(ws, project) if ws else None,
#             "at_risk_projects": cls._at_risk_projects(ws, limit=4) if ws else [],
#             "team_workload": cls._team_workload(ws, limit=6) if ws else [],
#             "user_open_tasks": cls._user_open_tasks(user, limit=5),
#             "now": timezone.now().isoformat(),
#         }
#         return ctx
#
#     # ------ helpers internes ------
#     @staticmethod
#     def _user_summary(user):
#         return {
#             "id": user.pk,
#             "name": user.get_full_name() or user.username,
#             "first_name": user.first_name or user.username,
#             "email": user.email,
#         }
#
#     @staticmethod
#     def _infer_workspace(user):
#         membership = (
#             dm.TeamMembership.objects.filter(user=user)
#             .select_related("workspace")
#             .order_by("-created_at")
#             .first()
#         )
#         if membership and membership.workspace_id:
#             return membership.workspace
#         # Sinon premier workspace owné
#         return dm.Workspace.objects.filter(owner=user).first()
#
#     @staticmethod
#     def _workspace_summary(ws):
#         active_projects = ws.projects.exclude(
#             status__in=[dm.Project.Status.DONE, dm.Project.Status.CANCELLED]
#         ).filter(is_archived=False)
#         return {
#             "id": ws.pk,
#             "name": ws.name,
#             "active_projects_count": active_projects.count(),
#             "members_count": ws.memberships.count(),
#         }
#
#     @staticmethod
#     def _active_sprint(ws, project=None):
#         qs = dm.Sprint.objects.filter(
#             workspace=ws,
#             status=dm.Sprint.Status.ACTIVE,
#             is_archived=False,
#         )
#         if project:
#             qs = qs.filter(project=project)
#         sprint = qs.order_by("-start_date").first()
#         if not sprint:
#             return None
#
#         today = timezone.localdate()
#         days_remaining = max((sprint.end_date - today).days, 0) if sprint.end_date else None
#         velocity_target = sprint.velocity_target or 0
#         velocity_completed = sprint.velocity_completed or 0
#         velocity_percent = 0
#         if velocity_target > 0:
#             velocity_percent = int((velocity_completed / velocity_target) * 100)
#
#         return {
#             "id": sprint.pk,
#             "name": sprint.name,
#             "number": sprint.number,
#             "project": sprint.project.name if sprint.project else "",
#             "days_remaining": days_remaining,
#             "velocity_target": velocity_target,
#             "velocity_completed": velocity_completed,
#             "velocity_percent": velocity_percent,
#             "total_story_points": sprint.total_story_points,
#             "completed_story_points": sprint.completed_story_points,
#             "remaining_story_points": sprint.remaining_story_points,
#         }
#
#     @staticmethod
#     def _at_risk_projects(ws, limit=4):
#         today = timezone.localdate()
#         qs = (
#             ws.projects.filter(is_archived=False)
#             .exclude(status__in=[dm.Project.Status.DONE, dm.Project.Status.CANCELLED])
#             .order_by("-risk_score")
#         )
#         out = []
#         for p in qs[:limit]:
#             days_late = 0
#             if p.target_date and p.target_date < today and p.status != dm.Project.Status.DONE:
#                 days_late = (today - p.target_date).days
#             out.append({
#                 "id": p.pk,
#                 "name": p.name,
#                 "status": p.get_status_display(),
#                 "priority": p.get_priority_display(),
#                 "risk_score": p.risk_score,
#                 "ai_risk_label": p.ai_risk_label or "—",
#                 "progress_percent": p.progress_percent,
#                 "target_date": p.target_date.isoformat() if p.target_date else None,
#                 "days_late": days_late,
#             })
#         return out
#
#     @staticmethod
#     def _team_workload(ws, limit=6):
#         # Charge = somme des heures restantes (estimate - spent) sur tâches non terminées
#         rows = (
#             dm.Task.objects.filter(
#                 workspace=ws, is_archived=False, assignee__isnull=False,
#             )
#             .exclude(status__in=[dm.Task.Status.DONE, dm.Task.Status.CANCELLED])
#             .values("assignee_id", "assignee__username", "assignee__first_name", "assignee__last_name")
#             .annotate(
#                 open_tasks=Count("id"),
#                 est_hours=Sum("estimate_hours"),
#                 spent_hours=Sum("spent_hours"),
#             )
#             .order_by("-open_tasks")[:limit]
#         )
#         out = []
#         for r in rows:
#             est = r["est_hours"] or Decimal("0")
#             spent = r["spent_hours"] or Decimal("0")
#             remaining = max(est - spent, Decimal("0"))
#             label = (
#                 (r["assignee__first_name"] or "") + " " + (r["assignee__last_name"] or "")
#             ).strip() or r["assignee__username"]
#             out.append({
#                 "user_id": r["assignee_id"],
#                 "user_label": label,
#                 "open_tasks": r["open_tasks"],
#                 "estimate_hours": float(est),
#                 "remaining_hours": float(remaining),
#             })
#         return out
#
#     @staticmethod
#     def _user_open_tasks(user, limit=5):
#         qs = (
#             dm.Task.objects.filter(assignee=user, is_archived=False)
#             .exclude(status__in=[dm.Task.Status.DONE, dm.Task.Status.CANCELLED])
#             .select_related("project")
#             .order_by("due_date", "-priority")[:limit]
#         )
#         return [
#             {
#                 "id": t.pk,
#                 "title": t.title,
#                 "project": t.project.name if t.project else "",
#                 "status": t.get_status_display(),
#                 "priority": t.get_priority_display(),
#                 "due_date": t.due_date.isoformat() if t.due_date else None,
#             }
#             for t in qs
#         ]
#
#
# # =========================================================================
# # Service principal
# # =========================================================================
# class AIChatService:
#
#     SYSTEM_PROMPT = (
#         "Tu es DevFlow AI, l'assistant intégré dans la plateforme DevFlow "
#         "(gestion de projet + finance TJM + IA). Tu réponds en français, de "
#         "manière concise (2-5 phrases sauf si on te demande un rapport), "
#         "avec des chiffres précis quand ils sont fournis dans le contexte. "
#         "Tu ne donnes JAMAIS de conseil hors-sujet. Tu ne hallucines pas de "
#         "données : si une info manque dans le contexte fourni, dis-le. "
#         "Tu peux utiliser des emojis sobres (📊 ⚠️ ✅) et du gras simple "
#         "(<strong>...</strong>) — pas de markdown lourd."
#     )
#
#     INTENT_KEYWORDS = {
#         "sprint_analysis": ["analyse du sprint", "sprint analysis", "vélocité"],
#         "risk_projects": ["projets à risque", "risk", "à risque"],
#         "report": ["générer un rapport", "rapport"],
#         "team_workload": ["charge équipe", "workload"],
#     }
#
#     # ---------------------------------------------------------------------
#     # Public API
#     # ---------------------------------------------------------------------
#     @classmethod
#     def get_or_create_session(cls, user, session_id=None, workspace=None, project=None):
#         if session_id:
#             session = (
#                 dm.AIChatSession.objects.filter(pk=session_id, user=user, is_active=True).first()
#             )
#             if session:
#                 return session
#         return dm.AIChatSession.objects.create(
#             user=user,
#             workspace=workspace or DevFlowContextBuilder._infer_workspace(user),
#             project=project,
#         )
#
#     @classmethod
#     def welcome_message(cls, user, workspace=None) -> dict:
#         """
#         Construit le message d'accueil dynamique avec les vraies données.
#         Pas d'appel IA ici — on veut une réponse instantanée à l'ouverture.
#         """
#         ctx = DevFlowContextBuilder.build_for_user(user, workspace=workspace)
#         sprint = ctx.get("active_sprint")
#         risks = ctx.get("at_risk_projects") or []
#         first_name = ctx["user"]["first_name"]
#
#         parts = [f"👋 Bonjour {first_name} ! Je suis DevFlow AI."]
#
#         if sprint:
#             parts.append(
#                 f"<br><br>Je vois que le <strong>{sprint['name']}</strong> "
#                 f"est en cours{' avec ' + str(sprint['days_remaining']) + ' jours restants' if sprint['days_remaining'] is not None else ''}. "
#                 f"La vélocité est à <strong>{sprint['velocity_percent']}%</strong>"
#                 + (" — légèrement en dessous de la cible." if sprint['velocity_percent'] < 80 else " — au-dessus de la cible.")
#             )
#         elif ctx.get("workspace"):
#             ws = ctx["workspace"]
#             parts.append(
#                 f"<br><br>Aucun sprint actif. {ws['active_projects_count']} projet(s) actifs sur le workspace <strong>{ws['name']}</strong>."
#             )
#
#         late_projects = [p for p in risks if p.get("days_late", 0) > 0]
#         if late_projects:
#             top = late_projects[0]
#             parts.append(
#                 f"<br><br>⚠️ <strong>{top['name']}</strong> est en retard ({top['days_late']} jour(s)). "
#                 "Voulez-vous que j'analyse les risques ?"
#             )
#         elif risks:
#             high = [p for p in risks if p.get("risk_score", 0) >= 60]
#             if high:
#                 parts.append(
#                     f"<br><br>⚠️ <strong>{high[0]['name']}</strong> présente un risque élevé (score {high[0]['risk_score']})."
#                 )
#
#         return {
#             "html": "".join(parts),
#             "context": ctx,
#             "suggestions": [
#                 "📊 Analyse du sprint",
#                 "⚠️ Projets à risque",
#                 "📋 Générer un rapport",
#                 "👥 Charge équipe",
#             ],
#         }
#
#     @classmethod
#     def process_user_message(
#         cls,
#         user,
#         message: str,
#         session_id=None,
#         workspace=None,
#         project=None,
#     ) -> ChatTurnResult:
#         message = (message or "").strip()
#         if not message:
#             raise ValueError("Message vide")
#
#         session = cls.get_or_create_session(
#             user=user, session_id=session_id, workspace=workspace, project=project,
#         )
#
#         # Persiste le message utilisateur
#         user_msg = dm.AIChatMessage.objects.create(
#             session=session,
#             role=dm.AIChatMessage.Role.USER,
#             content=message[:5000],
#             intent=cls._detect_intent(message),
#         )
#
#         # Construit le contexte DevFlow réel
#         ctx = DevFlowContextBuilder.build_for_user(
#             user=user, workspace=session.workspace, project=session.project,
#         )
#
#         # Si intent reconnu, on tente d'abord la réponse déterministe
#         # (toujours disponible, basée sur les données réelles)
#         deterministic = cls._answer_intent(user_msg.intent, ctx)
#
#         # Puis on enrichit avec l'IA si elle est disponible
#         ai_text, used_provider, model_name, tokens = cls._call_ai(
#             session=session, ctx=ctx, current_message=message,
#             seed=deterministic,
#         )
#
#         final_text = ai_text or deterministic or (
#             "Je n'ai pas réussi à interpréter votre demande. Pouvez-vous reformuler ?"
#         )
#
#         assistant_msg = dm.AIChatMessage.objects.create(
#             session=session,
#             role=dm.AIChatMessage.Role.ASSISTANT,
#             content=final_text[:8000],
#             intent=user_msg.intent,
#             used_provider=used_provider,
#             used_model=model_name,
#             tokens_used=tokens,
#             context_payload=ctx,
#         )
#
#         # Met à jour le titre de la session si vide
#         if not session.title:
#             session.title = (message[:120]) or "Conversation IA"
#             session.save(update_fields=["title", "updated_at"])
#
#         return ChatTurnResult(
#             session=session,
#             user_message=user_msg,
#             assistant_message=assistant_msg,
#             used_provider=used_provider,
#             context=ctx,
#         )
#
#     # ---------------------------------------------------------------------
#     # Intent detection
#     # ---------------------------------------------------------------------
#     @classmethod
#     def _detect_intent(cls, message: str) -> str:
#         m = (message or "").lower()
#         for intent, keywords in cls.INTENT_KEYWORDS.items():
#             for kw in keywords:
#                 if kw in m:
#                     return intent
#         return ""
#
#     # ---------------------------------------------------------------------
#     # Deterministic answers (toujours dispo, factuels)
#     # ---------------------------------------------------------------------
#     @classmethod
#     def _answer_intent(cls, intent: str, ctx: dict) -> str:
#         if intent == "sprint_analysis":
#             return cls._answer_sprint_analysis(ctx)
#         if intent == "risk_projects":
#             return cls._answer_risk_projects(ctx)
#         if intent == "team_workload":
#             return cls._answer_team_workload(ctx)
#         if intent == "report":
#             return cls._answer_report(ctx)
#         return ""
#
#     @staticmethod
#     def _answer_sprint_analysis(ctx: dict) -> str:
#         sprint = ctx.get("active_sprint")
#         if not sprint:
#             return "ℹ️ Aucun sprint actif détecté sur votre workspace."
#         lines = [
#             f"📊 <strong>Analyse — {sprint['name']}</strong>",
#             f"• Vélocité : <strong>{sprint['velocity_completed']}/{sprint['velocity_target']} pts</strong> ({sprint['velocity_percent']}%)",
#             f"• Story points : {sprint['completed_story_points']}/{sprint['total_story_points']} terminés, {sprint['remaining_story_points']} restants",
#         ]
#         if sprint['days_remaining'] is not None:
#             lines.append(f"• Jours restants : <strong>{sprint['days_remaining']}</strong>")
#         if sprint['velocity_percent'] < 70:
#             lines.append("⚠️ Vélocité en retrait — envisagez de réduire le scope ou ré-allouer une ressource.")
#         elif sprint['velocity_percent'] >= 95:
#             lines.append("✅ Sprint en avance — pensez à pré-charger des items du backlog.")
#         return "<br>".join(lines)
#
#     @staticmethod
#     def _answer_risk_projects(ctx: dict) -> str:
#         risks = ctx.get("at_risk_projects") or []
#         if not risks:
#             return "✅ Aucun projet à risque détecté."
#         lines = ["⚠️ <strong>Projets à risque</strong>"]
#         for p in risks:
#             late_str = f" · en retard de {p['days_late']}j" if p.get("days_late", 0) > 0 else ""
#             lines.append(
#                 f"• <strong>{p['name']}</strong> — {p['ai_risk_label']} (score {p['risk_score']}), "
#                 f"{p['progress_percent']}% terminé{late_str}"
#             )
#         return "<br>".join(lines)
#
#     @staticmethod
#     def _answer_team_workload(ctx: dict) -> str:
#         rows = ctx.get("team_workload") or []
#         if not rows:
#             return "ℹ️ Aucune charge à reporter actuellement."
#         lines = ["👥 <strong>Charge équipe (tâches ouvertes)</strong>"]
#         for r in rows:
#             lines.append(
#                 f"• {r['user_label']} : <strong>{r['open_tasks']} tâches</strong>, "
#                 f"{r['remaining_hours']:.0f}h restantes"
#             )
#         return "<br>".join(lines)
#
#     @staticmethod
#     def _answer_report(ctx: dict) -> str:
#         ws = ctx.get("workspace") or {}
#         sprint = ctx.get("active_sprint")
#         risks = ctx.get("at_risk_projects") or []
#         lines = [f"📋 <strong>Rapport hebdo — {ws.get('name', '—')}</strong>"]
#         lines.append(f"• Projets actifs : {ws.get('active_projects_count', 0)}")
#         if sprint:
#             lines.append(
#                 f"• Sprint actif : {sprint['name']} ({sprint['velocity_percent']}% vélocité, "
#                 f"{sprint['days_remaining']}j restants)"
#             )
#         lines.append(f"• Projets à risque : {len(risks)}")
#         if risks:
#             top = risks[0]
#             lines.append(f"  ↳ Le plus critique : {top['name']} (risque {top['ai_risk_label']})")
#         return "<br>".join(lines)
#
#     # ---------------------------------------------------------------------
#     # Provider IA
#     # ---------------------------------------------------------------------
#     @classmethod
#     def _call_ai(cls, session, ctx, current_message, seed=""):
#         provider = get_ai_provider()
#         if not provider.is_available():
#             return seed, "heuristic", "", 0
#
#         # Historique des derniers messages (max 10)
#         history = list(
#             session.messages.order_by("-created_at")[:10]
#         )
#         history.reverse()
#
#         messages = [AIMessage(role="system", content=cls.SYSTEM_PROMPT)]
#
#         # Ajoute le contexte DevFlow comme bloc system
#         import json
#         messages.append(AIMessage(
#             role="system",
#             content="Contexte DevFlow temps-réel:\n" + json.dumps(ctx, ensure_ascii=False, default=str)[:4000],
#         ))
#
#         for h in history[:-1]:  # tous sauf le message courant déjà persisté
#             role_map = {
#                 dm.AIChatMessage.Role.USER: "user",
#                 dm.AIChatMessage.Role.ASSISTANT: "assistant",
#                 dm.AIChatMessage.Role.SYSTEM: "system",
#             }
#             messages.append(AIMessage(role=role_map.get(h.role, "user"), content=h.content))
#
#         # Si on a une réponse heuristique, on la passe en seed pour orienter l'IA
#         user_content = current_message
#         if seed:
#             user_content = (
#                 f"{current_message}\n\n[Pré-analyse heuristique disponible :\n{seed}\n]"
#                 "\nUtilise ces faits comme base et reformule de manière naturelle, en français, en restant concis."
#             )
#         messages.append(AIMessage(role="user", content=user_content))
#
#         try:
#             response = provider.generate(messages=messages, temperature=0.4)
#             return response.text.strip(), provider.name, response.model or "", response.tokens_used or 0
#         except Exception as exc:
#             logger.warning("AI chat call failed: %s", exc)
#             return seed, "heuristic", "", 0
